#!/usr/bin/env python3
"""
Gap Analysis Approfondie: RAW vs FINAL
Analyse méthodique sheet par sheet pour identifier tous les éléments manquants
"""

import openpyxl
from pathlib import Path
import yaml
from rich.console import Console
from rich.table import Table
from rich import box
import logging

console = Console()
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class DeepGapAnalyzer:
    def __init__(self, raw_path: Path, final_path: Path, assumptions_path: Path):
        self.raw_path = raw_path
        self.final_path = final_path

        logger.info(f"📂 Chargement RAW: {raw_path.name}")
        self.raw_wb = openpyxl.load_workbook(raw_path, data_only=False)
        logger.info(f"✓ RAW: {len(self.raw_wb.sheetnames)} sheets")

        logger.info(f"📂 Chargement FINAL: {final_path.name}")
        self.final_wb = openpyxl.load_workbook(final_path, data_only=False)
        logger.info(f"✓ FINAL: {len(self.final_wb.sheetnames)} sheets")

        logger.info(f"📂 Chargement assumptions: {assumptions_path.name}")
        with open(assumptions_path, 'r', encoding='utf-8') as f:
            self.assumptions = yaml.safe_load(f)
        logger.info(f"✓ Assumptions chargées\n")

        self.gaps = []
        self.missing_elements = []
        self.recommendations = []

    def analyze_sheet_content(self, sheet_name: str):
        """Analyser en profondeur le contenu d'un sheet"""
        console.print(f"\n[bold cyan]═══ {sheet_name} ═══[/bold cyan]")

        if sheet_name not in self.raw_wb.sheetnames:
            console.print(f"[yellow]⚠️ Sheet '{sheet_name}' absent du RAW[/yellow]")
            return

        raw_ws = self.raw_wb[sheet_name]

        if sheet_name not in self.final_wb.sheetnames:
            console.print(f"[red]❌ Sheet '{sheet_name}' MANQUANT dans FINAL[/red]")
            self.gaps.append({
                'sheet': sheet_name,
                'type': 'sheet_missing',
                'severity': 'HIGH',
                'description': f"Sheet entier manquant"
            })
            return

        final_ws = self.final_wb[sheet_name]

        # Comparer dimensions
        raw_rows = raw_ws.max_row
        raw_cols = raw_ws.max_column
        final_rows = final_ws.max_row
        final_cols = final_ws.max_column

        console.print(f"Dimensions: RAW={raw_rows}×{raw_cols}, FINAL={final_rows}×{final_cols}")

        # Extraire sections importantes (première colonne)
        raw_labels = self._extract_labels(raw_ws, max_row=min(100, raw_rows))
        final_labels = self._extract_labels(final_ws, max_row=min(100, final_rows))

        # Identifier labels manquants
        missing_labels = []
        for label in raw_labels:
            if label and label not in final_labels:
                # Vérifier que ce n'est pas juste une variation mineure
                if not any(self._similar(label, fl) for fl in final_labels):
                    missing_labels.append(label)

        if missing_labels:
            console.print(f"\n[yellow]⚠️ {len(missing_labels)} label(s) manquant(s):[/yellow]")
            for label in missing_labels[:10]:  # Limiter à 10 pour la lisibilité
                console.print(f"  • {label}")
                self.missing_elements.append({
                    'sheet': sheet_name,
                    'type': 'label_missing',
                    'label': label,
                    'severity': 'MEDIUM'
                })

            if len(missing_labels) > 10:
                console.print(f"  ... et {len(missing_labels) - 10} autres")
        else:
            console.print("[green]✓ Tous les labels principaux présents[/green]")

        # Analyser formules
        raw_formulas = self._count_formulas(raw_ws)
        final_formulas = self._count_formulas(final_ws)

        if raw_formulas != final_formulas:
            console.print(f"[yellow]⚠️ Formules: RAW={raw_formulas}, FINAL={final_formulas}[/yellow]")
            if final_formulas < raw_formulas:
                self.gaps.append({
                    'sheet': sheet_name,
                    'type': 'formulas_lost',
                    'severity': 'MEDIUM',
                    'description': f"{raw_formulas - final_formulas} formules perdues"
                })
        else:
            console.print(f"[green]✓ Formules préservées ({raw_formulas})[/green]")

    def _extract_labels(self, ws, max_row=100):
        """Extraire les labels de la colonne A"""
        labels = []
        for row in range(1, max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value and isinstance(cell_value, str):
                # Nettoyer
                label = cell_value.strip()
                if len(label) > 0 and not label.startswith('='):
                    labels.append(label)
        return labels

    def _similar(self, s1: str, s2: str) -> bool:
        """Vérifier si deux strings sont similaires (pour gérer variations mineures)"""
        s1_clean = s1.lower().replace(' ', '').replace('_', '')
        s2_clean = s2.lower().replace(' ', '').replace('_', '')
        return s1_clean in s2_clean or s2_clean in s1_clean

    def _count_formulas(self, ws):
        """Compter le nombre de formules dans un sheet"""
        count = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    count += 1
        return count

    def analyze_parametres_completeness(self):
        """Vérifier que TOUTES les assumptions sont visibles dans Paramètres"""
        console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   ANALYSE COMPLÉTUDE PARAMÈTRES vs ASSUMPTIONS.YAML[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        if 'Paramètres' not in self.final_wb.sheetnames:
            console.print("[red]❌ Sheet Paramètres absent![/red]")
            return

        ws = self.final_wb['Paramètres']

        # Extraire tout le contenu de Paramètres
        params_content = []
        for row in range(1, min(100, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row, col)
                if cell.value:
                    params_content.append(str(cell.value))

        params_text = ' '.join(params_content).lower()

        # Vérifier couverture des sections principales
        table = Table(title="Couverture Assumptions → Paramètres", box=box.ROUNDED)
        table.add_column("Section YAML", style="cyan")
        table.add_column("Éléments Clés", style="white")
        table.add_column("Visible dans Paramètres", style="white")
        table.add_column("Status", style="white")

        # 1. Pricing
        pricing = self.assumptions.get('pricing', {})
        hackathon_price = pricing.get('hackathon', {}).get('base_price', 0)
        factory_price = pricing.get('factory', {}).get('base_price', 0)

        hackathon_visible = str(hackathon_price) in params_text or '18000' in params_text
        factory_visible = str(factory_price) in params_text or '75000' in params_text

        table.add_row(
            "pricing.hackathon",
            f"Prix base: {hackathon_price}€",
            "Oui" if hackathon_visible else "Non",
            "✅" if hackathon_visible else "❌"
        )
        table.add_row(
            "pricing.factory",
            f"Prix base: {factory_price}€",
            "Oui" if factory_visible else "Non",
            "✅" if factory_visible else "❌"
        )

        # 2. Hub pricing
        hub = pricing.get('enterprise_hub', {})
        for tier in ['starter', 'business', 'enterprise']:
            price = hub.get(tier, {}).get('monthly_price', 0)
            visible = str(price) in params_text
            table.add_row(
                f"pricing.hub.{tier}",
                f"Prix mensuel: {price}€",
                "Oui" if visible else "Non",
                "✅" if visible else "❌"
            )
            if not visible:
                self.missing_elements.append({
                    'sheet': 'Paramètres',
                    'type': 'assumption_missing',
                    'section': f'pricing.hub.{tier}',
                    'value': price,
                    'severity': 'HIGH'
                })

        # 3. Conversion rates
        sales_assumptions = self.assumptions.get('sales_assumptions', {})
        factory_conv = sales_assumptions.get('factory', {}).get('conversion_rate', 0)

        conv_visible = 'conversion' in params_text and ('0.3' in params_text or '30%' in params_text or str(factory_conv) in params_text)
        table.add_row(
            "sales.factory.conversion",
            f"Taux: {factory_conv*100}%",
            "Oui" if conv_visible else "Non",
            "✅" if conv_visible else "❌"
        )

        if not conv_visible:
            self.missing_elements.append({
                'sheet': 'Paramètres',
                'type': 'assumption_missing',
                'section': 'sales.factory.conversion',
                'value': factory_conv,
                'severity': 'HIGH'
            })

        # 4. Churn rates
        hub_churn_monthly = sales_assumptions.get('enterprise_hub', {}).get('churn_monthly', 0)
        hub_churn_annual = hub_churn_monthly * 12

        churn_visible = 'churn' in params_text
        table.add_row(
            "sales.hub.churn",
            f"Mensuel: {hub_churn_monthly*100}%, Annuel: {hub_churn_annual*100}%",
            "Oui" if churn_visible else "Non",
            "✅" if churn_visible else "❌"
        )

        if not churn_visible:
            self.missing_elements.append({
                'sheet': 'Paramètres',
                'type': 'assumption_missing',
                'section': 'sales.hub.churn',
                'value': hub_churn_monthly,
                'severity': 'HIGH'
            })

        # 5. Financial KPIs
        financial_kpis = self.assumptions.get('financial_kpis', {})
        arr_m14_target = financial_kpis.get('target_arr_dec_2026', 0)

        arr_visible = 'arr' in params_text and ('800' in params_text or str(arr_m14_target) in params_text)
        table.add_row(
            "financial_kpis.target_arr_m14",
            f"ARR M14: {arr_m14_target}€",
            "Oui" if arr_visible else "Non",
            "✅" if arr_visible else "❌"
        )

        # 6. Team assumptions
        costs = self.assumptions.get('costs', {})
        social_charges = costs.get('social_charges_rate', 0)

        charges_visible = 'charge' in params_text and ('45%' in params_text or '0.45' in params_text or str(social_charges) in params_text)
        table.add_row(
            "costs.social_charges_rate",
            f"Taux: {social_charges*100}%",
            "Oui" if charges_visible else "Non",
            "✅" if charges_visible else "❌"
        )

        if not charges_visible:
            self.missing_elements.append({
                'sheet': 'Paramètres',
                'type': 'assumption_missing',
                'section': 'costs.social_charges_rate',
                'value': social_charges,
                'severity': 'MEDIUM'
            })

        # 7. Hypothèses volumes
        hackathons_per_month = sales_assumptions.get('hackathon', {}).get('volumes_monthly', {})
        if hackathons_per_month and isinstance(hackathons_per_month, dict):
            # Extraire les 12 premiers mois
            volumes = [hackathons_per_month.get(f'm{i}', 0) for i in range(1, 13)]
            avg_hackathons = sum(volumes) / len(volumes) if volumes else 0
            volumes_visible = 'hackathon' in params_text or 'volume' in params_text
            table.add_row(
                "sales.hackathon.volumes",
                f"Moy mensuelle: {avg_hackathons:.1f}",
                "Oui" if volumes_visible else "Non",
                "⚠️" if volumes_visible else "❌"
            )

        console.print(table)

        # Recommandations
        missing_count = len([e for e in self.missing_elements if e.get('sheet') == 'Paramètres'])
        if missing_count > 0:
            console.print(f"\n[red]❌ {missing_count} assumption(s) importante(s) manquante(s) dans Paramètres[/red]")
        else:
            console.print("\n[green]✅ Toutes les assumptions clés sont visibles dans Paramètres[/green]")

    def analyze_all_sheets(self):
        """Analyser tous les sheets du RAW"""
        console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   ANALYSE SHEET PAR SHEET[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")

        for sheet_name in self.raw_wb.sheetnames:
            self.analyze_sheet_content(sheet_name)

    def generate_recommendations(self):
        """Générer recommandations d'amélioration"""
        console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   RECOMMANDATIONS D'AMÉLIORATION[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        # Grouper par priorité
        high_priority = []
        medium_priority = []
        low_priority = []

        # Analyser les gaps identifiés
        for gap in self.gaps:
            if gap['severity'] == 'HIGH':
                high_priority.append(gap)
            elif gap['severity'] == 'MEDIUM':
                medium_priority.append(gap)
            else:
                low_priority.append(gap)

        for elem in self.missing_elements:
            if elem['severity'] == 'HIGH':
                high_priority.append(elem)
            elif elem['severity'] == 'MEDIUM':
                medium_priority.append(elem)
            else:
                low_priority.append(elem)

        # Afficher recommandations
        if high_priority:
            console.print(f"[bold red]🔴 HAUTE PRIORITÉ ({len(high_priority)} items)[/bold red]\n")
            for i, item in enumerate(high_priority[:10], 1):
                if item['type'] == 'assumption_missing':
                    console.print(f"{i}. [red]Ajouter dans Paramètres:[/red] {item['section']} = {item['value']}")
                elif item['type'] == 'label_missing':
                    console.print(f"{i}. [red]Label manquant dans {item['sheet']}:[/red] {item['label']}")
                else:
                    console.print(f"{i}. [red]{item['sheet']}:[/red] {item['description']}")

        if medium_priority:
            console.print(f"\n[bold yellow]🟡 MOYENNE PRIORITÉ ({len(medium_priority)} items)[/bold yellow]\n")
            for i, item in enumerate(medium_priority[:10], 1):
                if item['type'] == 'assumption_missing':
                    console.print(f"{i}. [yellow]Ajouter dans Paramètres:[/yellow] {item['section']} = {item['value']}")
                elif item['type'] == 'label_missing':
                    console.print(f"{i}. [yellow]Label manquant dans {item['sheet']}:[/yellow] {item['label']}")
                else:
                    console.print(f"{i}. [yellow]{item['sheet']}:[/yellow] {item.get('description', item.get('label', ''))}")

        if low_priority:
            console.print(f"\n[bold green]🟢 BASSE PRIORITÉ ({len(low_priority)} items)[/bold green]")
            console.print(f"(Détails disponibles si nécessaire)")

        # Recommandations spécifiques
        console.print("\n[bold cyan]💡 AMÉLIORATIONS RECOMMANDÉES:[/bold cyan]\n")

        recommendations = [
            {
                'priority': 'HIGH',
                'title': 'Enrichir Paramètres avec toutes les assumptions critiques',
                'description': 'Ajouter sections visibles pour: prix Hub par tier, taux conversion, churn rates, volumes hackathons/mois',
                'impact': 'Transparence totale pour investisseurs et équipe'
            },
            {
                'priority': 'HIGH',
                'title': 'Vérifier formules manquantes',
                'description': 'Certains sheets ont perdu des formules - vérifier si intentionnel ou bug',
                'impact': 'Intégrité des calculs Excel'
            },
            {
                'priority': 'MEDIUM',
                'title': 'Ajouter labels descriptifs manquants',
                'description': f'{len([e for e in self.missing_elements if e["type"] == "label_missing"])} labels du RAW absents - évaluer pertinence',
                'impact': 'Complétude et clarté'
            },
        ]

        for i, rec in enumerate(recommendations, 1):
            priority_color = {'HIGH': 'red', 'MEDIUM': 'yellow', 'LOW': 'green'}[rec['priority']]
            console.print(f"{i}. [{priority_color}][{rec['priority']}][/{priority_color}] {rec['title']}")
            console.print(f"   → {rec['description']}")
            console.print(f"   💥 Impact: {rec['impact']}\n")

    def generate_summary(self):
        """Générer résumé final"""
        console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   RÉSUMÉ GAP ANALYSIS[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        high_count = len([x for x in (self.gaps + self.missing_elements) if x['severity'] == 'HIGH'])
        medium_count = len([x for x in (self.gaps + self.missing_elements) if x['severity'] == 'MEDIUM'])
        low_count = len([x for x in (self.gaps + self.missing_elements) if x['severity'] == 'LOW'])

        table = Table(box=box.ROUNDED)
        table.add_column("Métrique", style="cyan")
        table.add_column("Valeur", style="white")

        table.add_row("Sheets RAW", str(len(self.raw_wb.sheetnames)))
        table.add_row("Sheets FINAL", str(len(self.final_wb.sheetnames)))
        table.add_row("Sheets nouveaux", str(len(self.final_wb.sheetnames) - len(self.raw_wb.sheetnames)))
        table.add_row("", "")
        table.add_row("Gaps HAUTE priorité", f"🔴 {high_count}")
        table.add_row("Gaps MOYENNE priorité", f"🟡 {medium_count}")
        table.add_row("Gaps BASSE priorité", f"🟢 {low_count}")
        table.add_row("", "")
        table.add_row("Total problèmes identifiés", str(high_count + medium_count + low_count))

        console.print(table)

        if high_count == 0:
            console.print("\n[bold green]✅ Aucun problème HAUTE priorité![/bold green]")
        else:
            console.print(f"\n[bold red]⚠️ {high_count} problème(s) HAUTE priorité à corriger[/bold red]")

        console.print("\n[cyan]→ Voir recommandations ci-dessus pour plan d'action[/cyan]")


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   GAP ANALYSIS APPROFONDIE: RAW vs FINAL[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    final_file = base_path / "data" / "outputs" / "BP_50M_FINAL_Nov2025-Dec2029.xlsx"
    assumptions_file = base_path / "data" / "structured" / "assumptions.yaml"

    analyzer = DeepGapAnalyzer(raw_file, final_file, assumptions_file)

    # 1. Analyser tous les sheets
    analyzer.analyze_all_sheets()

    # 2. Analyser complétude Paramètres
    analyzer.analyze_parametres_completeness()

    # 3. Générer recommandations
    analyzer.generate_recommendations()

    # 4. Résumé
    analyzer.generate_summary()

    console.print("\n[bold green]✅ GAP ANALYSIS TERMINÉE[/bold green]\n")


if __name__ == '__main__':
    main()
