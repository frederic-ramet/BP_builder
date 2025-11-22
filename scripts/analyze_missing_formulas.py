#!/usr/bin/env python3
"""
Analyse détaillée des formules manquantes entre RAW et TEMPLATE
Identifier exactement quelles cellules ont perdu leurs formules
"""

import openpyxl
from pathlib import Path
from rich.console import Console
from rich.table import Table
from rich import box
from collections import defaultdict

console = Console()

base_path = Path(__file__).parent.parent
raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
template_file = base_path / "data" / "outputs" / "BP_50M_TEMPLATE.xlsx"

console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
console.print("[bold cyan]   ANALYSE DÉTAILLÉE: Formules Manquantes (6.6%)[/bold cyan]")
console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

wb_raw = openpyxl.load_workbook(raw_file, data_only=False)
wb_template = openpyxl.load_workbook(template_file, data_only=False)

def get_formula_cells(ws):
    """Retourner dict {cell_ref: formula} pour toutes les formules"""
    formulas = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas[cell.coordinate] = cell.value
    return formulas

def analyze_formula_types(formulas):
    """Classifier les formules par type"""
    types = defaultdict(list)
    for cell, formula in formulas.items():
        formula_upper = formula.upper()
        if 'SUM(' in formula_upper:
            types['SUM'].append(cell)
        elif 'VLOOKUP(' in formula_upper or 'XLOOKUP(' in formula_upper:
            types['LOOKUP'].append(cell)
        elif 'IF(' in formula_upper:
            types['IF'].append(cell)
        elif 'SUMIF' in formula_upper or 'SUMIFS' in formula_upper:
            types['SUMIF/S'].append(cell)
        elif 'INDEX(' in formula_upper or 'MATCH(' in formula_upper:
            types['INDEX/MATCH'].append(cell)
        elif formula_upper.startswith('=') and '+' in formula_upper:
            types['ADDITION'].append(cell)
        elif formula_upper.startswith('=') and '*' in formula_upper:
            types['MULTIPLICATION'].append(cell)
        else:
            types['OTHER'].append(cell)
    return types

# Analyser le sheet le plus impacté: "Charges de personnel et FG"
console.print("[bold yellow]═══ FOCUS: Charges de personnel et FG ═══[/bold yellow]\n")

sheet_name = "Charges de personnel et FG"
ws_raw = wb_raw[sheet_name]
ws_template = wb_template[sheet_name]

formulas_raw = get_formula_cells(ws_raw)
formulas_template = get_formula_cells(ws_template)

console.print(f"[cyan]RAW:[/cyan] {len(formulas_raw)} formules")
console.print(f"[cyan]TEMPLATE:[/cyan] {len(formulas_template)} formules")
console.print(f"[red]PERDUES:[/red] {len(formulas_raw) - len(formulas_template)} formules\n")

# Identifier les cellules qui ont perdu leurs formules
lost_cells = set(formulas_raw.keys()) - set(formulas_template.keys())
gained_cells = set(formulas_template.keys()) - set(formulas_raw.keys())

console.print(f"[red]Formules perdues:[/red] {len(lost_cells)}")
console.print(f"[green]Formules ajoutées:[/green] {len(gained_cells)}\n")

# Analyser les types de formules perdues
if lost_cells:
    lost_formulas = {cell: formulas_raw[cell] for cell in lost_cells}
    lost_types = analyze_formula_types(lost_formulas)

    console.print("[bold yellow]═══ TYPES DE FORMULES PERDUES ═══[/bold yellow]\n")

    table = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
    table.add_column("Type Formule", style="cyan", width=20)
    table.add_column("Nombre", justify="right", width=10)
    table.add_column("% du Total", justify="right", width=12)
    table.add_column("Exemples Cellules", width=30)

    for ftype, cells in sorted(lost_types.items(), key=lambda x: len(x[1]), reverse=True):
        pct = (len(cells) / len(lost_cells)) * 100
        examples = ", ".join(list(cells)[:5])
        if len(cells) > 5:
            examples += "..."
        table.add_row(ftype, str(len(cells)), f"{pct:.1f}%", examples)

    console.print(table)

# Analyser les zones de perte
console.print("\n[bold yellow]═══ ZONES GÉOGRAPHIQUES DES PERTES ═══[/bold yellow]\n")

def get_zone(cell_ref):
    """Déterminer la zone Excel (colonne)"""
    import re
    match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if match:
        col = match.group(1)
        row = int(match.group(2))
        return col, row
    return None, None

zones = defaultdict(list)
for cell in lost_cells:
    col, row = get_zone(cell)
    if col:
        zones[col].append(row)

# Top 10 colonnes avec le plus de pertes
top_zones = sorted(zones.items(), key=lambda x: len(x[1]), reverse=True)[:10]

table_zones = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
table_zones.add_column("Colonne", style="cyan", width=10)
table_zones.add_column("Formules Perdues", justify="right", width=18)
table_zones.add_column("Lignes", width=40)

for col, rows in top_zones:
    rows_sorted = sorted(rows)
    if len(rows_sorted) > 5:
        rows_str = f"{rows_sorted[0]}-{rows_sorted[-1]} ({len(rows_sorted)} lignes)"
    else:
        rows_str = ", ".join(map(str, rows_sorted))
    table_zones.add_row(col, str(len(rows)), rows_str)

console.print(table_zones)

# Échantillon de formules perdues
console.print("\n[bold yellow]═══ ÉCHANTILLON FORMULES PERDUES (10 premiers) ═══[/bold yellow]\n")

table_samples = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED, expand=True)
table_samples.add_column("Cellule", style="cyan", width=10)
table_samples.add_column("Formule RAW", style="yellow", width=60)
table_samples.add_column("Valeur TEMPLATE", style="green", width=20)

for i, cell in enumerate(sorted(lost_cells)[:10]):
    formula_raw = formulas_raw[cell]

    # Récupérer la valeur dans TEMPLATE
    try:
        template_val = ws_template[cell].value
        if template_val is None:
            template_val = "(vide)"
        elif isinstance(template_val, (int, float)):
            template_val = f"{template_val:,.0f}"
        else:
            template_val = str(template_val)[:20]
    except:
        template_val = "N/A"

    table_samples.add_row(cell, formula_raw[:60], template_val)

console.print(table_samples)

# Analyser Fundings aussi
console.print("\n\n[bold yellow]═══ FOCUS: Fundings (2 formules perdues) ═══[/bold yellow]\n")

sheet_name = "Fundings"
ws_raw = wb_raw[sheet_name]
ws_template = wb_template[sheet_name]

formulas_raw = get_formula_cells(ws_raw)
formulas_template = get_formula_cells(ws_template)

lost_cells_fundings = set(formulas_raw.keys()) - set(formulas_template.keys())

if lost_cells_fundings:
    table_fundings = Table(show_header=True, header_style="bold magenta", box=box.ROUNDED)
    table_fundings.add_column("Cellule", style="cyan", width=10)
    table_fundings.add_column("Formule RAW", style="yellow", width=70)

    for cell in sorted(lost_cells_fundings):
        formula = formulas_raw[cell]
        table_fundings.add_row(cell, formula)

    console.print(table_fundings)

# CONCLUSION
console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
console.print("[bold cyan]   CONCLUSION: Nature des 6.6% Manquants[/bold cyan]")
console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

console.print(f"[bold]Total formules perdues:[/bold] {len(formulas_raw) - len(formulas_template)} sur {len(formulas_raw)}")
console.print(f"[bold]Soit:[/bold] {((len(formulas_raw) - len(formulas_template)) / len(formulas_raw) * 100):.1f}%\n")

console.print("[yellow]Hypothèses:[/yellow]")
console.print("  1. Simplification YAML: Formules remplacées par données pilotées")
console.print("  2. Consolidation: Formules redondantes éliminées")
console.print("  3. Possible bug: À vérifier si critiques\n")

console.print("[green]Action recommandée:[/green]")
console.print("  → Vérifier si valeurs numériques correctes dans TEMPLATE")
console.print("  → Comparer calculs manuels RAW vs TEMPLATE")
console.print("  → Restaurer formules si impact métier")
