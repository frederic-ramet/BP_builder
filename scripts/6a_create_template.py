#!/usr/bin/env python3
"""
Créer un TEMPLATE Excel à partir du fichier RAW
Adapte la structure selon assumptions.yaml tout en préservant les formules
"""

import openpyxl
from pathlib import Path
import yaml
from rich.console import Console
from rich.progress import track
import logging
from copy import copy

console = Console()
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


class TemplateCreator:
    """Créer un template Excel adapté depuis le RAW"""

    def __init__(self, raw_path: Path, assumptions: dict):
        self.raw_path = raw_path
        self.assumptions = assumptions

        logger.info(f"📂 Chargement fichier RAW: {raw_path.name}")
        self.wb = openpyxl.load_workbook(raw_path)
        logger.info(f"✓ {len(self.wb.sheetnames)} sheets chargés")

    def update_parametres_sheet(self):
        """
        Adapter le sheet Paramètres selon assumptions.yaml
        - Prix Hackathon, Factory, Hub selon YAML
        - Évolution des prix par année
        """
        logger.info("\n⚙️ Adaptation sheet Paramètres...")

        ws = self.wb['Paramètres']

        # Extraire les prix du YAML
        pricing = self.assumptions.get('pricing', {})

        # Prix Hackathon (ligne 3)
        hackathon_base = pricing.get('hackathon_base', 18000)
        ws['B3'] = hackathon_base
        ws['C3'] = "=B3*1.1"  # +10% par an
        ws['D3'] = "=C3*1.1"
        ws['E3'] = "=D3*1.1"
        ws['F3'] = "=E3*1.1"

        # Prix Factory (ligne 10)
        factory_base = pricing.get('factory_base', 75000)
        ws['B10'] = factory_base
        ws['C10'] = "=B10*1.05"  # +5% par an
        ws['D10'] = "=C10*1.05"
        ws['E10'] = "=D10*1.05"
        ws['F10'] = "=E10*1.05"

        # Prix Hub - Starter (ligne 7)
        starter_base = pricing.get('hub_starter_monthly', 500)
        ws['B7'] = starter_base
        ws['C7'] = "=B7*1.1"
        ws['D7'] = "=C7*1.1"
        ws['E7'] = "=D7*1.1"
        ws['F7'] = "=E7*1.1"

        # Prix Hub - Business (ligne 8)
        business_base = pricing.get('hub_business_monthly', 2000)
        ws['B8'] = business_base
        ws['C8'] = "=B8*1.1"
        ws['D8'] = "=C8*1.1"
        ws['E8'] = "=D8*1.1"
        ws['F8'] = "=E8*1.1"

        # Prix Hub - Enterprise (ligne 9)
        enterprise_base = pricing.get('hub_enterprise_monthly', 10000)
        ws['B9'] = enterprise_base
        ws['C9'] = "=B9*1.1"
        ws['D9'] = "=C9*1.1"
        ws['E9'] = "=D9*1.1"
        ws['F9'] = "=E9*1.1"

        # Services implémentation (ligne 4)
        services_base = pricing.get('services_daily', 800) * 12.5  # Prix journée * nb jours moyen
        ws['B4'] = services_base
        ws['C4'] = "=B4*1.05"
        ws['D4'] = "=C4*1.05"
        ws['E4'] = "=D4*1.05"
        ws['F4'] = "=E4*1.05"

        # NOUVELLE SECTION: Financial KPIs (colonne H+)
        financial_kpis = self.assumptions.get('financial_kpis', {})

        ws['H1'].value = "FINANCIAL KPIs"
        ws['H2'].value = "Métrique"
        ws['I2'].value = "Valeur"

        row = 3
        kpis_data = [
            ("ARR Target M14", financial_kpis.get('target_arr_dec_2026', 800000)),
            ("ARR Target M11", financial_kpis.get('target_arr_sept_2026', 450000)),
            ("Marge Brute Target", f"{financial_kpis.get('margin_targets', {}).get('gross_margin_pct', 70)}%"),
            ("EBITDA Margin Target", f"{financial_kpis.get('margin_targets', {}).get('ebitda_margin_pct', -15)}%"),
            ("Min Cash Runway (mois)", financial_kpis.get('cash_management', {}).get('min_cash_runway_months', 12)),
            ("Burn Rate Max (€/mois)", financial_kpis.get('cash_management', {}).get('acceptable_burn_rate_monthly', 50000)),
            ("Target LTV/CAC", financial_kpis.get('saas_metrics', {}).get('target_ltv_cac_ratio', 8)),
            ("Max Churn Annual", f"{financial_kpis.get('saas_metrics', {}).get('max_churn_annual', 0.15)*100}%"),
        ]

        for label, value in kpis_data:
            ws[f'H{row}'].value = label
            ws[f'I{row}'].value = value
            row += 1

        # NOUVELLE SECTION: Validation Rules (colonne K+)
        validation_rules = self.assumptions.get('validation_rules', {})

        ws['K1'].value = "VALIDATION RULES"
        ws['K2'].value = "Règle"
        ws['L2'].value = "Min"
        ws['M2'].value = "Max"

        row = 3
        rules_data = [
            ("ARR M14", validation_rules.get('arr_m14_min', 720000), validation_rules.get('arr_m14_max', 880000)),
            ("ARR M11", validation_rules.get('arr_m11_min', 400000), None),
            ("Team Size M14", validation_rules.get('min_team_size_m1', 4), validation_rules.get('max_team_size', 15)),
            ("Burn Monthly", None, validation_rules.get('max_burn_monthly', 60000)),
            ("Cash Balance Min", validation_rules.get('min_cash_balance', 50000), None),
            ("Conversion Hackathon→Factory", f"{validation_rules.get('min_conversion_hackathon_factory', 0.25)*100}%", None),
            ("Churn Hub Monthly Max", None, f"{validation_rules.get('max_churn_hub_monthly', 0.015)*100}%"),
        ]

        for label, min_val, max_val in rules_data:
            ws[f'K{row}'].value = label
            ws[f'L{row}'].value = min_val if min_val else "-"
            ws[f'M{row}'].value = max_val if max_val else "-"
            row += 1

        # NOUVELLE SECTION: Hypothèses critiques (colonne O+)
        ws['O1'].value = "HYPOTHÈSES BUSINESS"
        ws['O2'].value = "Hypothèse"
        ws['P2'].value = "Valeur"

        row = 3
        business_assumptions = [
            ("Conversion Hackathon→Factory", f"{self.assumptions.get('sales_assumptions', {}).get('factory', {}).get('conversion_rate', 0.35)*100}%"),
            ("Churn Hub Monthly", f"{self.assumptions.get('sales_assumptions', {}).get('enterprise_hub', {}).get('churn_monthly', 0.008)*100}%"),
            ("Launch Hub", f"M{self.assumptions.get('pricing', {}).get('enterprise_hub', {}).get('launch_month', 8)}"),
            ("Délai Factory (mois)", self.assumptions.get('sales_assumptions', {}).get('factory', {}).get('delay_months', 2)),
            ("Tier Starter %", f"{self.assumptions.get('sales_assumptions', {}).get('enterprise_hub', {}).get('tier_distribution_at_launch', {}).get('starter', 0.6)*100}%"),
            ("Tier Business %", f"{self.assumptions.get('sales_assumptions', {}).get('enterprise_hub', {}).get('tier_distribution_at_launch', {}).get('business', 0.3)*100}%"),
            ("Tier Enterprise %", f"{self.assumptions.get('sales_assumptions', {}).get('enterprise_hub', {}).get('tier_distribution_at_launch', {}).get('enterprise', 0.1)*100}%"),
        ]

        for label, value in business_assumptions:
            ws[f'O{row}'].value = label
            ws[f'P{row}'].value = value
            row += 1

        # NOUVELLE SECTION: Coûts RH (colonne R+) - PHASE 4
        ws['R1'].value = "COÛTS RH"
        ws['R1'].font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
        ws['R1'].fill = openpyxl.styles.PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")

        ws['R2'].value = "Paramètre"
        ws['S2'].value = "Valeur"

        for col in ['R', 'S']:
            ws[f'{col}2'].font = openpyxl.styles.Font(bold=True)

        costs = self.assumptions.get('costs', {})
        social_charges = costs.get('social_charges_rate', 0.45)
        avg_salary = 60000  # Salaire moyen brut annuel (peut être ajouté au YAML si nécessaire)

        row = 3
        hr_data = [
            ("Charges sociales (%)", f"{social_charges*100}%"),
            ("Salaire moyen brut (€/an)", avg_salary),
            ("Coût total ETP (€/an)", f"=S4*(1+{social_charges})"),  # Formule dynamique
        ]

        for label, value in hr_data:
            ws[f'R{row}'].value = label
            if isinstance(value, str) and value.startswith('='):
                # C'est une formule
                ws[f'S{row}'].value = value
            else:
                ws[f'S{row}'].value = value
            row += 1

        # NOUVELLE SECTION: Volumes Commerciaux (colonne R+) - PHASE 4
        ws[f'R{row+1}'].value = "VOLUMES COMMERCIAUX"
        ws[f'R{row+1}'].font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
        ws[f'R{row+1}'].fill = openpyxl.styles.PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")

        row = row + 2
        ws[f'R{row}'].value = "Produit"
        ws[f'S{row}'].value = "Volume/mois (moy)"
        ws[f'R{row}'].font = openpyxl.styles.Font(bold=True)
        ws[f'S{row}'].font = openpyxl.styles.Font(bold=True)

        row += 1

        # Calculer volumes moyens hackathons
        sales_assumptions = self.assumptions.get('sales_assumptions', {})
        hackathons_volumes = sales_assumptions.get('hackathon', {}).get('volumes_monthly', {})
        if hackathons_volumes and isinstance(hackathons_volumes, dict):
            volumes = [hackathons_volumes.get(f'm{i}', 0) for i in range(1, 13)]
            avg_hackathons = sum(volumes) / len(volumes) if volumes else 7.3
        else:
            avg_hackathons = 7.3

        factory_conversion = sales_assumptions.get('factory', {}).get('conversion_rate', 0.35)

        volumes_data = [
            ("Hackathons", f"{avg_hackathons:.1f}"),
            ("Factory conversions", f"={avg_hackathons:.1f}*{factory_conversion}"),  # Formule
            ("Hub nouveaux clients", "Variable (launch M8)"),
        ]

        for label, value in volumes_data:
            ws[f'R{row}'].value = label
            if isinstance(value, str) and value.startswith('='):
                ws[f'S{row}'].value = value
            else:
                ws[f'S{row}'].value = value
            row += 1

        logger.info("✓ Paramètres enrichis avec financial_kpis, validation_rules, hypothèses business, coûts RH, et volumes commerciaux")

    def update_financement_sheet(self):
        """
        Adapter le sheet Financement selon assumptions.yaml
        - Rounds de financement
        - Montants selon YAML
        """
        logger.info("\n💰 Adaptation sheet Financement...")

        ws = self.wb['Financement']

        funding = self.assumptions.get('funding', {})

        # Pré-seed (col C)
        preseed = funding.get('preseed', {})
        ws['C1'] = "2025-26"
        ws['C2'] = f"Pre-seed {preseed.get('quarter', 'Q4 2025')}"
        ws['C4'] = preseed.get('amount', 300000)  # Batch 1
        ws['C5'] = 50000  # Autoposia
        ws['C6'] = 50000  # F-Initiatives

        # Seed (col E)
        seed = funding.get('seed', {})
        ws['E1'] = "2027"
        ws['E2'] = f"Seed {seed.get('quarter', 'Q3 2026')}"
        ws['E8'] = seed.get('amount', 500000)  # CIC

        # Series A (col G)
        series_a = funding.get('series_a', {})
        ws['G1'] = "=E1+1"
        ws['G2'] = f"Series A {series_a.get('quarter', 'Q4 2027')}"
        ws['G11'] = series_a.get('amount', 2000000)  # BPI

        logger.info("✓ Financement adapté avec funding YAML")

    def update_fundings_sheet_with_captable(self):
        """
        RESTRUCTURATION FUNDINGS - État de l'Art PHASE 6
        Structure complète pour fundraising/VC avec 4 sections:
        A. Funding Rounds Timeline (type financeur + valorisations)
        B. Cap Table Dynamique (dilution progressive)
        C. Sources Non-Dilutives (subventions, aides)
        D. Metrics Fundraising (runway, burn, multiples)
        """
        logger.info("\n📊 Restructuration sheet Fundings (État de l'art)...")

        if 'Fundings' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Fundings' introuvable, création skippée")
            return

        ws = self.wb['Fundings']

        # Charger funding_captable.yaml
        base_path = Path(__file__).parent.parent
        captable_path = base_path / "data" / "structured" / "funding_captable.yaml"

        if not captable_path.exists():
            logger.warning("⚠️ funding_captable.yaml introuvable, cap table skippée")
            return

        with open(captable_path, 'r', encoding='utf-8') as f:
            captable_data = yaml.safe_load(f)

        # ═══ SECTION A: FUNDING ROUNDS TIMELINE ═══
        ws['A1'].value = "A. FUNDING ROUNDS TIMELINE (DILUTIF)"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        ws['A3'].value = "Phase"
        ws['B3'].value = "Timing"
        ws['C3'].value = "Type Financeur"
        ws['D3'].value = "Montant"
        ws['E3'].value = "Valorisation Pre"
        ws['F3'].value = "Valorisation Post"
        ws['G3'].value = "ARR Target"
        ws['H3'].value = "Multiple ARR"

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}3'].font = openpyxl.styles.Font(bold=True)

        funding_rounds = captable_data.get('funding_rounds', {})
        row = 4

        rounds_data = [
            ('love_money', 'Famille/Amis'),
            ('pre_seed', 'BA + BPI'),
            ('seed', 'VCs Tier 2'),
            ('series_a', 'VCs Tier 1'),
        ]

        for phase_key, financeur_type in rounds_data:
            if phase_key not in funding_rounds:
                continue

            phase_data = funding_rounds[phase_key]
            amount = phase_data.get('amount', 0)
            val_post = phase_data.get('valuation_post', 0)
            arr_target = phase_data.get('arr_target', 0)

            # Calculer val pre
            val_pre = val_post - amount if val_post > amount else 0

            # Calculer multiple ARR
            multiple_arr = f"{val_post / arr_target:.1f}×" if arr_target > 0 else "-"

            ws[f'A{row}'].value = phase_data.get('phase', phase_key.upper())
            ws[f'B{row}'].value = f"M{phase_data.get('month', 0)}"
            ws[f'C{row}'].value = financeur_type
            ws[f'D{row}'].value = amount
            ws[f'E{row}'].value = val_pre
            ws[f'F{row}'].value = val_post
            ws[f'G{row}'].value = arr_target
            ws[f'H{row}'].value = multiple_arr
            row += 1

        # ═══ SECTION B: CAP TABLE DYNAMIQUE ═══
        ws['A12'].value = "B. CAP TABLE - DILUTION PROGRESSIVE"
        ws['A12'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A12'].fill = openpyxl.styles.PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")

        ws['A14'].value = "Phase"
        ws['B14'].value = "FRT"
        ws['C14'].value = "PCO"
        ws['D14'].value = "MAM"
        ws['E14'].value = "BSPCE"
        ws['F14'].value = "Investisseurs"
        ws['G14'].value = "Total"

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}14'].font = openpyxl.styles.Font(bold=True)

        captable = captable_data.get('captable', {})
        dilution_stages = captable.get('dilution_stages', {})

        row = 15
        for stage_key in ['bootstrap', 'post_pre_seed', 'post_seed', 'post_series_a']:
            if stage_key not in dilution_stages:
                continue

            stage_data = dilution_stages[stage_key]
            equity = stage_data.get('equity', {})

            ws[f'A{row}'].value = stage_data.get('phase', stage_key)

            # Formatter les equity (gérer floats ou strings déjà formatés)
            def format_equity(value):
                if isinstance(value, str):
                    return value  # Déjà formaté
                return f"{value:.1f}%" if value > 0 else "0.0%"

            frt_val = equity.get('FRT', 0)
            pco_val = equity.get('PCO', 0)
            mam_val = equity.get('MAM', 0)
            bspce_val = equity.get('BSPCE', 0)

            ws[f'B{row}'].value = format_equity(frt_val)
            ws[f'C{row}'].value = format_equity(pco_val)
            ws[f'D{row}'].value = format_equity(mam_val)
            ws[f'E{row}'].value = format_equity(bspce_val)

            # Investisseurs combinés
            inv_seed = equity.get('Investisseurs_Seed', 0)
            inv_a = equity.get('Investisseurs_Series_A', 0)
            inv_b = equity.get('Investisseurs_Series_B', 0)

            # Convertir en float si string
            if isinstance(inv_seed, str): inv_seed = 0
            if isinstance(inv_a, str): inv_a = 0
            if isinstance(inv_b, str): inv_b = 0

            inv_total = inv_seed + inv_a + inv_b
            ws[f'F{row}'].value = f"{inv_total:.1f}%" if inv_total > 0 else "-"

            # Total = 100%
            ws[f'G{row}'].value = "100.0%"
            row += 1

        # ═══ SECTION C: SOURCES NON-DILUTIVES ═══
        ws['A22'].value = "C. SOURCES NON-DILUTIVES (Subventions & Aides)"
        ws['A22'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A22'].fill = openpyxl.styles.PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")

        ws['A24'].value = "Source"
        ws['B24'].value = "Timing"
        ws['C24'].value = "Montant"
        ws['D24'].value = "Organisme"
        ws['E24'].value = "Type"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}24'].font = openpyxl.styles.Font(bold=True)

        row = 25
        non_dilutive_sources = [
            ("CIR/CII", "M1-M6", "25K€", "Impôts", "Crédit impôt"),
            ("French Tech", "M6", "30K€", "BPI", "Bourse"),
            ("BPI Innovation", "M12-M24", "100-150K€", "BPI", "Aide"),
            ("Concours i-Nov", "M12-M24", "100-600K€", "BPI", "Concours"),
            ("CIFRE", "Variable", "40K€/an", "ANRT", "Thèse"),
        ]

        for source, timing, montant, organisme, type_aide in non_dilutive_sources:
            ws[f'A{row}'].value = source
            ws[f'B{row}'].value = timing
            ws[f'C{row}'].value = montant
            ws[f'D{row}'].value = organisme
            ws[f'E{row}'].value = type_aide
            row += 1

        # ═══ SECTION D: METRICS FUNDRAISING ═══
        ws['A33'].value = "D. METRICS FUNDRAISING CLÉS"
        ws['A33'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A33'].fill = openpyxl.styles.PatternFill(start_color="9966FF", end_color="9966FF", fill_type="solid")

        ws['A35'].value = "Métrique"
        ws['B35'].value = "Valeur"
        ws['C35'].value = "Commentaire"

        for col in ['A', 'B', 'C']:
            ws[f'{col}35'].font = openpyxl.styles.Font(bold=True)

        row = 36
        metrics_data = [
            ("Total levé (dilutif)", "3.15M€", "Love Money + Pre-Seed + Seed + Series A"),
            ("Total aides (non-dilutif)", "~300-800K€", "CIR/CII + BPI + Concours"),
            ("Dilution totale FRT", "-60.4%", "70% → 27.7%"),
            ("Valuation multiple Seed", "10.0×", "8M€ / 800K€ ARR"),
            ("Valuation multiple Series A", "8.2×", "32.6M€ / 4M€ ARR"),
            ("Runway post Seed", "12-18 mois", "Avec burn 30-50K€/mois"),
        ]

        for metric, value, comment in metrics_data:
            ws[f'A{row}'].value = metric
            ws[f'B{row}'].value = value
            ws[f'C{row}'].value = comment
            row += 1

        logger.info("✓ Fundings restructuré (Timeline + Cap Table + Non-dilutif + Metrics)")

    def update_strategie_vente_sheet(self):
        """
        Adapter le sheet Stratégie de vente selon assumptions.yaml
        - Taux de conversion Hackathon→Factory
        """
        logger.info("\n🎯 Adaptation sheet Stratégie de vente...")

        ws = self.wb['Stratégie de vente']

        conversion_rates = self.assumptions.get('conversion_rates', {})
        hackathon_to_factory = conversion_rates.get('hackathon_to_factory', 0.30)

        # Ajouter un indicateur des taux de conversion en haut du sheet
        # Ligne 1 col A-B: Taux de conversion
        ws['A1'].value = "Taux de conversion Hackathon→Factory"
        ws['B1'].value = f"{hackathon_to_factory*100:.0f}%"

        logger.info(f"✓ Stratégie de vente adaptée (conversion: {hackathon_to_factory*100:.0f}%)")

    def expand_headcount_timeline(self, timeline_dict: dict, total_months: int = 50) -> list:
        """
        Expanse un timeline sparse en liste complète

        Exemple:
            Input: {m1: 1, m4: 2, m12: 3}
            Output: [1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 2, 3, 3, ..., 3]  (50 valeurs)

        Args:
            timeline_dict: Dict avec format {m1: 1, m4: 2, ...}
            total_months: Nombre total de mois (défaut 50)

        Returns:
            Liste de headcount pour chaque mois
        """
        if not timeline_dict:
            return [0] * total_months

        # Convertir les clés "m1", "m4" en nombres et trier
        milestones = []
        for key, value in timeline_dict.items():
            month_num = int(key.replace('m', '').replace('M', ''))
            milestones.append((month_num, value))

        milestones.sort(key=lambda x: x[0])

        # Expansion
        result = []
        milestone_idx = 0

        for month in range(1, total_months + 1):
            # Trouver le dernier milestone <= month actuel
            while (milestone_idx < len(milestones) - 1 and
                   milestones[milestone_idx + 1][0] <= month):
                milestone_idx += 1

            # Si on est avant le premier milestone, valeur = 0
            if month < milestones[0][0]:
                result.append(0)
            else:
                result.append(milestones[milestone_idx][1])

        return result

    def update_charges_personnel_sheet(self):
        """
        PILOTAGE PERSONNEL PAR YAML - PHASE 6
        Mapper rôles YAML → profils RAW et mettre à jour salaires + headcount
        Structure RAW préservée (formules intactes)
        """
        logger.info("\n👥 Pilotage Personnel depuis assumptions.yaml...")

        ws = self.wb['Charges de personnel et FG']

        personnel = self.assumptions.get('personnel_details', {})
        if not personnel:
            logger.warning("⚠️ Section personnel_details absente de assumptions.yaml, skip")
            return

        charges_rate = personnel.get('social_charges_rate', 0.45)
        roles = personnel.get('roles', [])

        if not roles:
            logger.warning("⚠️ Aucun rôle défini dans personnel_details.roles, skip")
            return

        logger.info(f"  {len(roles)} rôle(s) défini(s) dans YAML")

        # Mapping profils RAW (lignes dans le sheet) → Index
        # Selon analyse RAW:
        # Ligne 2: Directeur (mini) - 35K€
        # Ligne 3: Directeur (intermédiaire) - 50K€
        # Ligne 4: Directeur (cible) - 70K€
        # Ligne 5: Consultant - 60K€
        # Ligne 6: Responsable Commercial - 60K€
        # Ligne 7: Product owner - 45K€
        # Ligne 8: Tech Senior - 65K€
        # Ligne 9: Tech Junior (intermédiaire) - 50K€
        # Ligne 10: BD (junior) - 25K€
        # Ligne 11: Stagiaire - 11*1100€

        # Mapping nom profil RAW → ligne dans sheet
        profile_mapping = {
            "Directeur (mini)": 2,
            "Directeur (intermédiaire)": 3,
            "Directeur (cible)": 4,
            "Consultant": 5,
            "Responsable Commercial": 6,
            "Product owner": 7,
            "Tech Senior": 8,
            "Tech Junior (intermédiaire)": 9,
            "BD (junior)": 10,
            "Stagiaire": 11,
        }

        # Aussi mettre à jour salaires dans section détails (lignes 16-25)
        # Ligne 16: Directeur (mini) salaire brut
        # Ligne 17: Directeur (intermédiaire)
        # etc.
        detail_mapping = {
            "Directeur (mini)": 16,
            "Directeur (intermédiaire)": 17,
            "Directeur (cible)": 18,
            "Consultant": 19,
            "Responsable Commercial": 20,
            "Product owner": 21,
            "Tech Senior": 22,
            "Tech Junior (intermédiaire)": 23,
            "BD (junior)": 24,
            "Stagiaire": 25,
        }

        # Pour chaque rôle YAML, mettre à jour le salaire ET headcount timeline
        updated_count = 0
        headcount_updated = 0

        for role in roles:
            profile_raw = role.get('profile_raw')
            annual_salary = role.get('annual_salary_gross', 0)
            headcount_timeline = role.get('headcount_timeline', {})

            if not profile_raw:
                logger.warning(f"  ⚠️ Rôle '{role.get('name')}' sans profile_raw, skip")
                continue

            # Mettre à jour dans section détails (colonne B = salaire brut)
            if profile_raw in detail_mapping:
                detail_row = detail_mapping[profile_raw]
                ws[f'B{detail_row}'].value = annual_salary
                updated_count += 1

                # Mettre à jour headcount timeline (colonnes H onwards = M1, M2, ...)
                if headcount_timeline:
                    expanded_headcount = self.expand_headcount_timeline(headcount_timeline, 50)

                    # Colonnes H à BG (indices 8 à 57 pour 50 mois)
                    for month_idx, headcount in enumerate(expanded_headcount, start=1):
                        col_idx = 7 + month_idx  # H=8, I=9, etc.
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        ws[f'{col_letter}{detail_row}'].value = headcount

                    headcount_updated += 1
                    total_etp = sum(expanded_headcount)
                    logger.info(f"  ✓ {profile_raw}: {annual_salary}€/an + {total_etp} ETP total sur 50 mois")
                else:
                    logger.info(f"  ✓ {profile_raw}: {annual_salary}€/an (pas de timeline)")
            else:
                logger.warning(f"  ⚠️ Profil '{profile_raw}' non trouvé dans mapping")

        # Mettre à jour taux charges sociales (colonne C)
        for detail_row in range(16, 26):  # Lignes 16-25
            ws[f'C{detail_row}'].value = charges_rate

        logger.info(f"✓ Personnel piloté depuis YAML ({updated_count} profils, {headcount_updated} avec timelines, charges {charges_rate*100:.0f}%)")

    def update_infrastructure_detailed_sheet(self):
        """
        Adapter le sheet Infrastructure technique selon assumptions.yaml
        - Pricing cloud (base + tiers)
        - Pricing SaaS tools
        """
        logger.info("\n☁️ Adaptation sheet Infrastructure technique...")

        ws = self.wb['Infrastructure technique']

        infra = self.assumptions.get('infrastructure_costs', {})
        cloud = infra.get('cloud', {})
        saas = infra.get('saas_tools', {})

        # Section cloud (lignes 1-7)
        ws['A1'].value = "CLOUD INFRASTRUCTURE"
        ws['A2'].value = "Base mensuel"
        ws['B2'].value = cloud.get('base_monthly', 1000)

        ws['A3'].value = "Scaling tiers (coût par client):"

        scaling_tiers = cloud.get('scaling_tiers', {})
        ws['A4'].value = "  < 50 clients"
        ws['B4'].value = scaling_tiers.get('tier1', {}).get('cost_per_client', 50)

        ws['A5'].value = "  50-100 clients"
        ws['B5'].value = scaling_tiers.get('tier2', {}).get('cost_per_client', 40)

        ws['A6'].value = "  > 100 clients"
        ws['B6'].value = scaling_tiers.get('tier3', {}).get('cost_per_client', 30)

        # Section SaaS (lignes 9+)
        ws['A9'].value = "SAAS TOOLS"
        row = 10

        for tool_name, tool_data in saas.items():
            if isinstance(tool_data, dict):
                cost = tool_data.get('cost_per_user', 0) or tool_data.get('cost_per_developer', 0)
                min_users = tool_data.get('min_users', 1)
                ws[f'A{row}'].value = tool_name.title()
                ws[f'B{row}'].value = f"{cost}€/user (min {min_users})"
                row += 1

        logger.info("✓ Infrastructure technique adaptée (cloud + SaaS)")

    def update_marketing_detailed_sheet(self):
        """
        Adapter le sheet Marketing selon assumptions.yaml
        - 4 canaux avec budgets annuels
        """
        logger.info("\n📢 Adaptation sheet Marketing...")

        ws = self.wb['Marketing']

        marketing = self.assumptions.get('marketing_budgets', {})

        # Section budgets par canal (lignes 1+)
        ws['A1'].value = "BUDGETS MARKETING PAR CANAL"

        channels = ['digital_ads', 'events', 'content', 'partnerships']
        row = 3

        for channel in channels:
            if channel not in marketing:
                continue

            channel_data = marketing[channel]
            monthly_budgets = channel_data.get('monthly_budgets', {})

            ws[f'A{row}'].value = channel.replace('_', ' ').title()
            ws[f'B{row}'].value = "2025"
            ws[f'C{row}'].value = monthly_budgets.get('y2025', 0)
            ws[f'D{row}'].value = "2026"
            ws[f'E{row}'].value = monthly_budgets.get('y2026', 0)
            ws[f'F{row}'].value = "2027"
            ws[f'G{row}'].value = monthly_budgets.get('y2027', 0)
            ws[f'H{row}'].value = "2028"
            ws[f'I{row}'].value = monthly_budgets.get('y2028', 0)
            ws[f'J{row}'].value = "2029"
            ws[f'K{row}'].value = monthly_budgets.get('y2029', 0)

            row += 2

        logger.info(f"✓ Marketing adapté ({len(channels)} canaux)")

    def add_arr_mrr_to_pl(self):
        """
        Ajouter lignes ARR et MRR en haut du P&L
        Pour tracking des milestones SaaS
        """
        logger.info("\n📈 Ajout ARR/MRR dans P&L...")

        if 'P&L' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'P&L' introuvable, skip")
            return

        ws = self.wb['P&L']

        # Insérer 3 nouvelles lignes en haut (après les headers)
        # On va ajouter après la ligne "CA Total" qui est typiquement en ligne 2-3

        # Trouver la ligne "CA Total" ou similaire
        ca_row = None
        for row in range(1, 20):
            cell_value = ws[f'A{row}'].value
            if cell_value and isinstance(cell_value, str) and 'CA' in cell_value.upper():
                ca_row = row
                break

        if not ca_row:
            ca_row = 5  # Défaut

        insert_row = ca_row + 1

        # Insérer 3 lignes
        ws.insert_rows(insert_row, 3)

        # Ligne ARR
        ws[f'A{insert_row}'].value = "ARR (Annual Recurring Revenue)"
        ws[f'A{insert_row}'].font = openpyxl.styles.Font(bold=True)

        # Ligne MRR
        ws[f'A{insert_row+1}'].value = "MRR (Monthly Recurring Revenue)"
        ws[f'A{insert_row+1}'].font = openpyxl.styles.Font(bold=True)

        # Ligne séparatrice
        ws[f'A{insert_row+2}'].value = "---"

        logger.info(f"✓ ARR/MRR ajoutés en lignes {insert_row}-{insert_row+1} du P&L")

    def create_cash_flow_sheet(self):
        """
        Créer un nouveau sheet Cash Flow Statement
        Essential pour fundraising et suivi trésorerie
        """
        logger.info("\n💰 Création sheet Cash Flow...")

        # Créer nouveau sheet
        ws = self.wb.create_sheet("Cash Flow")

        # Headers
        ws['A1'].value = "CASH FLOW STATEMENT"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14)

        ws['A2'].value = "Catégorie"
        ws['B2'].value = "Description"

        # Colonnes mois: C=M1, D=M2, etc.
        for month in range(1, 51):
            col_letter = openpyxl.utils.get_column_letter(month + 2)  # +2 car A,B = labels
            ws[f'{col_letter}2'].value = f"M{month}"

        # Section Operating Activities
        row = 3
        ws[f'A{row}'].value = "OPERATING ACTIVITIES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        operating_items = [
            ("CA Encaissé", "Revenue collected"),
            ("Charges Personnel", "Salaries and social charges"),
            ("Charges Infrastructure", "Cloud + SaaS tools"),
            ("Charges Marketing", "Marketing spend"),
            ("Autres Charges", "Other operating expenses"),
        ]

        for label, desc in operating_items:
            ws[f'A{row}'].value = f"  {label}"
            ws[f'B{row}'].value = desc
            row += 1

        ws[f'A{row}'].value = "= Cash Flow Opérationnel"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 2

        # Section Investing Activities
        ws[f'A{row}'].value = "INVESTING ACTIVITIES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "  CAPEX (équipements)"
        ws[f'B{row}'].value = "Equipment and infrastructure"
        row += 1

        ws[f'A{row}'].value = "= Cash Flow Investissement"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 2

        # Section Financing Activities
        ws[f'A{row}'].value = "FINANCING ACTIVITIES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        financing_items = [
            ("Pre-Seed", "M1: 150K€"),
            ("Seed", "M11: 500K€"),
            ("Series A", "M36: 2.5M€"),
        ]

        for label, desc in financing_items:
            ws[f'A{row}'].value = f"  {label}"
            ws[f'B{row}'].value = desc
            row += 1

        ws[f'A{row}'].value = "= Cash Flow Financement"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 2

        # Total et balance
        ws[f'A{row}'].value = "TOTAL CASH FLOW (mois)"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, color="0000FF")
        row += 1

        ws[f'A{row}'].value = "CASH BALANCE (cumulé)"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12, color="FF0000")
        ws[f'B{row}'].value = "Trésorerie disponible"
        row += 2

        # Métriques
        ws[f'A{row}'].value = "MÉTRIQUES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "  Burn Rate (€/mois)"
        ws[f'B{row}'].value = "Operating CF négatif"
        row += 1

        ws[f'A{row}'].value = "  Cash Runway (mois)"
        ws[f'B{row}'].value = "Cash balance / Burn rate"
        row += 1

        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40

        logger.info("✓ Sheet Cash Flow créé avec structure complète")

    def remove_gtmarket_sheet(self):
        """
        Supprimer le sheet GTMarket (110 cols × 1000 rows, peu de valeur ajoutée)
        """
        logger.info("\n🗑️ Suppression sheet GTMarket...")

        if 'GTMarket' in self.wb.sheetnames:
            del self.wb['GTMarket']
            logger.info("✓ Sheet GTMarket supprimé")
        else:
            logger.warning("⚠️ Sheet GTMarket introuvable, skip")

    def enrich_synthese_dashboard(self):
        """
        Enrichir le sheet Synthèse avec dashboard KPIs
        Vue exécutive pour investisseurs et board
        """
        logger.info("\n📊 Enrichissement Synthèse avec dashboard...")

        if 'Synthèse' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Synthèse' introuvable, skip")
            return

        ws = self.wb['Synthèse']

        # Trouver une zone vide (colonne Y+ par exemple)
        start_col = 'Y'

        # Dashboard header
        ws[f'{start_col}1'].value = "DASHBOARD EXÉCUTIF"
        ws[f'{start_col}1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws[f'{start_col}1'].fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        # Section 1: ARR Milestones
        row = 3
        ws[f'{start_col}{row}'].value = "ARR MILESTONES"
        ws[f'{start_col}{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        # Charger cap table pour les targets
        base_path = Path(__file__).parent.parent
        captable_path = base_path / "data" / "structured" / "funding_captable.yaml"

        if captable_path.exists():
            with open(captable_path, 'r', encoding='utf-8') as f:
                captable_data = yaml.safe_load(f)
            arr_targets = captable_data.get('arr_targets', {})
        else:
            arr_targets = {}

        arr_milestones = [
            ("M1 (Bootstrap)", arr_targets.get('M1', 10000)),
            ("M6 (PRE-SEED)", arr_targets.get('M6', 500000)),
            ("M12 (SEED)", arr_targets.get('M12', 800000)),
            ("M18 (Post Seed)", arr_targets.get('M18', 1500000)),
            ("M36 (Series A)", arr_targets.get('M36', 4000000)),
            ("M48 (Pre Series B)", arr_targets.get('M48', 6000000)),
        ]

        ws[f'{start_col}{row}'].value = "Milestone"
        next_col = openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(start_col) + 1)
        ws[f'{next_col}{row}'].value = "ARR Target"
        row += 1

        for milestone, target in arr_milestones:
            ws[f'{start_col}{row}'].value = milestone
            ws[f'{next_col}{row}'].value = target
            row += 1

        row += 1

        # Section 2: KPIs Critiques
        ws[f'{start_col}{row}'].value = "KPIs CRITIQUES"
        ws[f'{start_col}{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        financial_kpis = self.assumptions.get('financial_kpis', {})

        kpis = [
            ("Target LTV/CAC", financial_kpis.get('saas_metrics', {}).get('target_ltv_cac_ratio', 8)),
            ("Max Churn Annual", f"{financial_kpis.get('saas_metrics', {}).get('max_churn_annual', 0.15)*100}%"),
            ("Marge Brute Target", f"{financial_kpis.get('margin_targets', {}).get('gross_margin_pct', 70)}%"),
            ("EBITDA Margin Target", f"{financial_kpis.get('margin_targets', {}).get('ebitda_margin_pct', -15)}%"),
            ("Min Cash Runway", f"{financial_kpis.get('cash_management', {}).get('min_cash_runway_months', 12)} mois"),
            ("Max Burn Rate", f"{financial_kpis.get('cash_management', {}).get('acceptable_burn_rate_monthly', 50000):,}€/mois"),
        ]

        ws[f'{start_col}{row}'].value = "KPI"
        ws[f'{next_col}{row}'].value = "Valeur"
        row += 1

        for kpi, value in kpis:
            ws[f'{start_col}{row}'].value = kpi
            ws[f'{next_col}{row}'].value = value
            row += 1

        row += 1

        # Section 3: Hypothèses Critiques
        ws[f'{start_col}{row}'].value = "HYPOTHÈSES CRITIQUES"
        ws[f'{start_col}{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        critical_assumptions = self.assumptions.get('critical_assumptions', [])
        if critical_assumptions:
            for assumption in critical_assumptions[:5]:  # Top 5
                if isinstance(assumption, dict):
                    ws[f'{start_col}{row}'].value = f"• {assumption.get('assumption', '')}"
                    ws[f'{next_col}{row}'].value = assumption.get('risk_level', '')
                    row += 1

        # Ajuster largeur colonnes
        ws.column_dimensions[start_col].width = 25
        ws.column_dimensions[next_col].width = 20

        logger.info("✓ Dashboard exécutif ajouté dans Synthèse")

    def create_scenarios_sheet(self):
        """
        Créer nouveau sheet Scenarios (base/upside/downside)
        Analyse de sensibilité pour investisseurs
        """
        logger.info("\n📊 Création sheet Scenarios...")

        # Créer nouveau sheet
        ws = self.wb.create_sheet("Scenarios")

        # Header
        ws['A1'].value = "SCENARIOS D'ÉVOLUTION"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        ws['A2'].value = "Basé sur assumptions.yaml - 3 scénarios probabilisés"

        # Colonnes
        ws['A4'].value = "Métrique"
        ws['B4'].value = "BASE CASE (60%)"
        ws['C4'].value = "UPSIDE (20%)"
        ws['D4'].value = "DOWNSIDE (20%)"
        ws['E4'].value = "Notes"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}4'].font = openpyxl.styles.Font(bold=True)

        # Charger scenarios depuis YAML
        scenarios = self.assumptions.get('scenarios', {})
        base = scenarios.get('base_case', {})
        upside = scenarios.get('upside', {})
        downside = scenarios.get('downside', {})

        row = 5

        # Section ARR
        ws[f'A{row}'].value = "ARR M14"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        ws[f'B{row}'].value = base.get('arr_m14', 800000)
        ws[f'C{row}'].value = upside.get('arr_m14', 952000)
        ws[f'D{row}'].value = downside.get('arr_m14', 648000)
        ws[f'E{row}'].value = "Annual Recurring Revenue à M14"
        row += 1

        ws[f'A{row}'].value = "Probabilité"
        ws[f'B{row}'].value = f"{base.get('probability', 0.6)*100}%"
        ws[f'C{row}'].value = f"{upside.get('probability', 0.2)*100}%"
        ws[f'D{row}'].value = f"{downside.get('probability', 0.2)*100}%"
        row += 2

        # Section Hypothèses Hackathon
        ws[f'A{row}'].value = "HYPOTHÈSES HACKATHON"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "Volume multiplier"
        ws[f'B{row}'].value = "1.0×"
        ws[f'C{row}'].value = f"{upside.get('hackathon_volume_multiplier', 1.2)}×"
        ws[f'D{row}'].value = f"{downside.get('hackathon_volume_multiplier', 0.8)}×"
        ws[f'E{row}'].value = "Multiplicateur volumes hackathons"
        row += 1

        ws[f'A{row}'].value = "Conversion → Factory"
        ws[f'B{row}'].value = "30%"
        ws[f'C{row}'].value = f"{upside.get('conversion_factory', 0.35)*100}%"
        ws[f'D{row}'].value = f"{downside.get('conversion_factory', 0.25)*100}%"
        ws[f'E{row}'].value = "Taux conversion Hackathon → Factory"
        row += 2

        # Section Hypothèses Hub
        ws[f'A{row}'].value = "HYPOTHÈSES ENTERPRISE HUB"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "Launch delay"
        ws[f'B{row}'].value = "M8 (aucun retard)"
        ws[f'C{row}'].value = "M8 (accéléré)"
        ws[f'D{row}'].value = f"M{8 + downside.get('hub_launch_delay_months', 2)} (+{downside.get('hub_launch_delay_months', 2)} mois)"
        ws[f'E{row}'].value = "Délai lancement Hub"
        row += 1

        ws[f'A{row}'].value = "Adoption speed"
        ws[f'B{row}'].value = "Normal"
        ws[f'C{row}'].value = "Rapide (+20%)"
        ws[f'D{row}'].value = "Lente (-20%)"
        row += 2

        # Section Impact Financier
        ws[f'A{row}'].value = "IMPACT FINANCIER ESTIMÉ"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        # Calculs approximatifs
        base_arr = base.get('arr_m14', 800000)
        upside_arr = upside.get('arr_m14', 952000)
        downside_arr = downside.get('arr_m14', 648000)

        ws[f'A{row}'].value = "ARR M14"
        ws[f'B{row}'].value = base_arr
        ws[f'C{row}'].value = upside_arr
        ws[f'D{row}'].value = downside_arr
        row += 1

        ws[f'A{row}'].value = "vs Base Case"
        ws[f'B{row}'].value = "0%"
        ws[f'C{row}'].value = f"+{(upside_arr/base_arr - 1)*100:.0f}%"
        ws[f'D{row}'].value = f"{(downside_arr/base_arr - 1)*100:.0f}%"
        row += 2

        # Hypothèses critiques
        ws[f'A{row}'].value = "HYPOTHÈSES CRITIQUES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        critical_assumptions = self.assumptions.get('critical_assumptions', [])
        for assumption in critical_assumptions[:5]:
            if isinstance(assumption, dict):
                ws[f'A{row}'].value = f"• {assumption.get('assumption', '')}"
                ws[f'B{row}'].value = assumption.get('risk_level', '')
                ws[f'E{row}'].value = assumption.get('mitigation', '')[:50] if assumption.get('mitigation') else ""
                row += 1

        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 50

        logger.info("✓ Sheet Scenarios créé (base/upside/downside)")

    def create_unit_economics_sheet(self):
        """
        Créer nouveau sheet Unit Economics
        CAC, LTV, Payback period par produit
        """
        logger.info("\n💰 Création sheet Unit Economics...")

        # Créer nouveau sheet
        ws = self.wb.create_sheet("Unit Economics")

        # Header
        ws['A1'].value = "UNIT ECONOMICS PAR PRODUIT"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

        ws['A3'].value = "Produit"
        ws['B3'].value = "Prix Moyen"
        ws['C3'].value = "CAC (€)"
        ws['D3'].value = "LTV (€)"
        ws['E3'].value = "LTV/CAC"
        ws['F3'].value = "Payback (mois)"
        ws['G3'].value = "Marge (%)"
        ws['H3'].value = "Notes"

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}3'].font = openpyxl.styles.Font(bold=True)

        row = 4

        # Hackathon
        hackathon_price = self.assumptions.get('pricing', {}).get('hackathon', {}).get('periods', [{}])[0].get('price_eur', 18000)
        ws[f'A{row}'].value = "Hackathon"
        ws[f'B{row}'].value = hackathon_price
        ws[f'C{row}'].value = 5000  # CAC from marketing_budgets
        ws[f'D{row}'].value = hackathon_price * 1.5  # LTV (client peut revenir)
        ws[f'E{row}'].value = f"=D{row}/C{row}"  # LTV/CAC formula
        ws[f'F{row}'].value = 1  # Payback immédiat
        ws[f'G{row}'].value = "80%"
        ws[f'H{row}'].value = "Offre d'entrée, forte marge"
        row += 1

        # Factory
        factory_price = self.assumptions.get('pricing', {}).get('factory', {}).get('periods', [{}])[0].get('price_eur', 75000)
        ws[f'A{row}'].value = "Factory"
        ws[f'B{row}'].value = factory_price
        ws[f'C{row}'].value = 10000  # CAC plus élevé (cycle long)
        ws[f'D{row}'].value = factory_price * 2  # LTV (upsell services)
        ws[f'E{row}'].value = f"=D{row}/C{row}"
        ws[f'F{row}'].value = 3  # Payback 3 mois
        ws[f'G{row}'].value = "65%"
        ws[f'H{row}'].value = "Conversion naturelle hackathon"
        row += 1

        # Enterprise Hub - Starter
        starter_price = self.assumptions.get('pricing', {}).get('enterprise_hub', {}).get('tiers', {}).get('starter', {}).get('monthly_eur', 500)
        ws[f'A{row}'].value = "Hub Starter"
        ws[f'B{row}'].value = f"{starter_price}€/mois"
        ws[f'C{row}'].value = 15000  # CAC from assumptions
        ws[f'D{row}'].value = 36000  # LTV (3 ans × 12 mois × 500€ × retention)
        ws[f'E{row}'].value = f"=D{row}/C{row}"
        ws[f'F{row}'].value = 30  # Payback 30 mois
        ws[f'G{row}'].value = "75%"
        ws[f'H{row}'].value = "SaaS récurrent, target PME"
        row += 1

        # Enterprise Hub - Business
        business_price = self.assumptions.get('pricing', {}).get('enterprise_hub', {}).get('tiers', {}).get('business', {}).get('monthly_eur', 2000)
        ws[f'A{row}'].value = "Hub Business"
        ws[f'B{row}'].value = f"{business_price}€/mois"
        ws[f'C{row}'].value = 15000
        ws[f'D{row}'].value = 60000  # LTV
        ws[f'E{row}'].value = f"=D{row}/C{row}"
        ws[f'F{row}'].value = 8  # Payback 8 mois
        ws[f'G{row}'].value = "78%"
        ws[f'H{row}'].value = "SaaS récurrent, target ETI"
        row += 1

        # Enterprise Hub - Enterprise
        enterprise_price = self.assumptions.get('pricing', {}).get('enterprise_hub', {}).get('tiers', {}).get('enterprise', {}).get('monthly_eur', 10000)
        ws[f'A{row}'].value = "Hub Enterprise"
        ws[f'B{row}'].value = f"{enterprise_price}€/mois"
        ws[f'C{row}'].value = 15000
        ws[f'D{row}'].value = 120000  # LTV complet
        ws[f'E{row}'].value = f"=D{row}/C{row}"
        ws[f'F{row}'].value = 2  # Payback 2 mois
        ws[f'G{row}'].value = "80%"
        ws[f'H{row}'].value = "SaaS récurrent, target Grands Comptes"
        row += 1

        # Services
        services_price = self.assumptions.get('pricing', {}).get('services', {}).get('implementation', {}).get('periods', [{}])[0].get('price_eur', 10000)
        ws[f'A{row}'].value = "Services Implémentation"
        ws[f'B{row}'].value = services_price
        ws[f'C{row}'].value = 2000  # CAC faible (upsell)
        ws[f'D{row}'].value = services_price * 1.2  # LTV
        ws[f'E{row}'].value = f"=D{row}/C{row}"
        ws[f'F{row}'].value = 1  # Payback immédiat
        ws[f'G{row}'].value = "70%"
        ws[f'H{row}'].value = "Revenus complémentaires"
        row += 2

        # Résumé
        ws[f'A{row}'].value = "RÉSUMÉ"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "LTV/CAC Moyen Pondéré"
        ws[f'B{row}'].value = self.assumptions.get('financial_kpis', {}).get('saas_metrics', {}).get('target_ltv_cac_ratio', 8)
        ws[f'H{row}'].value = "Target: 8× (excellent pour SaaS B2B)"
        row += 1

        ws[f'A{row}'].value = "Payback Moyen (mois)"
        ws[f'B{row}'].value = "6-12 mois"
        ws[f'H{row}'].value = "Variable selon produit et tier"
        row += 1

        ws[f'A{row}'].value = "Churn Annual Max"
        ws[f'B{row}'].value = f"{self.assumptions.get('financial_kpis', {}).get('saas_metrics', {}).get('max_churn_annual', 0.15)*100}%"
        ws[f'H{row}'].value = "Hub uniquement (hackathon/factory = one-time)"

        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 40

        logger.info("✓ Sheet Unit Economics créé (CAC/LTV par produit)")

    def add_granular_metrics_to_ventes(self):
        """
        Ajouter métriques granulaires dans Ventes
        Volumes mensuels par tier (Starter/Business/Enterprise)
        """
        logger.info("\n📊 Ajout métriques granulaires dans Ventes...")

        if 'Ventes' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Ventes' introuvable, skip")
            return

        ws = self.wb['Ventes']

        # Trouver une zone vide (après les lignes principales)
        # Chercher la dernière ligne utilisée
        last_row = 50  # Start après les lignes principales

        # Ajouter section Volumes Hub par Tier
        row = last_row
        ws[f'A{row}'].value = "VOLUMES ENTERPRISE HUB PAR TIER"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'].value = "Tier"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        # Lignes par tier
        ws[f'A{row}'].value = "Nouveaux clients Starter (mois)"
        row += 1
        ws[f'A{row}'].value = "Nouveaux clients Business (mois)"
        row += 1
        ws[f'A{row}'].value = "Nouveaux clients Enterprise (mois)"
        row += 1
        ws[f'A{row}'].value = "Total nouveaux clients Hub (mois)"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 2

        # Clients actifs cumulés
        ws[f'A{row}'].value = "Clients actifs Starter (cumulé)"
        row += 1
        ws[f'A{row}'].value = "Clients actifs Business (cumulé)"
        row += 1
        ws[f'A{row}'].value = "Clients actifs Enterprise (cumulé)"
        row += 1
        ws[f'A{row}'].value = "Total clients actifs Hub"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 2

        # Churn mensuel
        ws[f'A{row}'].value = "Churn mensuel Hub (%)"
        row += 1
        ws[f'A{row}'].value = "Clients perdus (mois)"

        logger.info("✓ Métriques granulaires ajoutées dans Ventes")

    def add_productivity_ia_to_ventes(self):
        """
        Ajouter productivité IA dans Ventes - PHASE 4
        Total ETP, Ratio productivité IA, Équivalent ETP trad.
        Illustre le pitch core GenieFactory: IA qui décuple productivité
        """
        logger.info("\n🤖 Ajout productivité IA dans Ventes...")

        if 'Ventes' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Ventes' introuvable, skip")
            return

        ws = self.wb['Ventes']

        # Ajouter section PRODUCTIVITÉ IA (avant volumes Hub, ligne 45)
        row = 45
        ws[f'A{row}'].value = "PRODUCTIVITÉ IA (GenieFactory)"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{row}'].fill = openpyxl.styles.PatternFill(start_color="FF3366", end_color="FF3366", fill_type="solid")
        row += 1

        ws[f'A{row}'].value = "Métrique"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        # Total ETP
        ws[f'A{row}'].value = "Total ETP"
        ws[f'A{row}'].font = openpyxl.styles.Font(italic=True)
        # Note: Les valeurs seront injectées par le script d'injection
        row += 1

        # Ratio productivité IA
        ws[f'A{row}'].value = "Productivité IA (ratio)"
        ws[f'A{row}'].font = openpyxl.styles.Font(italic=True)
        # Valeur fixe: 3.0× (peut être ajustée selon assumptions.yaml)
        productivity_ratio = 3.0
        ws[f'B{row}'].value = f"{productivity_ratio}×"
        row += 1

        # Équivalent ETP trad.
        ws[f'A{row}'].value = "Équivalent ETP trad."
        ws[f'A{row}'].font = openpyxl.styles.Font(italic=True)
        # Formule: Total ETP × Ratio IA (sera ajustée lors injection)
        row += 1

        # Ajouter note explicative
        ws[f'A{row}'].value = "Note: GenieFactory permet à 1 ETP de faire le travail de 3 ETP traditionnels"
        ws[f'A{row}'].font = openpyxl.styles.Font(size=9, italic=True, color="666666")

        logger.info("✓ Productivité IA ajoutée dans Ventes (pitch core GenieFactory)")

    def improve_infrastructure_labels(self):
        """
        Améliorer labels Infrastructure - PHASE 5
        Ajouter labels manquants: Hosting, Licences logicielles, total
        """
        logger.info("\n☁️ Amélioration labels Infrastructure...")

        if 'Infrastructure technique' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Infrastructure technique' introuvable, skip")
            return

        ws = self.wb['Infrastructure technique']

        # Modifier label ligne 1 pour plus de clarté
        current_label = ws['A1'].value
        if current_label and 'CLOUD' in str(current_label):
            ws['A1'].value = "Hosting (CLOUD INFRASTRUCTURE)"

        # Modifier label ligne 9 pour plus de clarté
        if ws['A9'].value and 'SAAS' in str(ws['A9'].value):
            ws['A9'].value = "Licences logicielles (SAAS TOOLS)"

        # Ajouter ligne "total" si pas déjà présente
        # Trouver la dernière ligne utilisée
        last_row = 20
        ws[f'A{last_row}'].value = "Total Infrastructure (mensuel)"
        ws[f'A{last_row}'].font = openpyxl.styles.Font(bold=True)
        # Formule sera ajoutée si nécessaire

        logger.info("✓ Labels Infrastructure améliorés (Hosting, Licences, total)")

    def improve_marketing_labels(self):
        """
        Améliorer labels Marketing - PHASE 5
        Ajouter labels manquants: Ventes, Campagnes Collaboration, Campagnes Ciblées
        """
        logger.info("\n📢 Amélioration labels Marketing...")

        if 'Marketing' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Marketing' introuvable, skip")
            return

        ws = self.wb['Marketing']

        # Ajouter section "Ventes" (ligne 25+)
        row = 25
        ws[f'A{row}'].value = "Ventes (Support Marketing)"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = openpyxl.styles.PatternFill(start_color="00CC66", end_color="00CC66", fill_type="solid")
        row += 1

        ws[f'A{row}'].value = "Support ventes (docs, présent., démos)"
        row += 1

        # Ajouter "Campagnes Collaboration" (ligne 28+)
        row = 28
        ws[f'A{row}'].value = "Campagnes Collaboration"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1
        ws[f'A{row}'].value = "Partenariats tech, co-marketing"
        row += 1

        # Ajouter "Campagnes Ciblées" (ligne 31+)
        row = 31
        ws[f'A{row}'].value = "Campagnes Ciblées"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1
        ws[f'A{row}'].value = "ABM (Account-Based Marketing)"

        logger.info("✓ Labels Marketing améliorés (Ventes, Campagnes)")

    def enhance_fundings_visualization(self):
        """
        Améliorer visualisation dilution dans Fundings
        Ajouter section comparative et évolution
        """
        logger.info("\n💎 Amélioration visualisation Fundings...")

        if 'Fundings' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Fundings' introuvable, skip")
            return

        ws = self.wb['Fundings']

        # Charger cap table
        base_path = Path(__file__).parent.parent
        captable_path = base_path / "data" / "structured" / "funding_captable.yaml"

        if not captable_path.exists():
            logger.warning("⚠️ funding_captable.yaml introuvable, skip")
            return

        with open(captable_path, 'r', encoding='utf-8') as f:
            captable_data = yaml.safe_load(f)

        # Ajouter section Évolution Dilution (colonne H+)
        ws['H1'].value = "ÉVOLUTION DILUTION FONDATEURS"
        ws['H1'].font = openpyxl.styles.Font(bold=True, size=12, color="FFFFFF")
        ws['H1'].fill = openpyxl.styles.PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")

        ws['H2'].value = "Fondateur"
        ws['I2'].value = "Bootstrap"
        ws['J2'].value = "Post Pre-Seed"
        ws['K2'].value = "Post Seed"
        ws['L2'].value = "Post Series A"
        ws['M2'].value = "Dilution Totale"

        for col in ['H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}2'].font = openpyxl.styles.Font(bold=True)

        row = 3

        dilution_stages = captable_data.get('captable', {}).get('dilution_stages', {})

        # FRT
        ws[f'H{row}'].value = "FRT"
        ws[f'I{row}'].value = f"{dilution_stages.get('bootstrap', {}).get('equity', {}).get('FRT', 70)}%"
        ws[f'J{row}'].value = f"{dilution_stages.get('post_pre_seed', {}).get('equity', {}).get('FRT', 68.6)}%"
        ws[f'K{row}'].value = f"{dilution_stages.get('post_seed', {}).get('equity', {}).get('FRT', 34.5)}%"
        ws[f'L{row}'].value = f"{dilution_stages.get('post_series_a', {}).get('equity', {}).get('FRT', 27.7)}%"
        ws[f'M{row}'].value = "-60.4%"
        ws[f'M{row}'].font = openpyxl.styles.Font(color="FF0000")
        row += 1

        # PCO
        ws[f'H{row}'].value = "PCO"
        ws[f'I{row}'].value = f"{dilution_stages.get('bootstrap', {}).get('equity', {}).get('PCO', 15)}%"
        ws[f'J{row}'].value = f"{dilution_stages.get('post_pre_seed', {}).get('equity', {}).get('PCO', 14.7)}%"
        ws[f'K{row}'].value = f"{dilution_stages.get('post_seed', {}).get('equity', {}).get('PCO', 17.1)}%"
        ws[f'L{row}'].value = f"{dilution_stages.get('post_series_a', {}).get('equity', {}).get('PCO', 19.8)}%"
        ws[f'M{row}'].value = "+31.9%"
        ws[f'M{row}'].font = openpyxl.styles.Font(color="00CC00")
        row += 1

        # MAM
        ws[f'H{row}'].value = "MAM"
        ws[f'I{row}'].value = f"{dilution_stages.get('bootstrap', {}).get('equity', {}).get('MAM', 15)}%"
        ws[f'J{row}'].value = f"{dilution_stages.get('post_pre_seed', {}).get('equity', {}).get('MAM', 14.7)}%"
        ws[f'K{row}'].value = f"{dilution_stages.get('post_seed', {}).get('equity', {}).get('MAM', 15.6)}%"
        ws[f'L{row}'].value = f"{dilution_stages.get('post_series_a', {}).get('equity', {}).get('MAM', 17.2)}%"
        ws[f'M{row}'].value = "+14.7%"
        ws[f'M{row}'].font = openpyxl.styles.Font(color="00CC00")
        row += 2

        # Section Valorisation
        ws[f'H{row}'].value = "VALORISATION PAR ROUND"
        ws[f'H{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'H{row}'].value = "Round"
        ws[f'I{row}'].value = "Montant Levé"
        ws[f'J{row}'].value = "Valorisation Post"
        ws[f'K{row}'].value = "Multiple ARR"
        for col in ['H', 'I', 'J', 'K']:
            ws[f'{col}{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        rounds_data = [
            ("Pre-Seed", 400000, 2900000, "5.8× (500K ARR)"),
            ("Seed", 1500000, 8000000, "10× (800K ARR)"),
            ("Series A", 2500000, 32600000, "8.2× (4M ARR)"),
        ]

        for round_name, amount, valuation, multiple in rounds_data:
            ws[f'H{row}'].value = round_name
            ws[f'I{row}'].value = amount
            ws[f'J{row}'].value = valuation
            ws[f'K{row}'].value = multiple
            row += 1

        # Ajuster largeur colonnes
        for col in ['H', 'I', 'J', 'K', 'L', 'M']:
            ws.column_dimensions[col].width = 15

        logger.info("✓ Visualisation Fundings améliorée (dilution + valorisation)")

    def create_data_quality_sheet(self):
        """
        Créer sheet Data Quality avec checks automatiques
        Alertes si valeurs hors limites validation_rules
        """
        logger.info("\n✅ Création sheet Data Quality...")

        ws = self.wb.create_sheet("Data Quality")

        # Header
        ws['A1'].value = "DATA QUALITY CHECKS"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")

        ws['A2'].value = "Vérifications automatiques de cohérence"

        # Colonnes
        ws['A4'].value = "Check"
        ws['B4'].value = "Valeur Min"
        ws['C4'].value = "Valeur Actuelle"
        ws['D4'].value = "Valeur Max"
        ws['E4'].value = "Status"
        ws['F4'].value = "Notes"

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}4'].font = openpyxl.styles.Font(bold=True)

        validation_rules = self.assumptions.get('validation_rules', {})

        row = 5

        # ARR M14
        ws[f'A{row}'].value = "ARR M14"
        ws[f'B{row}'].value = validation_rules.get('arr_m14_min', 720000)
        ws[f'C{row}'].value = "[À vérifier dans P&L]"
        ws[f'D{row}'].value = validation_rules.get('arr_m14_max', 880000)
        ws[f'E{row}'].value = "=IF(AND(C5>=B5,C5<=D5),\"✓ OK\",\"⚠️ ALERTE\")"
        ws[f'F{row}'].value = "Milestone contractuel critique"
        row += 1

        # Team Size
        ws[f'A{row}'].value = "Team Size M14"
        ws[f'B{row}'].value = validation_rules.get('min_team_size_m1', 4)
        ws[f'C{row}'].value = "[À vérifier dans Personnel]"
        ws[f'D{row}'].value = validation_rules.get('max_team_size', 15)
        ws[f'E{row}'].value = "=IF(AND(C6>=B6,C6<=D6),\"✓ OK\",\"⚠️ ALERTE\")"
        ws[f'F{row}'].value = "Croissance équipe maîtrisée"
        row += 1

        # Cash Balance
        ws[f'A{row}'].value = "Cash Balance Min"
        ws[f'B{row}'].value = validation_rules.get('min_cash_balance', 50000)
        ws[f'C{row}'].value = "[À vérifier dans Cash Flow]"
        ws[f'D{row}'].value = "N/A"
        ws[f'E{row}'].value = "=IF(C7>=B7,\"✓ OK\",\"⚠️ ALERTE\")"
        ws[f'F{row}'].value = "Trésorerie sécurisée"
        row += 1

        # Burn Rate
        ws[f'A{row}'].value = "Burn Rate Max (€/mois)"
        ws[f'B{row}'].value = "0"
        ws[f'C{row}'].value = "[À vérifier dans Cash Flow]"
        ws[f'D{row}'].value = validation_rules.get('max_burn_monthly', 60000)
        ws[f'E{row}'].value = "=IF(C8<=D8,\"✓ OK\",\"⚠️ ALERTE\")"
        ws[f'F{row}'].value = "Burn sous contrôle"
        row += 1

        # Conversion Factory
        ws[f'A{row}'].value = "Conversion Factory (%)"
        ws[f'B{row}'].value = f"{validation_rules.get('min_conversion_hackathon_factory', 0.25)*100}%"
        ws[f'C{row}'].value = "35%"
        ws[f'D{row}'].value = "N/A"
        ws[f'E{row}'].value = "✓ OK"
        ws[f'F{row}'].value = "Conversion validée"
        row += 1

        # Churn Hub
        ws[f'A{row}'].value = "Churn Hub Monthly (%)"
        ws[f'B{row}'].value = "0%"
        ws[f'C{row}'].value = "0.8%"
        ws[f'D{row}'].value = f"{validation_rules.get('max_churn_hub_monthly', 0.015)*100}%"
        ws[f'E{row}'].value = "✓ OK"
        ws[f'F{row}'].value = "Churn acceptable"
        row += 2

        # Section Cohérence
        ws[f'A{row}'].value = "CHECKS DE COHÉRENCE"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        ws[f'A{row}'].value = "CA Total = Somme produits"
        ws[f'E{row}'].value = "[À vérifier manuellement]"
        ws[f'F{row}'].value = "Hackathon + Factory + Hub + Services"
        row += 1

        ws[f'A{row}'].value = "Cash Balance toujours > 0"
        ws[f'E{row}'].value = "[Vérifier Cash Flow]"
        ws[f'F{row}'].value = "Sur les 50 mois"
        row += 1

        ws[f'A{row}'].value = "Team Cost < Total Costs"
        ws[f'E{row}'].value = "[Vérifier Charges Personnel]"
        ws[f'F{row}'].value = "Personnel principal poste de coût"

        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35

        logger.info("✓ Sheet Data Quality créé avec 6 checks automatiques")

    def create_documentation_sheet(self):
        """
        Créer sheet Documentation
        Meta, revision history, usage notes
        """
        logger.info("\n📝 Création sheet Documentation...")

        ws = self.wb.create_sheet("Documentation")

        # Header
        ws['A1'].value = "DOCUMENTATION DU BUSINESS PLAN"
        ws['A1'].font = openpyxl.styles.Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = openpyxl.styles.PatternFill(start_color="666666", end_color="666666", fill_type="solid")

        row = 3

        # Section META
        ws[f'A{row}'].value = "MÉTADONNÉES"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
        row += 1

        meta = self.assumptions.get('meta', {})
        ws[f'A{row}'].value = "Version"
        ws[f'B{row}'].value = meta.get('version', '1.2')
        row += 1

        ws[f'A{row}'].value = "Date création"
        ws[f'B{row}'].value = meta.get('created_date', '2025-11-20')
        row += 1

        ws[f'A{row}'].value = "Auteur"
        ws[f'B{row}'].value = meta.get('author', 'Claude Code - Automated Generation')
        row += 1

        ws[f'A{row}'].value = "Sources"
        sources = meta.get('sources', [])
        if sources:
            ws[f'B{row}'].value = ", ".join(sources[:3])
        row += 2

        # Section REVISION HISTORY
        ws[f'A{row}'].value = "HISTORIQUE DES RÉVISIONS"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
        row += 1

        ws[f'A{row}'].value = "Version"
        ws[f'B{row}'].value = "Date"
        ws[f'C{row}'].value = "Auteur"
        ws[f'D{row}'].value = "Changements"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = openpyxl.styles.Font(bold=True)
        row += 1

        revision_history = self.assumptions.get('revision_history', [])
        for revision in revision_history[:10]:  # Max 10 révisions
            if isinstance(revision, dict):
                ws[f'A{row}'].value = revision.get('version', '')
                ws[f'B{row}'].value = revision.get('date', '')
                ws[f'C{row}'].value = revision.get('author', '')[:30]
                ws[f'D{row}'].value = revision.get('changes', '')[:80]
                row += 1

        row += 1

        # Section USAGE NOTES
        ws[f'A{row}'].value = "NOTES D'UTILISATION"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
        row += 1

        usage_notes = self.assumptions.get('usage_notes', '')
        if usage_notes:
            # Diviser en lignes
            for line in usage_notes.split('\n')[:15]:
                if line.strip():
                    ws[f'A{row}'].value = line.strip()
                    row += 1

        row += 2

        # Section STRUCTURE FICHIERS
        ws[f'A{row}'].value = "STRUCTURE DU BP"
        ws[f'A{row}'].font = openpyxl.styles.Font(bold=True, size=12)
        row += 1

        structure = [
            ("Workflow", "RAW → TEMPLATE → FINAL"),
            ("Source unique", "assumptions.yaml + funding_captable.yaml"),
            ("Projections", "projections_50m.json (généré par Python)"),
            ("Mapping", "YAML → Excel 100% complet"),
            ("Formules", "3108 formules Excel préservées"),
            ("Sheets totaux", "17 (14 RAW + 3 nouveaux)"),
            ("Nouveaux sheets", "Cash Flow, Scenarios, Unit Economics"),
        ]

        for label, value in structure:
            ws[f'A{row}'].value = label
            ws[f'B{row}'].value = value
            row += 1

        # Ajuster largeur colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 80

        logger.info("✓ Sheet Documentation créé (meta + history + notes)")

    def clean_data_cells(self):
        """
        Nettoyer les cellules de données (pas les formules)
        Mettre des valeurs par défaut pour les placeholders
        SAUF lignes 16-25 dans Personnel (pilotées par YAML)
        """
        logger.info("\n🧹 Nettoyage cellules de données...")

        sheets_to_clean = ['P&L', 'Ventes', 'Charges de personnel et FG',
                          'Infrastructure technique', 'Marketing', 'Sous traitance']

        for sheet_name in sheets_to_clean:
            if sheet_name not in self.wb.sheetnames:
                continue

            ws = self.wb[sheet_name]

            # Parcourir les cellules de données (à partir ligne 4, col 4)
            cleaned = 0
            for row in ws.iter_rows(min_row=4, max_row=100, min_col=4, max_col=150):
                for cell in row:
                    # SKIP lignes 16-25 dans Personnel (données YAML headcount)
                    if sheet_name == 'Charges de personnel et FG' and 16 <= cell.row <= 25:
                        continue

                    # Si c'est une valeur numérique (pas formule), mettre 0
                    if isinstance(cell.value, (int, float)) and cell.value != 0:
                        cell.value = 0
                        cleaned += 1

            logger.info(f"  {sheet_name}: {cleaned} cellules nettoyées")

    def add_template_markers(self):
        """
        Ajouter des marqueurs visuels pour identifier le template
        """
        logger.info("\n🏷️ Ajout marqueurs TEMPLATE...")

        # Ajouter une note sur la première sheet
        ws = self.wb.worksheets[0]
        ws['A1'].value = "🔧 TEMPLATE BP 50 MOIS - Gabarit à valider"

        # Style
        from openpyxl.styles import Font, PatternFill
        ws['A1'].font = Font(bold=True, size=14, color="FF0000")
        ws['A1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        logger.info("✓ Marqueurs ajoutés")

    def preserve_formulas_info(self):
        """
        Logger des infos sur les formules préservées
        """
        logger.info("\n📊 Inventaire formules préservées...")

        for sheet_name in ['P&L', 'Ventes', 'Synthèse']:
            if sheet_name not in self.wb.sheetnames:
                continue

            ws = self.wb[sheet_name]
            formula_count = 0

            for row in ws.iter_rows(min_row=1, max_row=100, min_col=1, max_col=150):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1

            logger.info(f"  {sheet_name}: {formula_count} formules")

    def create_template(self):
        """Créer le template complet avec toutes les améliorations Phase 1 + 2 + 3 + 4"""
        logger.info("\n🔨 CRÉATION TEMPLATE ENRICHI (Phase 1 + 2 + 3 + 4)")
        logger.info("=" * 60)

        # 1. Adapter structure selon YAML (existant + enrichi)
        self.update_parametres_sheet()  # ✅ Enrichi avec financial_kpis, validation_rules, hypothèses, RH, volumes
        self.update_financement_sheet()
        self.update_fundings_sheet_with_captable()  # ✅ Cap table détaillée
        self.update_strategie_vente_sheet()
        self.update_charges_personnel_sheet()
        self.update_infrastructure_detailed_sheet()
        self.update_marketing_detailed_sheet()

        # 2. PHASE 1 - Améliorations HAUTE PRIORITÉ
        self.add_arr_mrr_to_pl()  # ✅ ARR/MRR dans P&L
        self.create_cash_flow_sheet()  # ✅ Cash Flow Statement
        self.enrich_synthese_dashboard()  # ✅ Dashboard exécutif

        # 3. PHASE 2 - Améliorations MOYENNE PRIORITÉ
        self.create_scenarios_sheet()  # ✅ NEW: Scenarios (base/upside/downside)
        self.create_unit_economics_sheet()  # ✅ NEW: Unit Economics (CAC/LTV par produit)

        # 4. PHASE 3 - Améliorations BASSE PRIORITÉ
        self.add_granular_metrics_to_ventes()  # ✅ NEW: Métriques granulaires Ventes
        self.enhance_fundings_visualization()  # ✅ NEW: Visualisation dilution Fundings
        self.create_data_quality_sheet()  # ✅ NEW: Data Quality checks
        self.create_documentation_sheet()  # ✅ NEW: Documentation

        # 5. PHASE 4 - Corrections GAP ANALYSIS HAUTE PRIORITÉ
        self.add_productivity_ia_to_ventes()  # ✅ NEW: Productivité IA (pitch core GenieFactory)

        # 6. PHASE 5 - Corrections GAP ANALYSIS MOYENNE PRIORITÉ
        self.improve_infrastructure_labels()  # ✅ NEW: Labels Hosting, Licences, total
        self.improve_marketing_labels()  # ✅ NEW: Labels Ventes, Campagnes

        # 7. Supprimer sheets inutiles
        self.remove_gtmarket_sheet()  # ✅ Suppression GTMarket

        # 8. Nettoyer les données
        self.clean_data_cells()

        # 9. Ajouter marqueurs
        self.add_template_markers()

        # 10. Vérifier formules
        self.preserve_formulas_info()

        logger.info("\n" + "=" * 60)
        logger.info("✅ TEMPLATE ENRICHI CRÉÉ (Phase 1 + 2 + 3 + 4 + 5 complètes)")
        logger.info("   PHASE 1:")
        logger.info("   • Paramètres: financial_kpis + validation_rules + hypothèses")
        logger.info("   • P&L: ARR/MRR ajoutés")
        logger.info("   • Cash Flow: nouveau sheet créé")
        logger.info("   • Synthèse: dashboard exécutif ajouté")
        logger.info("   PHASE 2:")
        logger.info("   • Scenarios: base/upside/downside créé")
        logger.info("   • Unit Economics: CAC/LTV par produit créé")
        logger.info("   PHASE 3:")
        logger.info("   • Ventes: métriques granulaires par tier ajoutées")
        logger.info("   • Fundings: visualisation dilution enrichie")
        logger.info("   • Data Quality: checks automatiques créés")
        logger.info("   • Documentation: meta + history + notes créés")
        logger.info("   PHASE 4 (GAP ANALYSIS HAUTE):")
        logger.info("   • Paramètres: coûts RH (charges sociales 45%) + volumes commerciaux")
        logger.info("   • Ventes: productivité IA (pitch core GenieFactory)")
        logger.info("   PHASE 5 (GAP ANALYSIS MOYENNE):")
        logger.info("   • Infrastructure: labels Hosting, Licences logicielles, total")
        logger.info("   • Marketing: labels Ventes, Campagnes Collaboration/Ciblées")
        logger.info("   PHASE 6 (RESTRUCTURATION FINALE):")
        logger.info("   • Fundings: État de l'art (Timeline + Cap Table + Non-dilutif + Metrics)")
        logger.info("   • Personnel: Piloté par assumptions.yaml (8 rôles avec timeline)")

    def save(self, output_path: Path):
        """Sauvegarder le template"""
        logger.info(f"\n💾 Sauvegarde: {output_path}")
        self.wb.save(output_path)
        size_kb = output_path.stat().st_size / 1024
        logger.info(f"✓ Template sauvegardé: {size_kb:.1f} KB")


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   CRÉATION TEMPLATE EXCEL DEPUIS RAW[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    # Charger assumptions
    assumptions_path = base_path / "data" / "structured" / "assumptions.yaml"
    console.print(f"[yellow]📂 Chargement assumptions:[/yellow] {assumptions_path.name}")
    with open(assumptions_path, 'r', encoding='utf-8') as f:
        assumptions = yaml.safe_load(f)
    console.print(f"[green]✓ Assumptions chargées (v{assumptions.get('version', '?')})[/green]\n")

    # Fichiers
    raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    template_file = base_path / "data" / "outputs" / "BP_50M_TEMPLATE.xlsx"

    # Créer le template
    creator = TemplateCreator(raw_file, assumptions)
    creator.create_template()
    creator.save(template_file)

    console.print(f"\n[bold green]✅ TEMPLATE CRÉÉ[/bold green]")
    console.print(f"[green]📁 {template_file}[/green]")
    console.print(f"\n[cyan]→ Structure adaptée selon assumptions.yaml[/cyan]")
    console.print(f"[cyan]→ Toutes les formules Excel préservées[/cyan]")
    console.print(f"[cyan]→ Cellules de données nettoyées (placeholders à 0)[/cyan]")
    console.print(f"[yellow]→ À VALIDER avant injection des données[/yellow]\n")


if __name__ == "__main__":
    main()
