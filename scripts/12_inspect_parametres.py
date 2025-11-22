#!/usr/bin/env python3
"""
Inspection détaillée du sheet Paramètres
Pour identifier exactement ce qui est présent et ce qui manque
"""

import openpyxl
from pathlib import Path
from rich.console import Console
from rich.table import Table
from rich import box

console = Console()


def inspect_parametres():
    base_path = Path(__file__).parent.parent
    final_file = base_path / "data" / "outputs" / "BP_50M_FINAL_Nov2025-Dec2029.xlsx"

    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   INSPECTION DÉTAILLÉE: Sheet Paramètres[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    wb = openpyxl.load_workbook(final_file, data_only=False)
    ws = wb['Paramètres']

    console.print(f"Dimensions: {ws.max_row} lignes × {ws.max_column} colonnes\n")

    # Extraire tout le contenu par colonnes
    console.print("[bold yellow]COLONNE A-B: Prix Produits[/bold yellow]")
    for row in range(1, 30):
        col_a = ws[f'A{row}'].value
        col_b = ws[f'B{row}'].value
        if col_a or col_b:
            console.print(f"  {row}: {col_a or '':<40} {col_b or ''}")

    console.print("\n[bold yellow]COLONNE C-D: ?[/bold yellow]")
    for row in range(1, 30):
        col_c = ws[f'C{row}'].value
        col_d = ws[f'D{row}'].value
        if col_c or col_d:
            console.print(f"  {row}: {col_c or '':<40} {col_d or ''}")

    console.print("\n[bold yellow]COLONNE E-F: ?[/bold yellow]")
    for row in range(1, 30):
        col_e = ws[f'E{row}'].value
        col_f = ws[f'F{row}'].value
        if col_e or col_f:
            console.print(f"  {row}: {col_e or '':<40} {col_f or ''}")

    console.print("\n[bold yellow]COLONNE H-I: Financial KPIs[/bold yellow]")
    for row in range(1, 30):
        col_h = ws[f'H{row}'].value
        col_i = ws[f'I{row}'].value
        if col_h or col_i:
            console.print(f"  {row}: {col_h or '':<40} {col_i or ''}")

    console.print("\n[bold yellow]COLONNE K-M: Validation Rules[/bold yellow]")
    for row in range(1, 30):
        col_k = ws[f'K{row}'].value
        col_l = ws[f'L{row}'].value
        col_m = ws[f'M{row}'].value
        if col_k or col_l or col_m:
            console.print(f"  {row}: {col_k or '':<30} {col_l or '':<15} {col_m or ''}")

    console.print("\n[bold yellow]COLONNE O-P: Hypothèses Business[/bold yellow]")
    for row in range(1, 30):
        col_o = ws[f'O{row}'].value
        col_p = ws[f'P{row}'].value
        if col_o or col_p:
            console.print(f"  {row}: {col_o or '':<40} {col_p or ''}")

    # Analyse des manques
    console.print("\n[bold red]═══ ÉLÉMENTS MANQUANTS IDENTIFIÉS ═══[/bold red]\n")

    manques = []

    # Vérifier taux conversion Factory
    factory_conv_found = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                if 'conversion' in cell.value.lower() and 'factory' in cell.value.lower():
                    factory_conv_found = True
                    break

    if not factory_conv_found:
        manques.append("❌ Taux de conversion Hackathon→Factory (35%)")

    # Vérifier taux charges sociales
    charges_found = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                if 'charge' in cell.value.lower() and 'social' in cell.value.lower():
                    charges_found = True
                    break

    if not charges_found:
        manques.append("❌ Taux charges sociales (45%)")

    # Vérifier prix Hub par tier
    hub_tiers_found = {'starter': False, 'business': False, 'enterprise': False}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                cell_lower = cell.value.lower()
                if 'hub' in cell_lower and 'starter' in cell_lower:
                    hub_tiers_found['starter'] = True
                if 'hub' in cell_lower and 'business' in cell_lower:
                    hub_tiers_found['business'] = True
                if 'hub' in cell_lower and 'enterprise' in cell_lower:
                    hub_tiers_found['enterprise'] = True

    for tier, found in hub_tiers_found.items():
        if not found:
            manques.append(f"⚠️ Prix Hub {tier.capitalize()} pas explicitement affiché")

    # Vérifier volumes hackathons
    volumes_found = False
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            if cell.value and isinstance(cell.value, str):
                if 'volume' in cell.value.lower() and 'hackathon' in cell.value.lower():
                    volumes_found = True
                    break

    if not volumes_found:
        manques.append("⚠️ Volumes hackathons mensuels (2-12 par mois)")

    for manque in manques:
        console.print(f"  {manque}")

    if not manques:
        console.print("[green]  ✅ Aucun manque critique identifié![/green]")

    console.print("\n")


if __name__ == '__main__':
    inspect_parametres()
