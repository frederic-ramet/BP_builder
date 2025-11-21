#!/usr/bin/env python3
"""
Valider le fichier Excel adapté
"""

import openpyxl
from pathlib import Path
import json
from rich.console import Console
from rich.table import Table
from rich import box

console = Console()


def validate_formulas_preserved(wb_source, wb_adapted):
    """Vérifier que les formules sont préservées"""
    console.print("\n[cyan]🔍 Validation: Formules Excel préservées[/cyan]")

    results = []

    for sheet_name in ['P&L', 'Ventes', 'Synthèse']:
        if sheet_name not in wb_source.sheetnames or sheet_name not in wb_adapted.sheetnames:
            continue

        ws_source = wb_source[sheet_name]
        ws_adapted = wb_adapted[sheet_name]

        # Compter les formules
        formulas_source = 0
        formulas_adapted = 0

        for row in ws_source.iter_rows(min_row=1, max_row=50, min_col=1, max_col=50):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas_source += 1

        for row in ws_adapted.iter_rows(min_row=1, max_row=50, min_col=1, max_col=50):
            for cell in row:
                val = cell.value
                if val and not isinstance(val, (str, int, float, bool, type(None))):
                    val = str(val)
                if isinstance(val, str) and val.startswith('='):
                    formulas_adapted += 1

        results.append({
            'sheet': sheet_name,
            'source': formulas_source,
            'adapted': formulas_adapted,
            'preserved': formulas_source == formulas_adapted
        })

    table = Table(box=box.ROUNDED)
    table.add_column("Sheet", style="cyan")
    table.add_column("Formules Source", justify="right")
    table.add_column("Formules Adaptées", justify="right")
    table.add_column("Status", justify="center")

    for r in results:
        status = "✅" if r['preserved'] else "❌"
        table.add_row(r['sheet'], str(r['source']), str(r['adapted']), status)

    console.print(table)

    all_preserved = all(r['preserved'] for r in results)
    if all_preserved:
        console.print("[green]✅ Toutes les formules préservées![/green]")
    else:
        console.print("[red]⚠️ Certaines formules ont été perdues[/red]")

    return all_preserved


def validate_data_injected(wb_adapted, projections):
    """Vérifier que les données Python ont été injectées"""
    console.print("\n[cyan]🔍 Validation: Données Python injectées[/cyan]")

    ws = wb_adapted['P&L']

    # Vérifier quelques valeurs clés
    checks = []

    # M1 (Col F): CA Total
    m1_ca_expected = projections[0]['revenue']['total']
    m1_ca_excel = ws['F2'].value
    if m1_ca_excel and not isinstance(m1_ca_excel, str):
        checks.append({
            'metric': 'M1 CA Total',
            'expected': m1_ca_expected,
            'actual': m1_ca_excel,
            'match': abs(m1_ca_expected - m1_ca_excel) < 1
        })

    # M14 (Col S): CA Total
    m14_ca_expected = projections[13]['revenue']['total']
    m14_ca_excel = ws['S2'].value
    if m14_ca_excel and not isinstance(m14_ca_excel, str):
        checks.append({
            'metric': 'M14 CA Total',
            'expected': m14_ca_expected,
            'actual': m14_ca_excel,
            'match': abs(m14_ca_expected - m14_ca_excel) < 1
        })

    # M50 (Col BC): CA Total
    m50_ca_expected = projections[49]['revenue']['total']
    m50_ca_excel = ws['BC2'].value
    if m50_ca_excel and not isinstance(m50_ca_excel, str):
        checks.append({
            'metric': 'M50 CA Total',
            'expected': m50_ca_expected,
            'actual': m50_ca_excel,
            'match': abs(m50_ca_expected - m50_ca_excel) < 1
        })

    table = Table(box=box.ROUNDED)
    table.add_column("Métrique", style="cyan")
    table.add_column("Attendu", justify="right")
    table.add_column("Excel", justify="right")
    table.add_column("Status", justify="center")

    for c in checks:
        status = "✅" if c['match'] else "❌"
        table.add_row(
            c['metric'],
            f"{c['expected']:,.0f}€",
            f"{c['actual']:,.0f}€",
            status
        )

    console.print(table)

    all_match = all(c['match'] for c in checks)
    if all_match:
        console.print("[green]✅ Données Python correctement injectées![/green]")
    else:
        console.print("[red]⚠️ Certaines données ne correspondent pas[/red]")

    return all_match


def validate_structure(wb_adapted):
    """Vérifier la structure du fichier"""
    console.print("\n[cyan]🔍 Validation: Structure du fichier[/cyan]")

    expected_sheets = [
        'Synthèse',
        'Stratégie de vente',
        'Financement',
        'P&L',
        'Paramètres',
        'GTMarket',
        'Ventes',
        'Sous traitance',
        'Charges de personnel et FG',
        'DIRECTION',
        'Infrastructure technique',
        'Fundings',
        '>>',
        'Positionnement',
        'Marketing'
    ]

    actual_sheets = wb_adapted.sheetnames

    table = Table(box=box.SIMPLE)
    table.add_column("Sheet", style="cyan")
    table.add_column("Present", justify="center")

    for sheet in expected_sheets:
        present = "✅" if sheet in actual_sheets else "❌"
        table.add_row(sheet, present)

    console.print(table)

    all_present = all(s in actual_sheets for s in expected_sheets)
    if all_present:
        console.print(f"[green]✅ Tous les 15 sheets présents![/green]")
    else:
        console.print(f"[red]⚠️ Certains sheets manquants[/red]")

    return all_present


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   VALIDATION FICHIER EXCEL ADAPTÉ[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")

    base_path = Path(__file__).parent.parent

    # Charger les fichiers
    source_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    adapted_file = base_path / "data" / "outputs" / "BP_50M_Adapted_Nov2025-Dec2029.xlsx"
    projections_file = base_path / "data" / "structured" / "projections_50m.json"

    console.print(f"\n[yellow]📂 Chargement fichiers...[/yellow]")
    wb_source = openpyxl.load_workbook(source_file, data_only=False)
    wb_adapted = openpyxl.load_workbook(adapted_file, data_only=False)

    with open(projections_file) as f:
        projections = json.load(f)

    console.print(f"[green]✓ Source: {len(wb_source.sheetnames)} sheets[/green]")
    console.print(f"[green]✓ Adapté: {len(wb_adapted.sheetnames)} sheets[/green]")
    console.print(f"[green]✓ Projections: {len(projections)} mois[/green]")

    # Validations
    v1 = validate_structure(wb_adapted)
    v2 = validate_formulas_preserved(wb_source, wb_adapted)
    v3 = validate_data_injected(wb_adapted, projections)

    # Résumé
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold]RÉSUMÉ VALIDATION:[/bold]")
    console.print(f"  Structure complète: {'✅' if v1 else '❌'}")
    console.print(f"  Formules préservées: {'✅' if v2 else '❌'}")
    console.print(f"  Données injectées: {'✅' if v3 else '❌'}")

    if v1 and v2 and v3:
        console.print("\n[bold green]✅ FICHIER EXCEL ADAPTÉ VALIDE![/bold green]")
        console.print("[green]→ Prêt pour utilisation[/green]")
        console.print("[green]→ Ouvrir dans Excel pour voir formules recalculer[/green]\n")
    else:
        console.print("\n[bold red]⚠️ VALIDATION INCOMPLÈTE[/bold red]")
        console.print("[red]→ Vérifier les erreurs ci-dessus[/red]\n")


if __name__ == "__main__":
    main()
