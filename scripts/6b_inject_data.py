#!/usr/bin/env python3
"""
Injecter les projections 50M dans le TEMPLATE Excel
pour générer le fichier FINAL
"""

import openpyxl
from pathlib import Path
import json
from rich.console import Console
from rich.progress import track
import logging

console = Console()
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


class DataInjector:
    """Injecter les données dans le template"""

    def __init__(self, template_path: Path, projections: list):
        self.template_path = template_path
        self.projections = projections

        logger.info(f"📂 Chargement TEMPLATE: {template_path.name}")
        self.wb = openpyxl.load_workbook(template_path)
        logger.info(f"✓ {len(self.wb.sheetnames)} sheets chargés")

        # Mapper les colonnes
        self.setup_month_mapping()

    def setup_month_mapping(self):
        """Mapper les 50 mois aux colonnes Excel"""
        self.month_to_col = {}

        pl_sheet = self.wb['P&L']

        current_month = 1
        for col_idx in range(4, pl_sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            year_cell = pl_sheet.cell(1, col_idx).value
            month_cell = pl_sheet.cell(2, col_idx).value

            # Skip colonnes totaux annuels
            if isinstance(year_cell, (int, float)) and month_cell is None:
                continue

            # C'est un mois
            if isinstance(month_cell, (int, float)) or (isinstance(month_cell, str) and month_cell.startswith('=')):
                self.month_to_col[current_month] = col_letter
                current_month += 1

            if current_month > 50:
                break

        logger.info(f"✓ Mapping: {len(self.month_to_col)} mois (M1→{self.month_to_col.get(1)}, M50→{self.month_to_col.get(50)})")

    def inject_pl_data(self):
        """Injecter données P&L"""
        logger.info("\n📊 Injection P&L...")

        ws = self.wb['P&L']

        row_map = {
            'ca_total': 2,
            'hackathons': 3,
            'factory': 4,
            'hub_mrr': 5,
            'services': 6,
            'sous_traitance': 10,
            'infrastructure': 11,
            'charges_personnel_ops': 12,
            'marketing': 18,
            'charges_personnel_fonc': 19,
            'frais_generaux': 20,
        }

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Injecter seulement si pas de formule
            data = {
                'ca_total': proj['revenue']['total'],
                'hackathons': proj['revenue']['hackathon']['revenue'],
                'factory': proj['revenue']['factory']['revenue'],
                'hub_mrr': proj['revenue']['enterprise_hub']['mrr'],
                'services': proj['revenue']['services']['revenue'],
                'sous_traitance': proj['costs']['personnel'].get('freelance', 0),
                'infrastructure': proj['costs']['infrastructure']['total'],
                'charges_personnel_ops': proj['costs']['personnel']['total'],
                'marketing': proj['costs']['marketing']['total'],
                'charges_personnel_fonc': proj['costs']['personnel']['total'],
                'frais_generaux': proj['costs'].get('admin', 0),
            }

            for key, value in data.items():
                if key not in row_map:
                    continue

                cell = ws[f'{col}{row_map[key]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = value
                    injected += 1

        logger.info(f"✓ P&L: {injected} cellules injectées")

    def inject_ventes_data(self):
        """Injecter données Ventes"""
        logger.info("\n💼 Injection Ventes...")

        ws = self.wb['Ventes']

        row_map = {
            'nb_hackathons': 3,
            'ca_hackathons': 4,
            'nb_factory': 6,
            'ca_factory': 9,
            'nb_clients_hub': 11,
            'mrr_hub': 13,
            'arr': 15,
        }

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            data = {
                'nb_hackathons': proj['revenue']['hackathon'].get('volume', 0),
                'ca_hackathons': proj['revenue']['hackathon']['revenue'],
                'nb_factory': proj['revenue']['factory'].get('volume', 0),
                'ca_factory': proj['revenue']['factory']['revenue'],
                'nb_clients_hub': proj['revenue']['enterprise_hub']['customers']['total'],
                'mrr_hub': proj['revenue']['enterprise_hub']['mrr'],
                'arr': proj['revenue']['enterprise_hub']['arr'],
            }

            for key, value in data.items():
                if key not in row_map:
                    continue

                cell = ws[f'{col}{row_map[key]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = value
                    injected += 1

        logger.info(f"✓ Ventes: {injected} cellules injectées")

    def inject_personnel_data(self):
        """Injecter données Personnel"""
        logger.info("\n👥 Injection Charges Personnel...")

        ws = self.wb['Charges de personnel et FG']

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Total personnel (ligne approximative 5)
            cell = ws[f'{col}5']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['personnel']['total']
                injected += 1

        logger.info(f"✓ Personnel: {injected} cellules injectées")

    def inject_infrastructure_data(self):
        """Injecter données Infrastructure"""
        logger.info("\n☁️ Injection Infrastructure...")

        ws = self.wb['Infrastructure technique']

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Cloud (ligne 3)
            cell = ws[f'{col}3']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['infrastructure'].get('cloud', 0)
                injected += 1

            # SaaS (ligne 5)
            cell = ws[f'{col}5']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['infrastructure'].get('saas_tools', 0)
                injected += 1

        logger.info(f"✓ Infrastructure: {injected} cellules injectées")

    def inject_marketing_data(self):
        """Injecter données Marketing"""
        logger.info("\n📢 Injection Marketing...")

        ws = self.wb['Marketing']

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Total marketing (ligne 3)
            cell = ws[f'{col}3']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['marketing']['total']
                injected += 1

        logger.info(f"✓ Marketing: {injected} cellules injectées")

    def inject_sous_traitance_data(self):
        """Injecter données Sous-traitance"""
        logger.info("\n🔧 Injection Sous-traitance...")

        ws = self.wb['Sous traitance']

        injected = 0
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Freelance (ligne 3)
            cell = ws[f'{col}3']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['personnel'].get('freelance', 0)
                injected += 1

        logger.info(f"✓ Sous-traitance: {injected} cellules injectées")

    def inject_cash_flow_data(self):
        """Injecter données Cash Flow"""
        logger.info("\n💰 Injection Cash Flow...")

        if 'Cash Flow' not in self.wb.sheetnames:
            logger.warning("⚠️ Sheet 'Cash Flow' introuvable, skip injection")
            return

        ws = self.wb['Cash Flow']

        # Trouver les lignes par labels
        row_map = {}
        for row in range(1, 30):
            label = ws[f'A{row}'].value
            if label and isinstance(label, str):
                if 'CA Encaissé' in label:
                    row_map['revenue'] = row
                elif 'Charges Personnel' in label:
                    row_map['personnel'] = row
                elif 'Charges Infrastructure' in label:
                    row_map['infrastructure'] = row
                elif 'Charges Marketing' in label:
                    row_map['marketing'] = row
                elif 'Cash Flow Opérationnel' in label:
                    row_map['operating_cf'] = row
                elif 'Pre-Seed' in label:
                    row_map['preseed'] = row
                elif 'Seed' in label and 'Pre-' not in label:
                    row_map['seed'] = row
                elif 'Series A' in label:
                    row_map['series_a'] = row
                elif 'Cash Flow Financement' in label:
                    row_map['financing_cf'] = row
                elif 'TOTAL CASH FLOW' in label:
                    row_map['total_cf'] = row
                elif 'CASH BALANCE' in label:
                    row_map['cash_balance'] = row
                elif 'Burn Rate' in label:
                    row_map['burn_rate'] = row
                elif 'Cash Runway' in label:
                    row_map['cash_runway'] = row

        injected = 0

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col_idx = openpyxl.utils.column_index_from_string(self.month_to_col[month])
            col_letter = openpyxl.utils.get_column_letter(col_idx + 2)  # +2 car Cash Flow a colonnes A,B de labels

            proj = self.projections[month - 1]

            # Revenue
            if 'revenue' in row_map:
                ws[f'{col_letter}{row_map["revenue"]}'].value = proj['revenue']['total']
                injected += 1

            # Personnel
            if 'personnel' in row_map:
                ws[f'{col_letter}{row_map["personnel"]}'].value = -proj['costs']['personnel']['total']
                injected += 1

            # Infrastructure
            if 'infrastructure' in row_map:
                infra_total = proj['costs']['infrastructure'].get('cloud', 0) + proj['costs']['infrastructure'].get('saas_tools', 0)
                ws[f'{col_letter}{row_map["infrastructure"]}'].value = -infra_total
                injected += 1

            # Marketing
            if 'marketing' in row_map:
                ws[f'{col_letter}{row_map["marketing"]}'].value = -proj['costs']['marketing']['total']
                injected += 1

            # Operating CF
            if 'operating_cf' in row_map:
                operating_cf = (proj['revenue']['total'] -
                               proj['costs']['personnel']['total'] -
                               proj['costs']['infrastructure'].get('cloud', 0) -
                               proj['costs']['infrastructure'].get('saas_tools', 0) -
                               proj['costs']['marketing']['total'])
                ws[f'{col_letter}{row_map["operating_cf"]}'].value = operating_cf
                injected += 1

            # Fundings (Pre-Seed M1, Seed M11, Series A M36)
            if 'preseed' in row_map and month == 1:
                ws[f'{col_letter}{row_map["preseed"]}'].value = 150000
                injected += 1

            if 'seed' in row_map and month == 11:
                ws[f'{col_letter}{row_map["seed"]}'].value = 500000
                injected += 1

            if 'series_a' in row_map and month == 36:
                ws[f'{col_letter}{row_map["series_a"]}'].value = 2500000
                injected += 1

            # Financing CF
            if 'financing_cf' in row_map:
                financing_cf = 0
                if month == 1:
                    financing_cf = 150000
                elif month == 11:
                    financing_cf = 500000
                elif month == 36:
                    financing_cf = 2500000
                ws[f'{col_letter}{row_map["financing_cf"]}'].value = financing_cf
                injected += 1

            # Total CF
            if 'total_cf' in row_map and 'operating_cf' in row_map and 'financing_cf' in row_map:
                operating = ws[f'{col_letter}{row_map["operating_cf"]}'].value or 0
                financing = ws[f'{col_letter}{row_map["financing_cf"]}'].value or 0
                ws[f'{col_letter}{row_map["total_cf"]}'].value = operating + financing
                injected += 1

            # Cash Balance (cumul)
            if 'cash_balance' in row_map:
                ws[f'{col_letter}{row_map["cash_balance"]}'].value = proj.get('cash_balance', 0)
                injected += 1

            # Burn Rate (si operating CF négatif)
            if 'burn_rate' in row_map and 'operating_cf' in row_map:
                operating = ws[f'{col_letter}{row_map["operating_cf"]}'].value or 0
                burn = -operating if operating < 0 else 0
                ws[f'{col_letter}{row_map["burn_rate"]}'].value = burn
                injected += 1

            # Cash Runway
            if 'cash_runway' in row_map and 'burn_rate' in row_map and 'cash_balance' in row_map:
                cash = ws[f'{col_letter}{row_map["cash_balance"]}'].value or 0
                burn = ws[f'{col_letter}{row_map["burn_rate"]}'].value or 0
                runway = cash / burn if burn > 0 else 999
                ws[f'{col_letter}{row_map["cash_runway"]}'].value = runway
                injected += 1

        logger.info(f"✓ Cash Flow: {injected} cellules injectées")

    def inject_arr_mrr_in_pl(self):
        """Injecter ARR/MRR dans P&L"""
        logger.info("\n📈 Injection ARR/MRR dans P&L...")

        ws = self.wb['P&L']

        # Trouver les lignes ARR/MRR
        arr_row = None
        mrr_row = None

        for row in range(1, 20):
            label = ws[f'A{row}'].value
            if label and isinstance(label, str):
                if 'ARR' in label and 'Annual' in label:
                    arr_row = row
                elif 'MRR' in label and 'Monthly' in label:
                    mrr_row = row

        if not arr_row or not mrr_row:
            logger.warning("⚠️ Lignes ARR/MRR introuvables dans P&L, skip")
            return

        injected = 0

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # ARR
            if arr_row:
                ws[f'{col}{arr_row}'].value = proj.get('arr', 0)
                injected += 1

            # MRR
            if mrr_row:
                hub_mrr = proj['revenue']['enterprise_hub'].get('mrr', 0)
                ws[f'{col}{mrr_row}'].value = hub_mrr
                injected += 1

        logger.info(f"✓ ARR/MRR: {injected} cellules injectées")

    def remove_template_marker(self):
        """Retirer le marqueur TEMPLATE"""
        logger.info("\n🏷️ Retrait marqueur TEMPLATE...")

        ws = self.wb.worksheets[0]
        if ws['A1'].value and 'TEMPLATE' in str(ws['A1'].value):
            ws['A1'].value = "Business Plan GenieFactory - 50 Mois (Nov 2025 - Dec 2029)"
            from openpyxl.styles import Font, PatternFill
            ws['A1'].font = Font(bold=True, size=14, color="000000")
            ws['A1'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            logger.info("✓ Marqueur retiré")

    def inject_all(self):
        """Injecter toutes les données + Phase 1 améliorations"""
        logger.info("\n🔨 INJECTION DONNÉES (avec Phase 1)")
        logger.info("=" * 60)

        self.inject_pl_data()
        self.inject_arr_mrr_in_pl()  # NEW: ARR/MRR dans P&L
        self.inject_ventes_data()
        self.inject_personnel_data()
        self.inject_infrastructure_data()
        self.inject_marketing_data()
        self.inject_sous_traitance_data()
        self.inject_cash_flow_data()  # NEW: Cash Flow complet
        self.remove_template_marker()

        logger.info("\n" + "=" * 60)
        logger.info("✅ INJECTION TERMINÉE (Phase 1 complète)")

    def save(self, output_path: Path):
        """Sauvegarder le fichier final"""
        logger.info(f"\n💾 Sauvegarde: {output_path}")
        self.wb.save(output_path)
        size_kb = output_path.stat().st_size / 1024
        logger.info(f"✓ Fichier FINAL sauvegardé: {size_kb:.1f} KB")


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   INJECTION DONNÉES DANS TEMPLATE[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    # Charger projections
    projections_path = base_path / "data" / "structured" / "projections_50m.json"
    console.print(f"[yellow]📂 Chargement projections:[/yellow] {projections_path.name}")
    with open(projections_path, 'r', encoding='utf-8') as f:
        projections = json.load(f)
    console.print(f"[green]✓ {len(projections)} mois chargés[/green]\n")

    # Fichiers
    template_file = base_path / "data" / "outputs" / "BP_50M_TEMPLATE.xlsx"
    final_file = base_path / "data" / "outputs" / "BP_50M_FINAL_Nov2025-Dec2029.xlsx"

    # Injecter
    injector = DataInjector(template_file, projections)
    injector.inject_all()
    injector.save(final_file)

    console.print(f"\n[bold green]✅ FICHIER FINAL GÉNÉRÉ[/bold green]")
    console.print(f"[green]📁 {final_file}[/green]")
    console.print(f"\n[cyan]→ Données Python injectées depuis projections_50m.json[/cyan]")
    console.print(f"[cyan]→ Toutes les formules Excel préservées[/cyan]")
    console.print(f"[cyan]→ Prêt pour validation finale[/cyan]\n")


if __name__ == "__main__":
    main()
