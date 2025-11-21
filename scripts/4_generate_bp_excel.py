#!/usr/bin/env python3
"""
GenieFactory BP 14 Mois - Script 4: Génération BP Excel
Crée BP_14M_Nov2025-Dec2026.xlsx avec formules Excel actives

Input:
  - data/structured/projections.json
  - data/structured/assumptions.yaml

Output:
  - data/outputs/BP_14M_Nov2025-Dec2026.xlsx (8 sheets avec formules)
"""

import json
import yaml
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# Configuration logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class BPExcelGenerator:
    """Générateur BP Excel avec formules"""

    def __init__(self, projections: List[Dict], assumptions: Dict):
        self.projections = projections
        self.assumptions = assumptions
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Supprimer sheet par défaut

        # Colonnes pour M1-M14 (F à S)
        self.month_cols = [get_column_letter(6 + i) for i in range(14)]  # F-S

    def create_styles(self):
        """Définir les styles réutilisables"""
        self.style_header = {
            'font': Font(bold=True, size=11, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

        self.style_total = {
            'font': Font(bold=True, size=10),
            'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'alignment': Alignment(horizontal='right')
        }

        self.style_currency = {
            'number_format': '#,##0 €',
            'alignment': Alignment(horizontal='right')
        }

        self.style_arr = {
            'font': Font(bold=True, color='00B050'),
            'number_format': '#,##0 €'
        }

    def apply_style(self, cell, style_dict):
        """Appliquer un style à une cellule"""
        for key, value in style_dict.items():
            setattr(cell, key, value)

    def create_pl_sheet(self):
        """Créer sheet P&L avec formules"""
        logger.info("📊 Création sheet P&L...")

        ws = self.wb.create_sheet("P&L")

        # Headers
        ws['A1'] = "Compte de Résultat"
        ws['A1'].font = Font(bold=True, size=14)

        # Ligne headers mois
        ws['E2'] = "Période"
        self.apply_style(ws['E2'], self.style_header)

        for idx, col in enumerate(self.month_cols):
            month_num = idx + 1
            ws[f'{col}2'] = f"M{month_num}"
            self.apply_style(ws[f'{col}2'], self.style_header)

        # REVENUS
        row = 3
        ws[f'A{row}'] = "CHIFFRE D'AFFAIRES"
        ws[f'A{row}'].font = Font(bold=True, size=11)

        # CA Total (formule SUM)
        row += 1
        ws[f'E{row}'] = "CA TOTAL"
        self.apply_style(ws[f'E{row}'], self.style_total)

        for idx, col in enumerate(self.month_cols):
            # Formule: SUM des lignes revenus (row+1 à row+4)
            ws[f'{col}{row}'] = f'=SUM({col}{row+1}:{col}{row+4})'
            self.apply_style(ws[f'{col}{row}'], {**self.style_currency, **self.style_total})

        # Détail revenus
        revenue_labels = [
            ('Hackathon', 'hackathon'),
            ('Factory', 'factory'),
            ('Enterprise Hub (MRR)', 'enterprise_hub'),
            ('Services', 'services')
        ]

        for label, key in revenue_labels:
            row += 1
            ws[f'E{row}'] = f"  {label}"

            for idx, col in enumerate(self.month_cols):
                if key == 'enterprise_hub':
                    value = self.projections[idx]['revenue'][key]['mrr']
                else:
                    value = self.projections[idx]['revenue'][key]['revenue']
                ws[f'{col}{row}'] = value
                self.apply_style(ws[f'{col}{row}'], self.style_currency)

        # Blank row
        row += 1

        # CHARGES
        row += 1
        ws[f'A{row}'] = "CHARGES"
        ws[f'A{row}'].font = Font(bold=True, size=11)

        # Charges Total
        row += 1
        charges_total_row = row
        ws[f'E{row}'] = "CHARGES TOTALES"
        self.apply_style(ws[f'E{row}'], self.style_total)

        for idx, col in enumerate(self.month_cols):
            ws[f'{col}{row}'] = f'=SUM({col}{row+1}:{col}{row+4})'
            self.apply_style(ws[f'{col}{row}'], {**self.style_currency, **self.style_total})

        # Détail charges
        cost_labels = [
            'Personnel',
            'Infrastructure',
            'Marketing',
            'Admin'
        ]

        for label in cost_labels:
            row += 1
            ws[f'E{row}'] = f"  {label}"

            for idx, col in enumerate(self.month_cols):
                if label == 'Personnel':
                    value = self.projections[idx]['costs']['personnel']['total']
                else:
                    value = self.projections[idx]['costs'][label.lower()]
                ws[f'{col}{row}'] = value
                self.apply_style(ws[f'{col}{row}'], self.style_currency)

        # Blank row
        row += 1

        # EBITDA
        row += 1
        ebitda_row = row
        ws[f'E{row}'] = "EBITDA"
        ws[f'E{row}'].font = Font(bold=True, size=11, color='FF0000')

        for idx, col in enumerate(self.month_cols):
            # Formule: CA Total - Charges Total
            ws[f'{col}{row}'] = f'={col}4-{col}{charges_total_row}'
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
            # Rouge si négatif
            ws[f'{col}{row}'].font = Font(bold=True, color='FF0000')

        # Burn rate
        row += 1
        ws[f'E{row}'] = "Burn Rate"
        for idx, col in enumerate(self.month_cols):
            ws[f'{col}{row}'] = f'=IF({col}{ebitda_row}<0,-{col}{ebitda_row},0)'
            self.apply_style(ws[f'{col}{row}'], self.style_currency)

        # ARR
        row += 1
        arr_row = row
        ws[f'E{row}'] = "ARR (Run Rate)"
        for idx, col in enumerate(self.month_cols):
            # ARR = MRR Hub × 12
            arr_value = self.projections[idx]['metrics']['arr']
            ws[f'{col}{row}'] = arr_value
            self.apply_style(ws[f'{col}{row}'], self.style_arr)

        # Cash position
        row += 1
        ws[f'E{row}'] = "Cash Position"
        for idx, col in enumerate(self.month_cols):
            month_idx = idx
            cash = self.projections[month_idx]['metrics']['cash']
            ws[f'{col}{row}'] = cash
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
            if cash < 50000:
                ws[f'{col}{row}'].font = Font(color='FF0000')

        # Ajuster largeurs colonnes
        ws.column_dimensions['E'].width = 25
        for col in self.month_cols:
            ws.column_dimensions[col].width = 12

        logger.info(f"✓ Sheet P&L créée ({row} rows)")

    def create_synthese_sheet(self):
        """Créer sheet Synthèse (dashboard)"""
        logger.info("📊 Création sheet Synthèse...")

        ws = self.wb.create_sheet("Synthèse", 0)  # Première position

        # Titre
        ws['A1'] = "GenieFactory - Business Plan 14 Mois"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = "Période: Nov 2025 - Dec 2026"
        ws['A2'].font = Font(size=11, italic=True)

        # KPIs clés
        row = 4
        ws[f'A{row}'] = "MÉTRIQUES CLÉS"
        ws[f'A{row}'].font = Font(bold=True, size=14)

        kpis = [
            ('ARR M14 (Dec 2026)', self.projections[13]['metrics']['arr']),
            ('ARR M11 (Sept 2026)', self.projections[10]['metrics']['arr']),
            ('CA Total 14 mois', sum(m['revenue']['total'] for m in self.projections)),
            ('EBITDA Total', sum(m['metrics']['ebitda'] for m in self.projections)),
            ('Burn Rate Max', max(m['metrics']['burn_rate'] for m in self.projections)),
            ('Équipe M14', self.projections[13]['metrics']['team_size']),
            ('Cash M14', self.projections[13]['metrics']['cash'])
        ]

        for idx, (label, value) in enumerate(kpis):
            row += 1
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            if 'ARR' in label:
                self.apply_style(ws[f'B{row}'], self.style_arr)
            else:
                self.apply_style(ws[f'B{row}'], self.style_currency)

        # Validation ARR M14
        row += 2
        ws[f'A{row}'] = "Validation ARR M14:"
        target = self.assumptions['financial_kpis']['target_arr_dec_2026']
        arr_m14 = self.projections[13]['metrics']['arr']

        if (target * 0.9) <= arr_m14 <= (target * 1.1):
            status = "✓ OK"
            color = '00B050'
        else:
            status = "⚠️ Hors target"
            color = 'FF0000'

        ws[f'B{row}'] = status
        ws[f'B{row}'].font = Font(bold=True, color=color)

        # Hypothèses principales
        row += 3
        ws[f'A{row}'] = "HYPOTHÈSES PRINCIPALES"
        ws[f'A{row}'].font = Font(bold=True, size=14)

        hypotheses = [
            ('Hackathon pricing M1-M6', f"{self.assumptions['pricing']['hackathon']['periods'][0]['price_eur']:,}€"),
            ('Factory pricing M1-M6', f"{self.assumptions['pricing']['factory']['periods'][0]['price_eur']:,}€"),
            ('Hub launch', f"M{self.assumptions['pricing']['enterprise_hub']['launch_month']}"),
            ('Conversion Hack→Factory', f"{self.assumptions['sales_assumptions']['factory']['conversion_rate']:.0%}"),
            ('Churn Hub annuel', f"{self.assumptions['pricing']['enterprise_hub']['churn_annual']:.0%}"),
            ('Pre-seed M1', f"{self.assumptions['timeline']['milestones'][0]['amount_eur']:,}€"),
            ('Seed M11', f"{self.assumptions['timeline']['milestones'][1]['amount_eur']:,}€")
        ]

        for label, value in hypotheses:
            row += 1
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        logger.info("✓ Sheet Synthèse créée")

    def add_charts_to_synthese(self, ws):
        """Ajouter graphiques à la sheet Synthèse"""
        logger.info("📈 Ajout graphiques...")

        # Graphique 1: Évolution ARR
        chart_arr = LineChart()
        chart_arr.title = "Évolution ARR (€)"
        chart_arr.style = 10
        chart_arr.y_axis.title = "ARR (€)"
        chart_arr.x_axis.title = "Mois"

        # Données depuis sheet Monitoring
        monitoring_sheet = self.wb['Monitoring']
        data = Reference(monitoring_sheet, min_col=6, min_row=4, max_col=19, max_row=4)  # ARR row
        cats = Reference(monitoring_sheet, min_col=6, min_row=2, max_col=19)  # Months
        chart_arr.add_data(data, titles_from_data=False)
        chart_arr.set_categories(cats)

        ws.add_chart(chart_arr, "D4")

        # Graphique 2: CA mensuel
        chart_ca = LineChart()
        chart_ca.title = "CA Mensuel (€)"
        chart_ca.style = 12
        chart_ca.y_axis.title = "CA (€)"
        chart_ca.x_axis.title = "Mois"

        # Données depuis sheet P&L
        pl_sheet = self.wb['P&L']
        data_ca = Reference(pl_sheet, min_col=6, min_row=4, max_col=19, max_row=4)  # CA Total row
        cats_ca = Reference(pl_sheet, min_col=6, min_row=2, max_col=19)
        chart_ca.add_data(data_ca, titles_from_data=False)
        chart_ca.set_categories(cats_ca)

        ws.add_chart(chart_ca, "D18")

        logger.info("✓ Graphiques ajoutés")

    def create_ventes_sheet(self):
        """Créer sheet Ventes (pipeline détaillé)"""
        logger.info("📊 Création sheet Ventes...")

        ws = self.wb.create_sheet("Ventes")

        # Headers
        ws['E1'] = "PIPELINE VENTES"
        ws['E1'].font = Font(bold=True, size=12)

        ws['E2'] = "Période"
        self.apply_style(ws['E2'], self.style_header)

        for idx, col in enumerate(self.month_cols):
            ws[f'{col}2'] = f"M{idx+1}"
            self.apply_style(ws[f'{col}2'], self.style_header)

        # Volumes
        row = 3
        volumes_data = [
            ('Hackathons (nb)', [p['revenue']['hackathon']['volume'] for p in self.projections]),
            ('Factory (nb)', [p['revenue']['factory']['volume'] for p in self.projections]),
            ('Hub nouveaux clients', [p['revenue']['enterprise_hub'].get('new_customers', 0) for p in self.projections]),
            ('Hub clients total', [p['revenue']['enterprise_hub']['customers']['total'] for p in self.projections])
        ]

        for label, values in volumes_data:
            row += 1
            ws[f'E{row}'] = label
            for idx, col in enumerate(self.month_cols):
                ws[f'{col}{row}'] = values[idx]
                ws[f'{col}{row}'].number_format = '0.0'

        ws.column_dimensions['E'].width = 25
        logger.info("✓ Sheet Ventes créée")

    def create_parametres_sheet(self):
        """Créer sheet Paramètres (pricing reference)"""
        logger.info("📊 Création sheet Paramètres...")

        ws = self.wb.create_sheet("Paramètres")

        ws['A1'] = "GRILLE TARIFAIRE"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        row = 3
        headers = ['Offre', 'M1-M6', 'M7-M14', 'Évolution']
        for idx, header in enumerate(headers, start=1):
            cell = ws.cell(row, idx, header)
            self.apply_style(cell, self.style_header)

        # Pricing data
        pricing_rows = [
            ('Hackathon',
             self.assumptions['pricing']['hackathon']['periods'][0]['price_eur'],
             self.assumptions['pricing']['hackathon']['periods'][1]['price_eur'],
             '+10%'),
            ('Factory',
             self.assumptions['pricing']['factory']['periods'][0]['price_eur'],
             self.assumptions['pricing']['factory']['periods'][1]['price_eur'],
             '+10%'),
            ('Hub Starter (MRR)',
             self.assumptions['pricing']['enterprise_hub']['tiers']['starter']['monthly_eur'],
             self.assumptions['pricing']['enterprise_hub']['tiers']['starter']['monthly_eur'],
             'Fixe'),
            ('Hub Business (MRR)',
             self.assumptions['pricing']['enterprise_hub']['tiers']['business']['monthly_eur'],
             self.assumptions['pricing']['enterprise_hub']['tiers']['business']['monthly_eur'],
             'Fixe'),
            ('Hub Enterprise (MRR)',
             self.assumptions['pricing']['enterprise_hub']['tiers']['enterprise']['monthly_eur'],
             self.assumptions['pricing']['enterprise_hub']['tiers']['enterprise']['monthly_eur'],
             'Fixe')
        ]

        for offre, m1_m6, m7_m14, evolution in pricing_rows:
            row += 1
            ws.cell(row, 1, offre)
            ws.cell(row, 2, m1_m6).number_format = '#,##0 €'
            ws.cell(row, 3, m7_m14).number_format = '#,##0 €'
            ws.cell(row, 4, evolution)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        logger.info("✓ Sheet Paramètres créée")

    def create_financement_sheet(self):
        """Créer sheet Financement"""
        logger.info("📊 Création sheet Financement...")

        ws = self.wb.create_sheet("Financement")

        ws['A1'] = "PLAN DE FINANCEMENT"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "Événement"
        ws[f'B{row}'] = "Mois"
        ws[f'C{row}'] = "Montant"
        for col in ['A', 'B', 'C']:
            self.apply_style(ws[f'{col}{row}'], self.style_header)

        # Milestones
        for milestone in self.assumptions['timeline']['milestones']:
            if 'amount_eur' in milestone:
                row += 1
                ws[f'A{row}'] = milestone['name']
                ws[f'B{row}'] = f"M{milestone['month']}"
                ws[f'C{row}'] = milestone['amount_eur']
                self.apply_style(ws[f'C{row}'], self.style_currency)

        # Breakdown Pre-seed
        row += 2
        ws[f'A{row}'] = "Détail Pre-seed:"
        ws[f'A{row}'].font = Font(bold=True)

        breakdown = self.assumptions['timeline']['milestones'][0]['breakdown']
        for source, amount in breakdown.items():
            row += 1
            ws[f'A{row}'] = f"  {source.replace('_', ' ').title()}"
            ws[f'C{row}'] = amount
            self.apply_style(ws[f'C{row}'], self.style_currency)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15

        logger.info("✓ Sheet Financement créée")

    def create_monitoring_sheet(self):
        """Créer sheet Monitoring (métriques SaaS)"""
        logger.info("📊 Création sheet Monitoring...")

        ws = self.wb.create_sheet("Monitoring")

        ws['A1'] = "MÉTRIQUES SAAS - Enterprise Hub"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['E2'] = "Mois"
        self.apply_style(ws['E2'], self.style_header)

        for idx, col in enumerate(self.month_cols):
            ws[f'{col}2'] = f"M{idx+1}"
            self.apply_style(ws[f'{col}2'], self.style_header)

        # Métriques
        row = 3
        metrics = [
            ('MRR', [p['metrics']['mrr'] for p in self.projections]),
            ('ARR', [p['metrics']['arr'] for p in self.projections]),
            ('Clients Starter', [p['revenue']['enterprise_hub']['customers']['starter'] for p in self.projections]),
            ('Clients Business', [p['revenue']['enterprise_hub']['customers']['business'] for p in self.projections]),
            ('Clients Enterprise', [p['revenue']['enterprise_hub']['customers']['enterprise'] for p in self.projections]),
            ('Clients Total', [p['revenue']['enterprise_hub']['customers']['total'] for p in self.projections])
        ]

        for label, values in metrics:
            row += 1
            ws[f'E{row}'] = label
            for idx, col in enumerate(self.month_cols):
                ws[f'{col}{row}'] = values[idx]
                if 'MRR' in label or 'ARR' in label:
                    self.apply_style(ws[f'{col}{row}'], self.style_arr)
                else:
                    ws[f'{col}{row}'].number_format = '0.0'

        ws.column_dimensions['E'].width = 20
        logger.info("✓ Sheet Monitoring créée")

    def generate(self) -> Workbook:
        """Générer le workbook complet"""
        logger.info("\n🔧 GÉNÉRATION BP EXCEL")
        logger.info("="*60)

        self.create_styles()
        self.create_synthese_sheet()
        self.create_pl_sheet()
        self.create_ventes_sheet()
        self.create_parametres_sheet()
        self.create_financement_sheet()
        self.create_monitoring_sheet()

        # Ajouter les graphiques après création de toutes les sheets
        self.add_charts_to_synthese(self.wb['Synthèse'])

        logger.info("\n✓ Workbook complet généré")
        return self.wb


def main():
    """Fonction principale"""
    logger.info("="*60)
    logger.info("🚀 GÉNÉRATION BP EXCEL - GenieFactory BP 14 Mois")
    logger.info("="*60)

    base_path = Path(__file__).parent.parent

    # Charger projections
    projections_path = base_path / "data" / "structured" / "projections.json"
    if not projections_path.exists():
        logger.error(f"❌ Fichier projections.json non trouvé: {projections_path}")
        logger.error("   Exécuter d'abord: python scripts/3_calculate_projections.py")
        return 1

    logger.info(f"📂 Chargement projections: {projections_path}")
    with open(projections_path, 'r', encoding='utf-8') as f:
        projections = json.load(f)

    # Charger assumptions
    assumptions_path = base_path / "data" / "structured" / "assumptions.yaml"
    logger.info(f"📂 Chargement assumptions: {assumptions_path}")
    with open(assumptions_path, 'r', encoding='utf-8') as f:
        assumptions = yaml.safe_load(f)

    # Générer Excel
    generator = BPExcelGenerator(projections, assumptions)
    wb = generator.generate()

    # Sauvegarder
    output_dir = base_path / "data" / "outputs"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "BP_14M_Nov2025-Dec2026.xlsx"

    wb.save(output_path)

    logger.info("\n" + "="*60)
    logger.info("✅ BP EXCEL GÉNÉRÉ")
    logger.info("="*60)
    logger.info(f"📁 Fichier créé: {output_path}")
    logger.info(f"📊 Sheets: {', '.join(wb.sheetnames)}")
    logger.info(f"💾 Taille: {output_path.stat().st_size / 1024:.1f} KB")

    logger.info("\n✓ Excel prêt à ouvrir dans MS Excel ou LibreOffice")
    logger.info("   → Prochaine étape: python scripts/5_update_bm_word.py")

    return 0


if __name__ == "__main__":
    exit(main())
