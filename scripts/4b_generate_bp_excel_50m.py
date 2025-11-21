#!/usr/bin/env python3
"""
GenieFactory BP 50 Mois - Script 4b: Génération BP Excel Complet
Crée BP_50M_Nov2025-Dec2029.xlsx avec 15 sheets et structure identique au source

Input:
  - data/structured/projections_50m.json
  - data/structured/assumptions.yaml

Output:
  - data/outputs/BP_50M_Nov2025-Dec2029.xlsx (15 sheets, ~122 colonnes P&L)
"""

import json
import yaml
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter

# Configuration logging
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class BPExcel50MGenerator:
    """Générateur BP Excel 50 mois - reproduction exacte structure source"""

    def __init__(self, projections: List[Dict], assumptions: Dict):
        self.projections = projections
        self.assumptions = assumptions
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Supprimer sheet par défaut

        # Structure colonnes comme source:
        # Colonnes C: Total 2025-2026
        # Colonnes D-Q: M1-M14 (Nov 25 - Dec 26)
        # Colonne R: Total 2027
        # Colonnes S-AD: M15-M26 (Jan 27 - Dec 27)
        # Colonne AE: Total 2028
        # Colonnes AF-AQ: M27-M38 (Jan 28 - Dec 28)
        # Colonne AR: Total 2029
        # Colonnes AS-BD: M39-M50 (Jan 29 - Dec 29)

        self.setup_column_structure()

    def setup_column_structure(self):
        """Définir la structure des colonnes pour les 50 mois + totaux annuels"""
        self.columns_map = {}

        # Colonne A: Labels
        # Colonne B: Notes/formules
        # Colonne C: Total 2025-2026

        col_idx = 4  # Commence à D

        # M1-M14 (Nov 2025 - Dec 2026)
        for month in range(1, 15):
            self.columns_map[month] = get_column_letter(col_idx)
            col_idx += 1

        # Colonne R: Total 2027
        self.columns_map['total_2027'] = get_column_letter(col_idx)
        col_idx += 1

        # M15-M26 (2027)
        for month in range(15, 27):
            self.columns_map[month] = get_column_letter(col_idx)
            col_idx += 1

        # Colonne AE: Total 2028
        self.columns_map['total_2028'] = get_column_letter(col_idx)
        col_idx += 1

        # M27-M38 (2028)
        for month in range(27, 39):
            self.columns_map[month] = get_column_letter(col_idx)
            col_idx += 1

        # Colonne AR: Total 2029
        self.columns_map['total_2029'] = get_column_letter(col_idx)
        col_idx += 1

        # M39-M50 (2029)
        for month in range(39, 51):
            self.columns_map[month] = get_column_letter(col_idx)
            col_idx += 1

        logger.info(f"✓ Structure colonnes définie: {len(self.columns_map)} colonnes")

    def create_styles(self):
        """Définir les styles réutilisables"""
        self.style_header_year = {
            'font': Font(bold=True, size=12, color='FFFFFF'),
            'fill': PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

        self.style_header_month = {
            'font': Font(bold=True, size=10, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        }

        self.style_total = {
            'font': Font(bold=True, size=10),
            'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
            'alignment': Alignment(horizontal='right')
        }

        self.style_currency = {
            'number_format': '#,##0 €',
            'alignment': Alignment(horizontal='right')
        }

        self.style_arr = {
            'font': Font(bold=True, color='00B050'),
            'number_format': '#,##0 €',
            'alignment': Alignment(horizontal='right')
        }

        self.style_section_header = {
            'font': Font(bold=True, size=11, color='FFFFFF'),
            'fill': PatternFill(start_color='548235', end_color='548235', fill_type='solid'),
            'alignment': Alignment(horizontal='left')
        }

    def apply_style(self, cell, style_dict):
        """Appliquer un style à une cellule"""
        for key, value in style_dict.items():
            setattr(cell, key, value)

    def create_pl_sheet(self):
        """Créer sheet P&L avec 50 mois (structure exacte source)"""
        logger.info("📊 Création sheet P&L (50 mois)...")

        ws = self.wb.create_sheet("P&L")

        # Titre
        ws['A1'] = "Compte de Résultat Prévisionnel - Nov 2025 à Dec 2029"
        ws['A1'].font = Font(bold=True, size=14)

        # Row 1: Années
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        # Row 2: Mois
        ws['A2'] = "Rubrique"
        ws['B2'] = "Notes"
        ws['C2'] = "Total 25-26"

        # Headers mois M1-M50
        for month in range(1, 51):
            col = self.columns_map[month]
            month_data = self.projections[month - 1]
            date_str = month_data['date']  # 2025-11 format
            month_num = int(date_str.split('-')[1])
            ws[f'{col}2'] = f"M{month_num}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        # Headers totaux annuels
        for total_col, year in [('R2', '2027'), ('AE2', '2028'), ('AR2', '2029')]:
            ws[total_col] = f"Total {year}"
            self.apply_style(ws[total_col], self.style_header_month)

        # === REVENUS ===
        row = 3
        ws[f'A{row}'] = "CHIFFRE D'AFFAIRES"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        # CA Hackathons
        ws[f'A{row}'] = "  Hackathons"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['hackathon']['revenue']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # CA Factory
        ws[f'A{row}'] = "  Factory Projects"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['factory']['revenue']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # CA Hub (MRR)
        ws[f'A{row}'] = "  Enterprise Hub (MRR)"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['enterprise_hub']['mrr']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # CA Services
        ws[f'A{row}'] = "  Services"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['services']['revenue']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Total CA
        row += 1
        ws[f'A{row}'] = "TOTAL CHIFFRE D'AFFAIRES"
        self.apply_style(ws[f'A{row}'], self.style_total)
        ca_total_row = row
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['total']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # === CHARGES ===
        row += 1
        ws[f'A{row}'] = "CHARGES D'EXPLOITATION"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        # Charges personnel
        ws[f'A{row}'] = "  Charges de personnel"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['costs']['personnel']['total']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Infrastructure
        ws[f'A{row}'] = "  Infrastructure technique"
        for month in range(1, 51):
            col = self.columns_map[month]
            infra = self.projections[month - 1]['costs']['infrastructure']
            value = infra if isinstance(infra, (int, float)) else infra['total']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Marketing
        ws[f'A{row}'] = "  Marketing & Commercial"
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            value = marketing if isinstance(marketing, (int, float)) else marketing['total']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Admin
        ws[f'A{row}'] = "  Frais généraux & Admin"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['costs']['admin']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Total charges
        row += 1
        ws[f'A{row}'] = "TOTAL CHARGES"
        self.apply_style(ws[f'A{row}'], self.style_total)
        charges_total_row = row
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['costs']['total']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # === RÉSULTAT ===
        row += 1
        ws[f'A{row}'] = "EBITDA"
        self.apply_style(ws[f'A{row}'], self.style_total)
        ebitda_row = row
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['metrics']['ebitda']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # ARR
        row += 1
        ws[f'A{row}'] = "ARR (Run Rate)"
        arr_row = row
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['metrics']['arr']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_arr)
        row += 1

        # Cash position
        ws[f'A{row}'] = "Cash Position"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['metrics']['cash']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Team size
        ws[f'A{row}'] = "Équipe (ETP)"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['metrics']['team_size']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0'
        row += 1

        # Largeurs colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12

        logger.info(f"✓ Sheet P&L créée: {row} lignes × 50 mois")

    def create_charges_personnel_sheet(self):
        """Créer sheet Charges de personnel et FG (détail par rôle)"""
        logger.info("👥 Création sheet Charges Personnel...")

        ws = self.wb.create_sheet("Charges Personnel")

        # Titre
        ws['A1'] = "Charges de Personnel et Frais Généraux - Détail par Rôle"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers similaires au P&L
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        # Row 2: Mois
        ws['A2'] = "Rôle / Poste"
        ws['B2'] = "Salaire Annuel"
        ws['C2'] = "Total 25-26"

        for month in range(1, 51):
            col = self.columns_map[month]
            ws[f'{col}2'] = f"M{month}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        # Vérifier si personnel_details existe
        if 'personnel_details' not in self.assumptions:
            logger.warning("⚠️ personnel_details non trouvé dans assumptions - sheet simplifiée")
            row = 3
            ws[f'A{row}'] = "Charges de personnel totales"
            for month in range(1, 51):
                col = self.columns_map[month]
                value = self.projections[month - 1]['costs']['personnel']['total']
                ws[f'{col}{row}'] = value
                self.apply_style(ws[f'{col}{row}'], self.style_currency)
            return

        # Détail par rôle
        personnel_details = self.assumptions['personnel_details']
        row = 3

        ws[f'A{row}'] = "SALAIRES BRUTS"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        # Pour chaque rôle
        roles_order = [
            'directeur_general', 'product_owner', 'tech_senior', 'tech_junior',
            'commercial', 'bd_junior', 'stagiaire', 'consultant'
        ]

        for role_name in roles_order:
            if role_name not in personnel_details['roles']:
                continue

            role_data = personnel_details['roles'][role_name]
            ws[f'A{row}'] = f"  {role_data['title']}"
            ws[f'B{row}'] = f"{role_data['salary_brut_annual']:,.0f} €"

            # Pour chaque mois, extraire le coût de ce rôle
            for month in range(1, 51):
                col = self.columns_map[month]
                costs_personnel = self.projections[month - 1]['costs']['personnel']

                if 'roles' in costs_personnel and role_name in costs_personnel['roles']:
                    value = costs_personnel['roles'][role_name]['cost_monthly']
                else:
                    value = 0

                ws[f'{col}{row}'] = value
                self.apply_style(ws[f'{col}{row}'], self.style_currency)

            row += 1

        # Total salaires bruts
        row += 1
        ws[f'A{row}'] = "TOTAL SALAIRES BRUTS"
        self.apply_style(ws[f'A{row}'], self.style_total)
        for month in range(1, 51):
            col = self.columns_map[month]
            costs_personnel = self.projections[month - 1]['costs']['personnel']
            value = costs_personnel.get('salary_brut', 0)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Charges sociales
        row += 1
        ws[f'A{row}'] = "CHARGES SOCIALES (45%)"
        ws[f'A{row}'].font = Font(bold=True)
        for month in range(1, 51):
            col = self.columns_map[month]
            costs_personnel = self.projections[month - 1]['costs']['personnel']
            value = costs_personnel.get('charges_sociales', 0)
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Freelances
        ws[f'A{row}'] = "Freelances / Consultants"
        for month in range(1, 51):
            col = self.columns_map[month]
            costs_personnel = self.projections[month - 1]['costs']['personnel']
            value = costs_personnel.get('freelance', 0)
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Total charges personnel
        row += 1
        ws[f'A{row}'] = "TOTAL CHARGES DE PERSONNEL"
        self.apply_style(ws[f'A{row}'], self.style_total)
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['costs']['personnel']['total']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True, size=11)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # FTE total
        row += 1
        ws[f'A{row}'] = "Effectif Total (ETP)"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['metrics']['team_size']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0.0'
        row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15

        logger.info(f"✓ Sheet Charges Personnel créée: {len(roles_order)} rôles")

    def create_infrastructure_sheet(self):
        """Créer sheet Infrastructure Technique (Cloud + SaaS)"""
        logger.info("☁️ Création sheet Infrastructure Technique...")

        ws = self.wb.create_sheet("Infrastructure")

        # Titre
        ws['A1'] = "Infrastructure Technique - Cloud & SaaS Tools"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        ws['A2'] = "Poste de coût"
        ws['B2'] = "Type"
        ws['C2'] = "Total 25-26"

        for month in range(1, 51):
            col = self.columns_map[month]
            ws[f'{col}2'] = f"M{month}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        row = 3

        # Cloud costs
        ws[f'A{row}'] = "INFRASTRUCTURE CLOUD"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "  Cloud (AWS/Azure)"
        ws[f'B{row}'] = "Variable"
        for month in range(1, 51):
            col = self.columns_map[month]
            infra = self.projections[month - 1]['costs']['infrastructure']
            if isinstance(infra, dict) and 'cloud' in infra:
                value = infra['cloud']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # SaaS tools
        row += 1
        ws[f'A{row}'] = "OUTILS SAAS"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "  SaaS Tools (Notion, Slack, etc.)"
        ws[f'B{row}'] = "Par user"
        for month in range(1, 51):
            col = self.columns_map[month]
            infra = self.projections[month - 1]['costs']['infrastructure']
            if isinstance(infra, dict) and 'saas_tools' in infra:
                value = infra['saas_tools']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # R&D externe
        ws[f'A{row}'] = "  R&D Externe"
        ws[f'B{row}'] = "Fixe"
        for month in range(1, 51):
            col = self.columns_map[month]
            infra = self.projections[month - 1]['costs']['infrastructure']
            if isinstance(infra, dict) and 'rd_external' in infra:
                value = infra['rd_external']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Total
        row += 1
        ws[f'A{row}'] = "TOTAL INFRASTRUCTURE"
        self.apply_style(ws[f'A{row}'], self.style_total)
        for month in range(1, 51):
            col = self.columns_map[month]
            infra = self.projections[month - 1]['costs']['infrastructure']
            value = infra if isinstance(infra, (int, float)) else infra.get('total', 0)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

        logger.info("✓ Sheet Infrastructure créée")

    def create_marketing_sheet(self):
        """Créer sheet Marketing (budget par canal)"""
        logger.info("📢 Création sheet Marketing...")

        ws = self.wb.create_sheet("Marketing")

        # Titre
        ws['A1'] = "Marketing & Acquisition - Budget par Canal"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        ws['A2'] = "Canal Marketing"
        ws['B2'] = "Type"
        ws['C2'] = "Total 25-26"

        for month in range(1, 51):
            col = self.columns_map[month]
            ws[f'{col}2'] = f"M{month}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        row = 3

        ws[f'A{row}'] = "BUDGET MARKETING"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        # Digital ads
        ws[f'A{row}'] = "  Digital Ads (Google, LinkedIn)"
        ws[f'B{row}'] = "Mensuel"
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            if isinstance(marketing, dict) and 'digital_ads' in marketing:
                value = marketing['digital_ads']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Events
        ws[f'A{row}'] = "  Events & Salons"
        ws[f'B{row}'] = "Trimestriel"
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            if isinstance(marketing, dict) and 'events' in marketing:
                value = marketing['events']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Content
        ws[f'A{row}'] = "  Content Marketing"
        ws[f'B{row}'] = "Mensuel"
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            if isinstance(marketing, dict) and 'content' in marketing:
                value = marketing['content']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Partnerships
        ws[f'A{row}'] = "  Partenariats"
        ws[f'B{row}'] = "Mensuel"
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            if isinstance(marketing, dict) and 'partnerships' in marketing:
                value = marketing['partnerships']
            else:
                value = 0
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Total
        row += 1
        ws[f'A{row}'] = "TOTAL MARKETING"
        self.apply_style(ws[f'A{row}'], self.style_total)
        for month in range(1, 51):
            col = self.columns_map[month]
            marketing = self.projections[month - 1]['costs']['marketing']
            value = marketing if isinstance(marketing, (int, float)) else marketing.get('total', 0)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

        logger.info("✓ Sheet Marketing créée")

    def create_ventes_sheet(self):
        """Créer sheet Ventes (pipeline commercial)"""
        logger.info("💼 Création sheet Ventes...")

        ws = self.wb.create_sheet("Ventes")

        # Titre
        ws['A1'] = "Prévisions de Ventes - Pipeline Commercial"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        ws['A2'] = "Segment / Métrique"
        ws['B2'] = "Prix unitaire"
        ws['C2'] = "Total 25-26"

        for month in range(1, 51):
            col = self.columns_map[month]
            ws[f'{col}2'] = f"M{month}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        row = 3

        # Hackathons
        ws[f'A{row}'] = "HACKATHONS"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "  Nombre de hackathons"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['hackathon']['volume']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0'
        row += 1

        ws[f'A{row}'] = "  CA Hackathons"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['hackathon']['revenue']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Factory
        row += 1
        ws[f'A{row}'] = "FACTORY PROJECTS"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "  Nombre de projets Factory"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['factory']['volume']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0.0'
        row += 1

        ws[f'A{row}'] = "  CA Factory"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['factory']['revenue']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Hub
        row += 1
        ws[f'A{row}'] = "ENTERPRISE HUB"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "  Clients Hub actifs"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['enterprise_hub']['customers']['total']
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0.0'
        row += 1

        ws[f'A{row}'] = "  Nouveaux clients Hub"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['enterprise_hub'].get('new_customers', 0)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].number_format = '0'
        row += 1

        ws[f'A{row}'] = "  MRR Hub"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['enterprise_hub']['mrr']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        ws[f'A{row}'] = "  ARR Hub"
        for month in range(1, 51):
            col = self.columns_map[month]
            value = self.projections[month - 1]['revenue']['enterprise_hub']['arr']
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_arr)
        row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15

        logger.info("✓ Sheet Ventes créée")

    def create_synthese_sheet(self):
        """Créer sheet Synthèse (dashboard annuel)"""
        logger.info("📊 Création sheet Synthèse...")

        ws = self.wb.create_sheet("Synthèse", 0)  # Insert at beginning

        # Titre
        ws['A1'] = "Business Plan GenieFactory - Synthèse 2025-2029"
        ws['A1'].font = Font(bold=True, size=16, color='1F4E78')

        row = 3
        ws[f'A{row}'] = "Vue Annuelle Consolidée"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2

        # Headers
        ws['A5'] = "Métrique"
        ws['B5'] = "2025-2026 (14M)"
        ws['C5'] = "2027"
        ws['D5'] = "2028"
        ws['E5'] = "2029"
        ws['F5'] = "TOTAL 50M"

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}5'], self.style_header_month)

        row = 6

        # CA total
        ws[f'A{row}'] = "Chiffre d'Affaires"
        ca_2025_2026 = sum(p['revenue']['total'] for p in self.projections[:14])
        ca_2027 = sum(p['revenue']['total'] for p in self.projections[14:26])
        ca_2028 = sum(p['revenue']['total'] for p in self.projections[26:38])
        ca_2029 = sum(p['revenue']['total'] for p in self.projections[38:50])
        ca_total = ca_2025_2026 + ca_2027 + ca_2028 + ca_2029

        ws[f'B{row}'] = ca_2025_2026
        ws[f'C{row}'] = ca_2027
        ws[f'D{row}'] = ca_2028
        ws[f'E{row}'] = ca_2029
        ws[f'F{row}'] = ca_total
        for col in ['B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
            ws[f'{col}{row}'].font = Font(bold=True)
        row += 1

        # ARR fin de période
        ws[f'A{row}'] = "ARR (fin période)"
        arr_m14 = self.projections[13]['metrics']['arr']
        arr_m26 = self.projections[25]['metrics']['arr']
        arr_m38 = self.projections[37]['metrics']['arr']
        arr_m50 = self.projections[49]['metrics']['arr']

        ws[f'B{row}'] = arr_m14
        ws[f'C{row}'] = arr_m26
        ws[f'D{row}'] = arr_m38
        ws[f'E{row}'] = arr_m50
        ws[f'F{row}'] = arr_m50  # Dernière valeur
        for col in ['B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}{row}'], self.style_arr)
        row += 1

        # Charges totales
        ws[f'A{row}'] = "Charges totales"
        charges_2025_2026 = sum(p['costs']['total'] for p in self.projections[:14])
        charges_2027 = sum(p['costs']['total'] for p in self.projections[14:26])
        charges_2028 = sum(p['costs']['total'] for p in self.projections[26:38])
        charges_2029 = sum(p['costs']['total'] for p in self.projections[38:50])
        charges_total = charges_2025_2026 + charges_2027 + charges_2028 + charges_2029

        ws[f'B{row}'] = charges_2025_2026
        ws[f'C{row}'] = charges_2027
        ws[f'D{row}'] = charges_2028
        ws[f'E{row}'] = charges_2029
        ws[f'F{row}'] = charges_total
        for col in ['B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # EBITDA
        ws[f'A{row}'] = "EBITDA"
        ws[f'B{row}'] = ca_2025_2026 - charges_2025_2026
        ws[f'C{row}'] = ca_2027 - charges_2027
        ws[f'D{row}'] = ca_2028 - charges_2028
        ws[f'E{row}'] = ca_2029 - charges_2029
        ws[f'F{row}'] = ca_total - charges_total
        for col in ['B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
            ws[f'{col}{row}'].font = Font(bold=True, color='00B050' if ws[f'{col}{row}'].value > 0 else 'C00000')
        row += 1

        # Cash fin de période
        ws[f'A{row}'] = "Cash (fin période)"
        ws[f'B{row}'] = self.projections[13]['metrics']['cash']
        ws[f'C{row}'] = self.projections[25]['metrics']['cash']
        ws[f'D{row}'] = self.projections[37]['metrics']['cash']
        ws[f'E{row}'] = self.projections[49]['metrics']['cash']
        ws[f'F{row}'] = self.projections[49]['metrics']['cash']
        for col in ['B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Équipe
        ws[f'A{row}'] = "Équipe (ETP)"
        ws[f'B{row}'] = self.projections[13]['metrics']['team_size']
        ws[f'C{row}'] = self.projections[25]['metrics']['team_size']
        ws[f'D{row}'] = self.projections[37]['metrics']['team_size']
        ws[f'E{row}'] = self.projections[49]['metrics']['team_size']
        for col in ['B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].number_format = '0'
        row += 1

        # Largeurs
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15

        logger.info("✓ Sheet Synthèse créée")

    def create_parametres_sheet(self):
        """Créer sheet Paramètres (pricing et assumptions)"""
        logger.info("⚙️ Création sheet Paramètres...")

        ws = self.wb.create_sheet("Paramètres")

        ws['A1'] = "Paramètres et Hypothèses Clés"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "PRICING"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        # Hackathon pricing
        ws[f'A{row}'] = "Hackathon"
        ws[f'B{row}'] = "20,000 €"
        row += 1

        ws[f'A{row}'] = "Factory Project"
        ws[f'B{row}'] = "82,000 €"
        row += 1

        ws[f'A{row}'] = "Hub Starter (mensuel)"
        ws[f'B{row}'] = "500 €"
        row += 1

        ws[f'A{row}'] = "Hub Business (mensuel)"
        ws[f'B{row}'] = "2,000 €"
        row += 1

        ws[f'A{row}'] = "Hub Enterprise (mensuel)"
        ws[f'B{row}'] = "10,000 €"
        row += 1

        # KPIs
        row += 2
        ws[f'A{row}'] = "KPIS CLES"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "ARR Target M14"
        ws[f'B{row}'] = f"{self.assumptions['financial_kpis']['target_arr_dec_2026']:,} €"
        row += 1

        ws[f'A{row}'] = "Churn Hub mensuel"
        ws[f'B{row}'] = "10%"
        row += 1

        ws[f'A{row}'] = "Conversion Hack→Factory"
        ws[f'B{row}'] = "30%"
        row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        logger.info("✓ Sheet Paramètres créée")

    def create_financement_sheet(self):
        """Créer sheet Financement"""
        logger.info("💰 Création sheet Financement...")

        ws = self.wb.create_sheet("Financement")

        ws['A1'] = "Plan de Financement"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "Tour"
        ws[f'B{row}'] = "Mois"
        ws[f'C{row}'] = "Montant"
        ws[f'D{row}'] = "Valorisation post"
        for col in ['A', 'B', 'C', 'D']:
            self.apply_style(ws[f'{col}{row}'], self.style_header_month)
        row += 1

        # Pre-seed
        ws[f'A{row}'] = "Pre-seed"
        ws[f'B{row}'] = "M1 (Nov 2025)"
        ws[f'C{row}'] = 250000
        ws[f'D{row}'] = 1500000
        self.apply_style(ws[f'C{row}'], self.style_currency)
        self.apply_style(ws[f'D{row}'], self.style_currency)
        row += 1

        # Seed
        ws[f'A{row}'] = "Seed"
        ws[f'B{row}'] = "M11 (Sept 2026)"
        ws[f'C{row}'] = 500000
        ws[f'D{row}'] = 4000000
        self.apply_style(ws[f'C{row}'], self.style_currency)
        self.apply_style(ws[f'D{row}'], self.style_currency)
        row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

        logger.info("✓ Sheet Financement créée")

    def create_strategie_vente_sheet(self):
        """Créer sheet Stratégie de vente"""
        logger.info("🎯 Création sheet Stratégie de vente...")

        ws = self.wb.create_sheet("Stratégie de vente")

        ws['A1'] = "Stratégie de Vente - Pipeline & Conversion"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "PHASES DE VENTE"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        phases = [
            ("Phase 1: Hackathon", "Découverte client, POC technique", "18-20K€"),
            ("Phase 2: Factory", "Projet structurant, 6-9 mois", "75-82K€"),
            ("Phase 3: Hub Subscription", "Récurrence, scaling", "500-10K€/mois"),
        ]

        ws['A4'] = "Phase"
        ws['B4'] = "Description"
        ws['C4'] = "Pricing"
        for col in ['A', 'B', 'C']:
            self.apply_style(ws[f'{col}4'], self.style_header_month)

        row = 5
        for phase, desc, price in phases:
            ws[f'A{row}'] = phase
            ws[f'B{row}'] = desc
            ws[f'C{row}'] = price
            row += 1

        row += 2
        ws[f'A{row}'] = "TAUX DE CONVERSION"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        conversions = [
            ("Hackathon → Factory", "30%"),
            ("Factory → Hub", "50%"),
            ("Hub Starter → Business", "20% (upgrade)"),
        ]

        for label, rate in conversions:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = rate
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20

        logger.info("✓ Sheet Stratégie de vente créée")

    def create_gtmarket_sheet(self):
        """Créer sheet GTMarket (Go-to-Market)"""
        logger.info("🚀 Création sheet GTMarket...")

        ws = self.wb.create_sheet("GTMarket")

        ws['A1'] = "Go-to-Market Strategy - Phases de Déploiement"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "PHASES DE DÉPLOIEMENT"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        phases = [
            ("Nov 2025 - Mars 2026", "Phase 1: Pre-seed", "Lancement Hackathons, premiers clients pilotes"),
            ("Avr 2026 - Sept 2026", "Phase 2: Traction", "Factory projects, préparation Hub"),
            ("Oct 2026 - Dec 2026", "Phase 3: Hub Launch", "Lancement Enterprise Hub, scaling MRR"),
            ("2027", "Phase 4: Scaling", "Croissance clients Hub, équipe 10+ ETP"),
            ("2028", "Phase 5: Consolidation", "ARR 5M€+, préparation Series A"),
            ("2029", "Phase 6: Expansion", "ARR 7M€+, expansion EU"),
        ]

        ws['A5'] = "Période"
        ws['B5'] = "Phase"
        ws['C5'] = "Objectifs"
        for col in ['A', 'B', 'C']:
            self.apply_style(ws[f'{col}5'], self.style_header_month)

        row = 6
        for period, phase, obj in phases:
            ws[f'A{row}'] = period
            ws[f'B{row}'] = phase
            ws[f'C{row}'] = obj
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50

        logger.info("✓ Sheet GTMarket créée")

    def create_sous_traitance_sheet(self):
        """Créer sheet Sous-traitance"""
        logger.info("🔧 Création sheet Sous-traitance...")

        ws = self.wb.create_sheet("Sous-traitance")

        ws['A1'] = "Coûts de Sous-traitance & Freelances"
        ws['A1'].font = Font(bold=True, size=14)

        # Headers
        ws['D1'] = "2025-2026"
        ws.merge_cells('D1:Q1')
        self.apply_style(ws['D1'], self.style_header_year)

        ws['S1'] = "2027"
        ws.merge_cells('S1:AD1')
        self.apply_style(ws['S1'], self.style_header_year)

        ws['AF1'] = "2028"
        ws.merge_cells('AF1:AQ1')
        self.apply_style(ws['AF1'], self.style_header_year)

        ws['AS1'] = "2029"
        ws.merge_cells('AS1:BD1')
        self.apply_style(ws['AS1'], self.style_header_year)

        ws['A2'] = "Type de prestation"
        ws['B2'] = "Description"
        ws['C2'] = "Total 25-26"

        for month in range(1, 51):
            col = self.columns_map[month]
            ws[f'{col}2'] = f"M{month}"
            self.apply_style(ws[f'{col}2'], self.style_header_month)

        row = 3

        ws[f'A{row}'] = "Freelances / Consultants"
        ws[f'B{row}'] = "Dev, Design, Conseil"
        for month in range(1, 51):
            col = self.columns_map[month]
            costs_personnel = self.projections[month - 1]['costs']['personnel']
            value = costs_personnel.get('freelance', 0)
            ws[f'{col}{row}'] = value
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        ws[f'A{row}'] = "Total Sous-traitance"
        self.apply_style(ws[f'A{row}'], self.style_total)
        for month in range(1, 51):
            col = self.columns_map[month]
            costs_personnel = self.projections[month - 1]['costs']['personnel']
            value = costs_personnel.get('freelance', 0)
            ws[f'{col}{row}'] = value
            ws[f'{col}{row}'].font = Font(bold=True)
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30

        logger.info("✓ Sheet Sous-traitance créée")

    def create_direction_sheet(self):
        """Créer sheet DIRECTION (scénarios management)"""
        logger.info("👔 Création sheet DIRECTION...")

        ws = self.wb.create_sheet("DIRECTION")

        ws['A1'] = "Équipe de Direction - Scénarios de Rémunération"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "SCÉNARIOS SALAIRES DIRECTION"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        scenarios = [
            ("Scénario Conservateur", "Salaires minimaux fondateurs", "Total: 200K€/an"),
            ("Scénario Réaliste", "Salaires marché -20%", "Total: 280K€/an"),
            ("Scénario Marché", "Salaires marché complets", "Total: 350K€/an"),
        ]

        ws['A5'] = "Scénario"
        ws['B5'] = "Description"
        ws['C5'] = "Budget annuel"
        for col in ['A', 'B', 'C']:
            self.apply_style(ws[f'{col}5'], self.style_header_month)

        row = 6
        for scenario, desc, budget in scenarios:
            ws[f'A{row}'] = scenario
            ws[f'B{row}'] = desc
            ws[f'C{row}'] = budget
            row += 1

        row += 2
        ws[f'A{row}'] = "ÉQUIPE DIRECTION 2025-2029"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        direction = [
            ("CEO", "50K€/an", "Fondateur, salaire progressif"),
            ("CTO", "45K€/an", "Fondateur, salaire progressif"),
            ("CPO", "40K€/an", "Fondateur, salaire progressif"),
            ("CMO", "45K€/an", "Recrutement 2027"),
        ]

        for role, salary, note in direction:
            ws[f'A{row}'] = role
            ws[f'B{row}'] = salary
            ws[f'C{row}'] = note
            row += 1

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40

        logger.info("✓ Sheet DIRECTION créée")

    def create_fundings_detailed_sheet(self):
        """Créer sheet Fundings (détaillé avec dilution)"""
        logger.info("💰 Création sheet Fundings (détaillé)...")

        ws = self.wb.create_sheet("Fundings")

        ws['A1'] = "Plan de Financement Détaillé - Levées et Dilution"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "TOURS DE FINANCEMENT"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws['A5'] = "Tour"
        ws['B5'] = "Date"
        ws['C5'] = "Montant"
        ws['D5'] = "Valorisation pré"
        ws['E5'] = "Valorisation post"
        ws['F5'] = "Dilution"
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            self.apply_style(ws[f'{col}5'], self.style_header_month)

        row = 6

        # Pre-seed
        ws[f'A{row}'] = "Pre-seed"
        ws[f'B{row}'] = "Nov 2025 (M1)"
        ws[f'C{row}'] = 250000
        ws[f'D{row}'] = 1250000
        ws[f'E{row}'] = 1500000
        ws[f'F{row}'] = "16.7%"
        for col in ['C', 'D', 'E']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Seed
        ws[f'A{row}'] = "Seed"
        ws[f'B{row}'] = "Sept 2026 (M11)"
        ws[f'C{row}'] = 500000
        ws[f'D{row}'] = 3500000
        ws[f'E{row}'] = 4000000
        ws[f'F{row}'] = "12.5%"
        for col in ['C', 'D', 'E']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        # Series A (prévisionnel)
        ws[f'A{row}'] = "Series A (prévisionnel)"
        ws[f'B{row}'] = "Dec 2028 (M38)"
        ws[f'C{row}'] = 2000000
        ws[f'D{row}'] = 18000000
        ws[f'E{row}'] = 20000000
        ws[f'F{row}'] = "10.0%"
        for col in ['C', 'D', 'E']:
            self.apply_style(ws[f'{col}{row}'], self.style_currency)
        row += 1

        row += 2
        ws[f'A{row}'] = "UTILISATION DES FONDS"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        utilisation = [
            ("Pre-seed (250K€)", "40% Produit, 30% Sales/Marketing, 20% Ops, 10% Légal"),
            ("Seed (500K€)", "30% Tech, 40% Commercial, 20% Marketing, 10% Admin"),
            ("Series A (2M€)", "35% R&D, 40% Go-to-market, 15% Intl, 10% Ops"),
        ]

        for tour, usage in utilisation:
            ws[f'A{row}'] = tour
            ws[f'B{row}'] = usage
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18

        logger.info("✓ Sheet Fundings (détaillé) créée")

    def create_navigation_sheet(self):
        """Créer sheet >> (navigation)"""
        logger.info("🧭 Création sheet Navigation...")

        ws = self.wb.create_sheet(">>")

        ws['A1'] = "Navigation - Accès Rapide aux Sheets"
        ws['A1'].font = Font(bold=True, size=16, color='1F4E78')

        row = 3
        ws[f'A{row}'] = "📊 SHEETS PRINCIPALES"
        ws[f'A{row}'].font = Font(bold=True, size=12, color='1F4E78')
        row += 2

        main_sheets = [
            ("1. Synthèse", "Dashboard annuel consolidé"),
            ("2. P&L", "Compte de résultat 50 mois"),
            ("3. Ventes", "Pipeline commercial détaillé"),
            ("4. Charges Personnel", "Détail par rôle et FTE"),
        ]

        for sheet_name, description in main_sheets:
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = description
            ws[f'A{row}'].font = Font(bold=True, color='4472C4')
            row += 1

        row += 1
        ws[f'A{row}'] = "💰 SHEETS FINANCIÈRES"
        ws[f'A{row}'].font = Font(bold=True, size=12, color='548235')
        row += 2

        finance_sheets = [
            ("5. Infrastructure", "Coûts Cloud + SaaS"),
            ("6. Marketing", "Budget par canal"),
            ("7. Sous-traitance", "Freelances & consultants"),
            ("8. Financement", "Plan de financement simple"),
            ("9. Fundings", "Levées détaillées + dilution"),
        ]

        for sheet_name, description in finance_sheets:
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = description
            row += 1

        row += 1
        ws[f'A{row}'] = "📈 SHEETS STRATÉGIE"
        ws[f'A{row}'].font = Font(bold=True, size=12, color='C00000')
        row += 2

        strategy_sheets = [
            ("10. Stratégie de vente", "Pipeline & conversion"),
            ("11. GTMarket", "Go-to-market phases"),
            ("12. DIRECTION", "Scénarios rémunération"),
            ("13. Positionnement", "Analyse concurrentielle"),
            ("14. Paramètres", "Pricing & KPIs"),
        ]

        for sheet_name, description in strategy_sheets:
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = description
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        logger.info("✓ Sheet Navigation créée")

    def create_positionnement_sheet(self):
        """Créer sheet Positionnement (analyse concurrentielle)"""
        logger.info("🎯 Création sheet Positionnement...")

        ws = self.wb.create_sheet("Positionnement")

        ws['A1'] = "Positionnement & Analyse Concurrentielle"
        ws['A1'].font = Font(bold=True, size=14)

        row = 3
        ws[f'A{row}'] = "DIFFÉRENCIATION GÉNIEFACTORY"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        differentiation = [
            ("USP 1", "Approche hybride Hackathon → Factory → Hub"),
            ("USP 2", "No-code/Low-code adapté secteurs régulés"),
            ("USP 3", "Accompagnement end-to-end de l'idée au scaling"),
        ]

        for usp, desc in differentiation:
            ws[f'A{row}'] = usp
            ws[f'B{row}'] = desc
            row += 1

        row += 2
        ws[f'A{row}'] = "MATRICE CONCURRENTIELLE"
        self.apply_style(ws[f'A{row}'], self.style_section_header)
        row += 1

        ws[f'A{row}'] = "Concurrent"
        ws[f'B{row}'] = "Positionnement"
        ws[f'C{row}'] = "Forces"
        ws[f'D{row}'] = "Faiblesses"
        for col in ['A', 'B', 'C', 'D']:
            self.apply_style(ws[f'{col}{row}'], self.style_header_month)
        row += 1

        competitors = [
            ("Agences digitales", "Services custom", "Expertise métier", "Coûts élevés, pas scalable"),
            ("Plateformes no-code", "Self-service", "Prix accessible", "Peu d'accompagnement"),
            ("ESN classiques", "Conseil + Dev", "Crédibilité", "Lenteur, innovation limitée"),
        ]

        for name, pos, strength, weakness in competitors:
            ws[f'A{row}'] = name
            ws[f'B{row}'] = pos
            ws[f'C{row}'] = strength
            ws[f'D{row}'] = weakness
            row += 1

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 35

        logger.info("✓ Sheet Positionnement créée")

    def generate(self):
        """Générer le workbook complet"""
        logger.info("\n🔨 Génération workbook BP 50 mois complet...")

        self.create_styles()

        # Créer les sheets dans l'ordre
        logger.info("\n📑 Création de 15 sheets complètes...")

        # 1. Synthèse (en premier pour être la première tab)
        self.create_synthese_sheet()

        # 2. Stratégie de vente (phases et conversions)
        self.create_strategie_vente_sheet()

        # 3. Financement simple (rounds principaux)
        self.create_financement_sheet()

        # 4. P&L (sheet principale détaillée 50M)
        self.create_pl_sheet()

        # 5. Paramètres (pricing et KPIs)
        self.create_parametres_sheet()

        # 6. GTMarket (phases déploiement)
        self.create_gtmarket_sheet()

        # 7. Ventes (pipeline commercial)
        self.create_ventes_sheet()

        # 8. Sous-traitance (coûts freelance)
        self.create_sous_traitance_sheet()

        # 9. Charges Personnel (détail par rôle)
        self.create_charges_personnel_sheet()

        # 10. DIRECTION (scénarios management)
        self.create_direction_sheet()

        # 11. Infrastructure (Cloud + SaaS)
        self.create_infrastructure_sheet()

        # 12. Fundings détaillé (dilution)
        self.create_fundings_detailed_sheet()

        # 13. >> (Navigation)
        self.create_navigation_sheet()

        # 14. Positionnement (analyse concurrentielle)
        self.create_positionnement_sheet()

        # 15. Marketing (budget par canal)
        self.create_marketing_sheet()

        logger.info("\n✓ Workbook complet généré - 15 sheets")
        logger.info(f"  Sheets: {len(self.wb.sheetnames)}")
        logger.info(f"  Ordre: {', '.join(self.wb.sheetnames)}")
        return self.wb


def main():
    """Fonction principale"""
    logger.info("="*60)
    logger.info("🚀 GÉNÉRATION BP EXCEL 50 MOIS - GenieFactory")
    logger.info("="*60)

    base_path = Path(__file__).parent.parent

    # Charger projections 50M
    projections_path = base_path / "data" / "structured" / "projections_50m.json"
    if not projections_path.exists():
        logger.error(f"❌ Fichier projections_50m.json non trouvé: {projections_path}")
        logger.error("   Exécuter d'abord: python scripts/3_calculate_projections.py")
        return 1

    logger.info(f"📂 Chargement projections: {projections_path}")
    with open(projections_path, 'r', encoding='utf-8') as f:
        projections = json.load(f)

    logger.info(f"✓ Projections chargées: {len(projections)} mois")

    # Charger assumptions
    assumptions_path = base_path / "data" / "structured" / "assumptions.yaml"
    logger.info(f"📂 Chargement assumptions: {assumptions_path}")
    with open(assumptions_path, 'r', encoding='utf-8') as f:
        assumptions = yaml.safe_load(f)

    logger.info(f"✓ Assumptions chargées (version {assumptions.get('version', '1.0')})")

    # Générer Excel
    generator = BPExcel50MGenerator(projections, assumptions)
    wb = generator.generate()

    # Sauvegarder
    output_path = base_path / "data" / "outputs" / "BP_50M_Nov2025-Dec2029.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb.save(output_path)

    logger.info("\n" + "="*60)
    logger.info("✅ BP EXCEL 50 MOIS GÉNÉRÉ")
    logger.info("="*60)
    logger.info(f"📁 Fichier: {output_path}")
    logger.info(f"📊 Taille: {output_path.stat().st_size / 1024:.1f} KB")
    logger.info(f"📑 Sheets: {len(wb.sheetnames)} - {', '.join(wb.sheetnames)}")

    logger.info("\n✓ Excel prêt à ouvrir dans MS Excel ou LibreOffice")
    logger.info(f"   → {len(wb.sheetnames)} sheets créés")
    logger.info("   → Couverture complète: 50 mois (Nov 2025 - Dec 2029)")

    return 0


if __name__ == "__main__":
    exit(main())
