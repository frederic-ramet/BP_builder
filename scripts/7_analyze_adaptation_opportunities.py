#!/usr/bin/env python3
"""
Analyser les autres onglets du RAW pour identifier les opportunités d'adaptation
"""

import openpyxl
from pathlib import Path
import yaml
from rich.console import Console
from rich.table import Table
from rich import box

console = Console()


def analyze_sheet_for_yaml_alignment(wb, sheet_name, assumptions):
    """Analyser un sheet et identifier les opportunités d'adaptation"""

    if sheet_name not in wb.sheetnames:
        return None

    ws = wb[sheet_name]
    opportunities = []

    if sheet_name == "Stratégie de vente":
        # Taux de conversion Hackathon→Factory→Hub
        conversion_rate = assumptions.get('conversion_rates', {}).get('hackathon_to_factory', 0.30)
        opportunities.append({
            'type': 'CONVERSION_RATES',
            'description': f'Taux conversion Hackathon→Factory: {conversion_rate*100:.0f}%',
            'yaml_path': 'conversion_rates.hackathon_to_factory',
            'current_sheet_value': 'À vérifier',
            'action': 'Adapter les taux de conversion selon YAML'
        })

    elif sheet_name == "GTMarket":
        # Timeline de déploiement par phases
        opportunities.append({
            'type': 'TIMELINE',
            'description': 'Timeline déploiement 2025-2029',
            'yaml_path': 'long_term_projections.years',
            'current_sheet_value': 'Phases mensuelles',
            'action': 'Synchroniser avec long_term_projections (ARR growth, team growth)'
        })

    elif sheet_name == "Charges de personnel et FG":
        # 8 rôles détaillés
        roles = assumptions.get('personnel_details', {}).get('roles', {})
        opportunities.append({
            'type': 'PERSONNEL_ROLES',
            'description': f'{len(roles)} rôles définis dans YAML',
            'yaml_path': 'personnel_details.roles',
            'current_sheet_value': 'À vérifier structure',
            'action': 'Adapter structure avec 8 rôles: DG, Commercial, Dev Senior, Dev Junior, etc.'
        })

        charges_rate = assumptions.get('personnel_details', {}).get('charges_sociales_rate', 0.45)
        opportunities.append({
            'type': 'CHARGES_SOCIALES',
            'description': f'Charges sociales: {charges_rate*100:.0f}%',
            'yaml_path': 'personnel_details.charges_sociales_rate',
            'current_sheet_value': 'À vérifier',
            'action': 'S\'assurer que le taux 45% est appliqué'
        })

    elif sheet_name == "DIRECTION":
        # Scénarios de salaire
        opportunities.append({
            'type': 'DIRECTOR_SALARY',
            'description': 'Scénarios salaire direction',
            'yaml_path': 'personnel_details.roles.directeur_general',
            'current_sheet_value': 'Multiples scénarios',
            'action': 'Valider cohérence avec YAML (50K€/an brut base)'
        })

    elif sheet_name == "Infrastructure technique":
        # Pricing cloud et SaaS
        infra = assumptions.get('infrastructure_costs', {})
        cloud_base = infra.get('cloud', {}).get('base_monthly', 1000)
        opportunities.append({
            'type': 'INFRASTRUCTURE_COSTS',
            'description': f'Cloud base: {cloud_base}€/mois + scaling tiers',
            'yaml_path': 'infrastructure_costs',
            'current_sheet_value': 'À vérifier',
            'action': 'Adapter pricing cloud et SaaS selon YAML'
        })

    elif sheet_name == "Marketing":
        # Budgets par canal
        marketing = assumptions.get('marketing_budgets', {})
        channels = ['digital_ads', 'events', 'content', 'partnerships']
        opportunities.append({
            'type': 'MARKETING_BUDGETS',
            'description': f'{len(channels)} canaux marketing avec budgets annuels',
            'yaml_path': 'marketing_budgets',
            'current_sheet_value': 'À vérifier',
            'action': 'Adapter budgets par canal et par année (2025-2029)'
        })

    elif sheet_name == "Fundings":
        # Funding rounds détaillés
        funding = assumptions.get('funding', {})
        opportunities.append({
            'type': 'FUNDING_DETAILS',
            'description': 'Détails rounds: Pre-seed, Seed, Series A avec dilution',
            'yaml_path': 'funding',
            'current_sheet_value': 'Structure complexe',
            'action': 'Adapter montants et valorisations selon YAML'
        })

    elif sheet_name == "Sous traitance":
        # Freelance costs et TJM
        opportunities.append({
            'type': 'FREELANCE_COSTS',
            'description': 'Coûts freelance par type (PM, Audit, Accompagnement, Projet)',
            'yaml_path': 'personnel_details (freelance)',
            'current_sheet_value': 'TJM détaillés',
            'action': 'Simplifier ou valider cohérence avec personnel_details.freelance'
        })

    elif sheet_name == "Positionnement":
        # Analyse concurrentielle
        opportunities.append({
            'type': 'COMPETITIVE_ANALYSIS',
            'description': 'Matrice de positionnement concurrentiel',
            'yaml_path': 'N/A',
            'current_sheet_value': 'Statique',
            'action': 'Pas d\'adaptation nécessaire (contenu stratégique statique)'
        })

    elif sheet_name == ">>":
        # Navigation
        opportunities.append({
            'type': 'NAVIGATION',
            'description': 'Index de navigation entre sheets',
            'yaml_path': 'N/A',
            'current_sheet_value': 'Liste sheets',
            'action': 'Valider que les 15 sheets sont listés correctement'
        })

    return opportunities


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   ANALYSE OPPORTUNITÉS D'ADAPTATION[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    # Charger fichiers
    raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    assumptions_file = base_path / "data" / "structured" / "assumptions.yaml"

    console.print(f"[yellow]📂 Chargement...[/yellow]")
    wb = openpyxl.load_workbook(raw_file, data_only=False)

    with open(assumptions_file, 'r', encoding='utf-8') as f:
        assumptions = yaml.safe_load(f)

    console.print(f"[green]✓ RAW: {len(wb.sheetnames)} sheets[/green]")
    console.print(f"[green]✓ YAML chargé[/green]\n")

    # Sheets à analyser (ceux non encore adaptés)
    sheets_to_analyze = [
        "Stratégie de vente",
        "GTMarket",
        "Ventes",
        "Sous traitance",
        "Charges de personnel et FG",
        "DIRECTION",
        "Infrastructure technique",
        "Marketing",
        "Fundings",
        "Positionnement",
        ">>"
    ]

    all_opportunities = []

    for sheet_name in sheets_to_analyze:
        opps = analyze_sheet_for_yaml_alignment(wb, sheet_name, assumptions)
        if opps:
            for opp in opps:
                opp['sheet'] = sheet_name
                all_opportunities.append(opp)

    # Afficher par priorité
    console.print("[bold]📋 OPPORTUNITÉS D'ADAPTATION PAR PRIORITÉ[/bold]\n")

    # Haute priorité: données structurées du YAML
    high_priority = [o for o in all_opportunities if o['type'] in [
        'PERSONNEL_ROLES', 'CHARGES_SOCIALES', 'INFRASTRUCTURE_COSTS',
        'MARKETING_BUDGETS', 'CONVERSION_RATES'
    ]]

    if high_priority:
        console.print("[bold red]🔴 HAUTE PRIORITÉ (données YAML structurées)[/bold red]\n")

        table = Table(box=box.ROUNDED)
        table.add_column("Sheet", style="cyan", no_wrap=True)
        table.add_column("Type", style="yellow")
        table.add_column("Description", style="white")
        table.add_column("Action", style="green")

        for opp in high_priority:
            table.add_row(
                opp['sheet'],
                opp['type'],
                opp['description'],
                opp['action'][:60] + "..." if len(opp['action']) > 60 else opp['action']
            )

        console.print(table)
        console.print()

    # Priorité moyenne
    medium_priority = [o for o in all_opportunities if o['type'] in [
        'TIMELINE', 'DIRECTOR_SALARY', 'FUNDING_DETAILS', 'FREELANCE_COSTS'
    ]]

    if medium_priority:
        console.print("[bold yellow]🟡 PRIORITÉ MOYENNE (validation cohérence)[/bold yellow]\n")

        table = Table(box=box.SIMPLE)
        table.add_column("Sheet", style="cyan", no_wrap=True)
        table.add_column("Type", style="yellow")
        table.add_column("Action", style="white")

        for opp in medium_priority:
            table.add_row(
                opp['sheet'],
                opp['type'],
                opp['action'][:70] + "..." if len(opp['action']) > 70 else opp['action']
            )

        console.print(table)
        console.print()

    # Basse priorité
    low_priority = [o for o in all_opportunities if o['type'] in [
        'COMPETITIVE_ANALYSIS', 'NAVIGATION'
    ]]

    if low_priority:
        console.print("[bold green]🟢 BASSE PRIORITÉ (pas d'adaptation nécessaire)[/bold green]\n")

        for opp in low_priority:
            console.print(f"  • {opp['sheet']}: {opp['action']}")

        console.print()

    # Résumé
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print(f"[bold]RÉSUMÉ:[/bold]")
    console.print(f"  🔴 Haute priorité: {len(high_priority)} adaptations")
    console.print(f"  🟡 Priorité moyenne: {len(medium_priority)} validations")
    console.print(f"  🟢 Basse priorité: {len(low_priority)} items")
    console.print(f"  📊 Total: {len(all_opportunities)} opportunités identifiées")

    if high_priority:
        console.print(f"\n[yellow]💡 Recommandation: Adapter les {len(high_priority)} sheets haute priorité[/yellow]")
        console.print(f"[yellow]   dans le script 6a_create_template.py[/yellow]\n")


if __name__ == "__main__":
    main()
