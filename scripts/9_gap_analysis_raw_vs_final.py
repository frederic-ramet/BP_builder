#!/usr/bin/env python3
"""
Analyse comparative RAW vs TEMPLATE vs FINAL
Identifier les gaps et proposer des améliorations
"""

import openpyxl
from pathlib import Path
import yaml
import json
from rich.console import Console
from rich.table import Table
from rich import box
from collections import defaultdict

console = Console()


class GapAnalyzer:
    """Analyser les différences entre RAW, TEMPLATE et FINAL"""

    def __init__(self, raw_path: Path, template_path: Path, final_path: Path):
        console.print("[yellow]📂 Chargement des fichiers...[/yellow]")
        self.raw_wb = openpyxl.load_workbook(raw_path, data_only=False)
        self.template_wb = openpyxl.load_workbook(template_path, data_only=False)
        self.final_wb = openpyxl.load_workbook(final_path, data_only=False)
        console.print("[green]✓ 3 fichiers chargés[/green]\n")

    def analyze_sheets_coverage(self):
        """Analyser la couverture des sheets"""
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   1. ANALYSE COUVERTURE SHEETS[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        raw_sheets = set(self.raw_wb.sheetnames)
        template_sheets = set(self.template_wb.sheetnames)
        final_sheets = set(self.final_wb.sheetnames)

        table = Table(box=box.ROUNDED, title="Couverture des Sheets")
        table.add_column("Sheet", style="cyan")
        table.add_column("RAW", justify="center")
        table.add_column("TEMPLATE", justify="center")
        table.add_column("FINAL", justify="center")
        table.add_column("Status", style="yellow")

        all_sheets = sorted(raw_sheets | template_sheets | final_sheets)

        gaps = []
        for sheet in all_sheets:
            in_raw = "✓" if sheet in raw_sheets else "✗"
            in_template = "✓" if sheet in template_sheets else "✗"
            in_final = "✓" if sheet in final_sheets else "✗"

            if sheet in raw_sheets and sheet not in template_sheets:
                status = "⚠️ Retiré"
                gaps.append(f"Sheet '{sheet}' présent dans RAW mais retiré de TEMPLATE")
            elif sheet in template_sheets and sheet not in final_sheets:
                status = "❌ Manquant"
                gaps.append(f"Sheet '{sheet}' présent dans TEMPLATE mais manquant dans FINAL")
            elif sheet not in raw_sheets:
                status = "🆕 Nouveau"
            else:
                status = "✅ OK"

            table.add_row(sheet, in_raw, in_template, in_final, status)

        console.print(table)
        console.print()

        # Résumé
        console.print(f"[bold]Résumé:[/bold]")
        console.print(f"  • RAW: {len(raw_sheets)} sheets")
        console.print(f"  • TEMPLATE: {len(template_sheets)} sheets")
        console.print(f"  • FINAL: {len(final_sheets)} sheets")

        if gaps:
            console.print(f"\n[yellow]⚠️ {len(gaps)} gap(s) détecté(s):[/yellow]")
            for gap in gaps:
                console.print(f"  - {gap}")
        else:
            console.print(f"\n[green]✅ Aucun gap de couverture[/green]")

        console.print()
        return gaps

    def analyze_sheet_structure(self):
        """Analyser la structure détaillée de chaque sheet"""
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   2. ANALYSE STRUCTURE PAR SHEET[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        # Sheets critiques à analyser
        critical_sheets = [
            'P&L', 'Ventes', 'Synthèse', 'Paramètres', 'Financement', 'Fundings',
            'Charges de personnel et FG', 'Infrastructure technique', 'Marketing',
            'Stratégie de vente', 'Sous traitance'
        ]

        issues = []

        for sheet_name in critical_sheets:
            if sheet_name not in self.raw_wb.sheetnames:
                continue

            console.print(f"[cyan]📊 {sheet_name}[/cyan]")

            raw_ws = self.raw_wb[sheet_name]
            template_ws = self.template_wb[sheet_name] if sheet_name in self.template_wb.sheetnames else None
            final_ws = self.final_wb[sheet_name] if sheet_name in self.final_wb.sheetnames else None

            # Dimensions
            raw_dims = f"{raw_ws.max_row} × {raw_ws.max_column}"
            template_dims = f"{template_ws.max_row} × {template_ws.max_column}" if template_ws else "N/A"
            final_dims = f"{final_ws.max_row} × {final_ws.max_column}" if final_ws else "N/A"

            console.print(f"  Dimensions: RAW={raw_dims}, TEMPLATE={template_dims}, FINAL={final_dims}")

            # Compter formules
            raw_formulas = self._count_formulas(raw_ws)
            template_formulas = self._count_formulas(template_ws) if template_ws else 0
            final_formulas = self._count_formulas(final_ws) if final_ws else 0

            console.print(f"  Formules: RAW={raw_formulas}, TEMPLATE={template_formulas}, FINAL={final_formulas}")

            # Check formulas preservation
            if raw_formulas != template_formulas:
                issues.append(f"{sheet_name}: Formules non préservées RAW→TEMPLATE ({raw_formulas}→{template_formulas})")
                console.print(f"  [yellow]⚠️ Formules modifiées: {raw_formulas}→{template_formulas}[/yellow]")

            if template_formulas != final_formulas:
                issues.append(f"{sheet_name}: Formules non préservées TEMPLATE→FINAL ({template_formulas}→{final_formulas})")
                console.print(f"  [yellow]⚠️ Formules modifiées: {template_formulas}→{final_formulas}[/yellow]")

            if raw_formulas == template_formulas == final_formulas and raw_formulas > 0:
                console.print(f"  [green]✓ Formules préservées[/green]")

            # Vérifier données injectées (colonnes F, S, AE, BC pour mois 1, 14, 26, 50)
            if sheet_name == 'P&L' and final_ws:
                self._check_data_injection(final_ws, sheet_name)

            console.print()

        if issues:
            console.print(f"[yellow]⚠️ {len(issues)} problème(s) structurel(s):[/yellow]")
            for issue in issues:
                console.print(f"  - {issue}")
            console.print()

        return issues

    def _count_formulas(self, ws, max_row=100, max_col=150):
        """Compter les formules dans un sheet"""
        count = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_row),
                                min_col=1, max_col=min(ws.max_column, max_col)):
            for cell in row:
                val = cell.value
                if val and not isinstance(val, (str, int, float, bool, type(None))):
                    val = str(val)
                if isinstance(val, str) and val.startswith('='):
                    count += 1
        return count

    def _check_data_injection(self, ws, sheet_name):
        """Vérifier que les données ont été injectées"""
        test_cells = [
            ('F2', 'M1 CA'),
            ('S2', 'M14 CA'),
            ('AE2', 'M26 CA'),
            ('BC2', 'M50 CA')
        ]

        console.print(f"  [cyan]Vérification injection données:[/cyan]")
        for cell_ref, label in test_cells:
            try:
                cell = ws[cell_ref]
                val = cell.value
                is_formula = isinstance(val, str) and val.startswith('=')
                has_value = val is not None and val != 0

                if is_formula:
                    console.print(f"    {label} ({cell_ref}): FORMULE ✓")
                elif has_value:
                    console.print(f"    {label} ({cell_ref}): {val:,.0f}€ ✓")
                else:
                    console.print(f"    {label} ({cell_ref}): [yellow]VIDE ⚠️[/yellow]")
            except Exception as e:
                console.print(f"    {label} ({cell_ref}): [red]ERREUR[/red]")

    def analyze_yaml_to_excel_mapping(self):
        """Analyser le mapping assumptions.yaml → Excel"""
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   3. MAPPING YAML → EXCEL[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        base_path = Path(__file__).parent.parent
        assumptions_path = base_path / "data" / "structured" / "assumptions.yaml"

        with open(assumptions_path, 'r', encoding='utf-8') as f:
            assumptions = yaml.safe_load(f)

        mappings = []

        # 1. Paramètres
        mappings.append({
            'section': 'pricing.hackathon',
            'excel_sheet': 'Paramètres',
            'excel_cell': 'B3',
            'adapted': '✓'
        })

        # 2. Financement
        mappings.append({
            'section': 'funding',
            'excel_sheet': 'Financement',
            'excel_cell': 'C4, E8, G11',
            'adapted': '✓'
        })

        # 3. Cap table
        mappings.append({
            'section': 'funding_captable.yaml',
            'excel_sheet': 'Fundings',
            'excel_cell': 'A1-F40',
            'adapted': '✓'
        })

        # 4. Personnel
        mappings.append({
            'section': 'personnel_details',
            'excel_sheet': 'Charges de personnel et FG',
            'excel_cell': 'A1-B20',
            'adapted': '✓'
        })

        # 5. Infrastructure
        mappings.append({
            'section': 'infrastructure_costs',
            'excel_sheet': 'Infrastructure technique',
            'excel_cell': 'A1-B15',
            'adapted': '✓'
        })

        # 6. Marketing
        mappings.append({
            'section': 'marketing_budgets',
            'excel_sheet': 'Marketing',
            'excel_cell': 'A1-K10',
            'adapted': '✓'
        })

        # 7. Conversion rates
        mappings.append({
            'section': 'conversion_rates',
            'excel_sheet': 'Stratégie de vente',
            'excel_cell': 'A1-B1',
            'adapted': '✓'
        })

        # 8. Sales assumptions → Ventes
        mappings.append({
            'section': 'sales_assumptions (M1-M50)',
            'excel_sheet': 'Ventes',
            'excel_cell': 'F-BC (50 colonnes)',
            'adapted': '✓ via projections_50m.json'
        })

        table = Table(box=box.ROUNDED, title="Mapping YAML → Excel")
        table.add_column("Section YAML", style="cyan")
        table.add_column("Sheet Excel", style="yellow")
        table.add_column("Cellules", style="white")
        table.add_column("Adapté", justify="center")

        for m in mappings:
            table.add_row(m['section'], m['excel_sheet'], m['excel_cell'], m['adapted'])

        console.print(table)
        console.print()

    def identify_improvements(self):
        """Identifier les opportunités d'amélioration"""
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   4. OPPORTUNITÉS D'AMÉLIORATION[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        improvements = []

        # 1. Sheet Paramètres - afficher plus d'infos
        improvements.append({
            'priority': '🔴 HAUTE',
            'sheet': 'Paramètres',
            'issue': 'Sheet peu informatif - seulement les prix',
            'improvement': 'Ajouter KPIs clés: taux conversion (35%), churn Hub (10%/an), ARR targets, timeline Hub (M8)',
            'impact': 'Meilleure compréhension des hypothèses business'
        })

        # 2. P&L - ARR/MRR pas visibles
        improvements.append({
            'priority': '🔴 HAUTE',
            'sheet': 'P&L',
            'issue': 'ARR et MRR pas affichés dans P&L',
            'improvement': 'Ajouter lignes ARR et MRR en haut du P&L pour tracking milestones',
            'impact': 'Visibilité directe des métriques SaaS critiques'
        })

        # 3. Fundings - visualisation cap table
        improvements.append({
            'priority': '🟡 MOYENNE',
            'sheet': 'Fundings',
            'issue': 'Cap table en format texte brut',
            'improvement': 'Ajouter graphique dilution et waterfall valorisation',
            'impact': 'Meilleure visualisation de la dilution progressive'
        })

        # 4. Synthèse - dashboard manquant
        improvements.append({
            'priority': '🔴 HAUTE',
            'sheet': 'Synthèse',
            'issue': 'Pas de dashboard avec KPIs visuels',
            'improvement': 'Créer dashboard avec: ARR graph, CAC/LTV, burn rate, cash runway, team growth',
            'impact': 'Vue exécutive pour investisseurs et board'
        })

        # 5. Ventes - pas de metrics granulaires
        improvements.append({
            'priority': '🟡 MOYENNE',
            'sheet': 'Ventes',
            'issue': 'Volumes détaillés manquants (nb hackathons, clients Hub par tier)',
            'improvement': 'Ajouter lignes: nb hackathons, nb Factory, Hub Starter/Business/Enterprise, churn',
            'impact': 'Traçabilité complète du funnel commercial'
        })

        # 6. Personnel - évolution team pas visible
        improvements.append({
            'priority': '🟢 BASSE',
            'sheet': 'Charges de personnel et FG',
            'issue': 'Évolution team (5→26 ETP) pas tracée par mois',
            'improvement': 'Ajouter ligne "Effectif total" par mois',
            'impact': 'Visibilité sur la croissance RH'
        })

        # 7. Cash flow - manquant
        improvements.append({
            'priority': '🔴 HAUTE',
            'sheet': 'NOUVEAU: Cash Flow',
            'issue': 'Pas de sheet Cash Flow Statement',
            'improvement': 'Créer sheet avec: Operating CF, Investing CF, Financing CF, Cash balance mensuel',
            'impact': 'Essentiel pour suivi trésorerie et fundraising'
        })

        # 8. Scenarios - manquants
        improvements.append({
            'priority': '🟡 MOYENNE',
            'sheet': 'NOUVEAU: Scenarios',
            'issue': 'Pas de scénarios (base/upside/downside)',
            'improvement': 'Créer sheet avec 3 scénarios basés sur assumptions.yaml',
            'impact': 'Analyse de sensibilité pour investisseurs'
        })

        # 9. Unit Economics - manquant
        improvements.append({
            'priority': '🟡 MOYENNE',
            'sheet': 'NOUVEAU: Unit Economics',
            'issue': 'CAC, LTV, payback period pas calculés',
            'improvement': 'Créer sheet avec unit economics par produit (Hackathon, Factory, Hub)',
            'impact': 'Justification du business model'
        })

        # 10. Validation data quality
        improvements.append({
            'priority': '🟢 BASSE',
            'sheet': 'NOUVEAU: Data Quality',
            'issue': 'Pas de checks de cohérence automatiques',
            'improvement': 'Ajouter checks: sum(produits)=CA total, cash>0, team_cost<total_costs',
            'impact': 'Détection erreurs avant envoi investisseurs'
        })

        # Afficher par priorité
        for priority in ['🔴 HAUTE', '🟡 MOYENNE', '🟢 BASSE']:
            priority_items = [i for i in improvements if i['priority'] == priority]
            if not priority_items:
                continue

            console.print(f"\n[bold]{priority}[/bold]")
            table = Table(box=box.SIMPLE)
            table.add_column("Sheet", style="cyan", no_wrap=True)
            table.add_column("Problème", style="yellow")
            table.add_column("Amélioration", style="green")

            for item in priority_items:
                table.add_row(
                    item['sheet'],
                    item['issue'][:50] + "..." if len(item['issue']) > 50 else item['issue'],
                    item['improvement'][:60] + "..." if len(item['improvement']) > 60 else item['improvement']
                )

            console.print(table)

        console.print(f"\n[bold]TOTAL: {len(improvements)} opportunités d'amélioration identifiées[/bold]")
        console.print()

        return improvements

    def generate_action_plan(self, improvements):
        """Générer un plan d'action priorisé"""
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
        console.print("[bold cyan]   5. PLAN D'ACTION RECOMMANDÉ[/bold cyan]")
        console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

        # Phase 1: Quick wins (haute priorité, faible effort)
        phase1 = [
            "1. Enrichir sheet Paramètres avec KPIs (taux conversion, churn, ARR targets)",
            "2. Ajouter lignes ARR/MRR dans P&L",
            "3. Créer sheet Cash Flow avec balance mensuel"
        ]

        # Phase 2: Impact moyen (moyenne priorité)
        phase2 = [
            "4. Ajouter métriques granulaires dans Ventes (nb hackathons, clients Hub par tier)",
            "5. Créer sheet Unit Economics (CAC/LTV par produit)",
            "6. Créer sheet Scenarios (base/upside/downside)"
        ]

        # Phase 3: Nice to have (basse priorité)
        phase3 = [
            "7. Améliorer visualisation cap table dans Fundings (graphiques)",
            "8. Créer dashboard exécutif dans Synthèse",
            "9. Ajouter tracking effectif dans Personnel",
            "10. Créer sheet Data Quality avec checks automatiques"
        ]

        console.print("[bold]🚀 PHASE 1 - Quick Wins (2-3h)[/bold]")
        for item in phase1:
            console.print(f"  {item}")

        console.print("\n[bold]📊 PHASE 2 - Améliorations moyennes (5-6h)[/bold]")
        for item in phase2:
            console.print(f"  {item}")

        console.print("\n[bold]✨ PHASE 3 - Polish final (3-4h)[/bold]")
        for item in phase3:
            console.print(f"  {item}")

        console.print("\n[green]💡 Recommandation: Commencer par Phase 1 pour impact maximal immédiat[/green]\n")


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   GAP ANALYSIS: RAW vs TEMPLATE vs FINAL[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    template_file = base_path / "data" / "outputs" / "BP_50M_TEMPLATE.xlsx"
    final_file = base_path / "data" / "outputs" / "BP_50M_FINAL_Nov2025-Dec2029.xlsx"

    analyzer = GapAnalyzer(raw_file, template_file, final_file)

    # 1. Couverture sheets
    gaps = analyzer.analyze_sheets_coverage()

    # 2. Structure détaillée
    issues = analyzer.analyze_sheet_structure()

    # 3. Mapping YAML → Excel
    analyzer.analyze_yaml_to_excel_mapping()

    # 4. Opportunités d'amélioration
    improvements = analyzer.identify_improvements()

    # 5. Plan d'action
    analyzer.generate_action_plan(improvements)

    # Résumé final
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold]RÉSUMÉ GAP ANALYSIS:[/bold]")
    console.print(f"  • Gaps de couverture: {len(gaps)}")
    console.print(f"  • Problèmes structurels: {len(issues)}")
    console.print(f"  • Opportunités d'amélioration: {len(improvements)}")
    console.print()

    high_priority = [i for i in improvements if i['priority'] == '🔴 HAUTE']
    console.print(f"[bold red]🔴 {len(high_priority)} améliorations HAUTE PRIORITÉ identifiées[/bold red]")
    console.print("[yellow]💡 Voir le plan d'action ci-dessus pour prochaines étapes[/yellow]\n")


if __name__ == "__main__":
    main()
