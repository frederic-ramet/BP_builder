#!/usr/bin/env python3
"""
Analyser structure Fundings et Personnel dans RAW
Pour restructuration état de l'art
"""

import openpyxl
from pathlib import Path
from rich.console import Console
from rich.table import Table
from rich import box

console = Console()

base_path = Path(__file__).parent.parent
raw_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"

console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
console.print("[bold cyan]   ANALYSE STRUCTURE RAW: Fundings & Personnel[/bold cyan]")
console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

wb = openpyxl.load_workbook(raw_file, data_only=False)

# ===== FUNDINGS =====
console.print("[bold yellow]═══ FUNDINGS (RAW) ═══[/bold yellow]\n")

if 'Fundings' in wb.sheetnames:
    ws_fundings = wb['Fundings']
    console.print(f"Dimensions: {ws_fundings.max_row}×{ws_fundings.max_column}\n")

    # Extraire structure (100 premières lignes)
    console.print("[cyan]Structure (50 premières lignes):[/cyan]")
    for row in range(1, min(51, ws_fundings.max_row + 1)):
        col_a = ws_fundings[f'A{row}'].value
        col_b = ws_fundings[f'B{row}'].value
        col_c = ws_fundings[f'C{row}'].value
        col_d = ws_fundings[f'D{row}'].value
        col_e = ws_fundings[f'E{row}'].value

        if col_a or col_b or col_c or col_d or col_e:
            # Tronquer si trop long
            col_a_str = str(col_a)[:40] if col_a else ""
            col_b_str = str(col_b)[:20] if col_b else ""
            col_c_str = str(col_c)[:20] if col_c else ""
            col_d_str = str(col_d)[:20] if col_d else ""
            col_e_str = str(col_e)[:20] if col_e else ""

            console.print(f"  {row:3d}: {col_a_str:<40} | {col_b_str:<20} | {col_c_str:<20} | {col_d_str:<20} | {col_e_str:<20}")

    # Compter formules
    formulas = 0
    for row in ws_fundings.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas += 1
    console.print(f"\n[green]Formules: {formulas}[/green]")
else:
    console.print("[red]Sheet Fundings absent du RAW[/red]")

# ===== PERSONNEL =====
console.print("\n\n[bold yellow]═══ CHARGES DE PERSONNEL ET FG (RAW) ═══[/bold yellow]\n")

if 'Charges de personnel et FG' in wb.sheetnames:
    ws_personnel = wb['Charges de personnel et FG']
    console.print(f"Dimensions: {ws_personnel.max_row}×{ws_personnel.max_column}\n")

    # Extraire structure (100 premières lignes, colonnes A-E)
    console.print("[cyan]Structure (100 premières lignes, colonnes A-E):[/cyan]")
    for row in range(1, min(101, ws_personnel.max_row + 1)):
        col_a = ws_personnel[f'A{row}'].value
        col_b = ws_personnel[f'B{row}'].value
        col_c = ws_personnel[f'C{row}'].value
        col_d = ws_personnel[f'D{row}'].value
        col_e = ws_personnel[f'E{row}'].value

        if col_a or col_b or col_c or col_d or col_e:
            # Tronquer si trop long
            col_a_str = str(col_a)[:50] if col_a else ""
            col_b_str = str(col_b)[:15] if col_b else ""
            col_c_str = str(col_c)[:15] if col_c else ""
            col_d_str = str(col_d)[:15] if col_d else ""
            col_e_str = str(col_e)[:15] if col_e else ""

            console.print(f"  {row:3d}: {col_a_str:<50} | {col_b_str:<15} | {col_c_str:<15} | {col_d_str:<15} | {col_e_str:<15}")

    # Compter formules
    formulas = 0
    for row in ws_personnel.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formulas += 1
    console.print(f"\n[green]Formules: {formulas}[/green]")

    # Identifier profils
    console.print("\n[cyan]Profils RH identifiés:[/cyan]")
    profils = []
    for row in range(1, min(201, ws_personnel.max_row + 1)):
        cell_a = ws_personnel[f'A{row}'].value
        if cell_a and isinstance(cell_a, str):
            # Chercher patterns de profils
            if any(keyword in cell_a.lower() for keyword in ['ceo', 'cto', 'directeur', 'développeur', 'commercial', 'bd', 'marketing', 'ops', 'product']):
                profils.append((row, cell_a))

    for row, profil in profils[:30]:  # Limiter à 30 premiers
        console.print(f"    Ligne {row:3d}: {profil}")

    if len(profils) > 30:
        console.print(f"    ... et {len(profils) - 30} autres profils")
else:
    console.print("[red]Sheet Personnel absent du RAW[/red]")

console.print("\n")
