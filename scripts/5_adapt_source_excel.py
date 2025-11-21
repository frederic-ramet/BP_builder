#!/usr/bin/env python3
"""
Adapter le fichier Excel source en injectant les projections 50M calculées en Python
tout en préservant toutes les formules Excel existantes
"""

import openpyxl
from pathlib import Path
import json
from rich.console import Console
from rich.progress import track
from datetime import datetime
import logging

console = Console()
logging.basicConfig(level=logging.INFO, format='[%(asctime)s] %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


class SourceExcelAdapter:
    """Adapter le fichier source avec les projections calculées"""

    def __init__(self, source_path: Path, projections: list, assumptions: dict):
        self.source_path = source_path
        self.projections = projections
        self.assumptions = assumptions

        # Charger le workbook source
        logger.info(f"📂 Chargement fichier source: {source_path.name}")
        self.wb = openpyxl.load_workbook(source_path)
        logger.info(f"✓ {len(self.wb.sheetnames)} sheets chargés")

        # Mapper la structure des colonnes (50 mois)
        self.setup_month_mapping()

    def setup_month_mapping(self):
        """
        Mapper les 50 mois aux colonnes Excel du fichier source

        Source structure (P&L):
        - Col C: 2025-2026 (header)
        - Col D: 11 (Nov 2025 = M1)
        - Col E: 12 (Dec 2025 = M2)
        - Col F: 1 (Jan 2026 = M3)
        - Col G: 2 (Feb 2026 = M4)
        - ...
        - Col Q: 12 (Dec 2026 = M14)
        - Col R: 2027 (total year)
        - Col S: 1 (Jan 2027 = M15)
        - ...
        """
        self.month_to_col = {}

        # Analyser le P&L pour comprendre la structure
        pl_sheet = self.wb['P&L']

        current_month = 1
        for col_idx in range(4, pl_sheet.max_column + 1):  # Partir de col D
            col_letter = openpyxl.utils.get_column_letter(col_idx)

            # Lire la ligne 1 (année) et ligne 2 (mois ou total)
            year_cell = pl_sheet.cell(1, col_idx).value
            month_cell = pl_sheet.cell(2, col_idx).value

            # Si c'est une colonne de total annuel, skip
            if isinstance(year_cell, (int, float)) and month_cell is None:
                logger.info(f"  Col {col_letter}: Total année {int(year_cell)}")
                continue

            # Si c'est un mois
            if isinstance(month_cell, (int, float)) or (isinstance(month_cell, str) and month_cell.startswith('=')):
                self.month_to_col[current_month] = col_letter
                current_month += 1

            if current_month > 50:
                break

        logger.info(f"✓ Mapping colonnes: {len(self.month_to_col)} mois mappés")
        logger.info(f"  M1 → Col {self.month_to_col.get(1)}")
        logger.info(f"  M14 → Col {self.month_to_col.get(14)}")
        logger.info(f"  M15 → Col {self.month_to_col.get(15)}")
        logger.info(f"  M26 → Col {self.month_to_col.get(26)}")
        logger.info(f"  M50 → Col {self.month_to_col.get(50)}")

    def adapt_pl_sheet(self):
        """Adapter le sheet P&L avec les projections"""
        logger.info("\n📊 Adaptation sheet P&L...")

        ws = self.wb['P&L']

        # Mapper les lignes importantes
        row_mapping = {
            'ca_total': 2,
            'hackathons': 3,
            'factory_projects': 4,
            'hub_mrr': 5,
            'services': 6,
            'depenses_ops': 9,
            'sous_traitance': 10,
            'infrastructure': 11,
            'charges_personnel_ops': 12,
            'resultat_ops': 14,
            'depenses_fonc': 17,
            'marketing': 18,
            'charges_personnel_fonc': 19,
            'frais_generaux': 20,
        }

        # Injecter les données pour chaque mois
        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # CA Total - Vérifier si formule existante
            ca_cell = ws[f'{col}{row_mapping["ca_total"]}']
            if not (isinstance(ca_cell.value, str) and ca_cell.value.startswith('=')):
                ca_cell.value = proj['revenue']['total']

            # Revenues individuels - seulement si pas de formule
            try:
                # Hackathon
                cell = ws[f'{col}{row_mapping["hackathons"]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = proj['revenue']['hackathon']['revenue']

                # Factory
                cell = ws[f'{col}{row_mapping["factory_projects"]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = proj['revenue']['factory']['revenue']

                # Enterprise Hub MRR
                cell = ws[f'{col}{row_mapping["hub_mrr"]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = proj['revenue']['enterprise_hub']['mrr']

                # Services
                cell = ws[f'{col}{row_mapping["services"]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = proj['revenue']['services']['revenue']
            except KeyError as e:
                pass  # Donnée manquante

            # Coûts
            for row_key, data_path in [
                ('sous_traitance', ['costs', 'personnel', 'freelance']),  # Freelance
                ('infrastructure', ['costs', 'infrastructure', 'total']),
                ('charges_personnel_ops', ['costs', 'personnel', 'total']),
                ('marketing', ['costs', 'marketing', 'total']),
                ('charges_personnel_fonc', ['costs', 'personnel', 'total']),
                ('frais_generaux', ['costs', 'admin'])
            ]:
                if row_key not in row_mapping:
                    continue

                cell = ws[f'{col}{row_mapping[row_key]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    value = proj
                    try:
                        for key in data_path:
                            value = value[key]
                        cell.value = value
                    except (KeyError, TypeError):
                        pass  # Donnée non disponible

        logger.info(f"✓ P&L adapté: {len(self.month_to_col)} mois × {len(row_mapping)} lignes")

    def adapt_ventes_sheet(self):
        """Adapter le sheet Ventes"""
        logger.info("\n💼 Adaptation sheet Ventes...")

        ws = self.wb['Ventes']

        # Les lignes importantes pour Ventes
        # On va injecter les nombres de clients, ARR, MRR, etc.
        row_mapping = {
            'nb_hackathons': 3,
            'ca_hackathons': 4,
            'nb_factory': 6,
            'ca_factory': 9,
            'nb_clients_hub': 11,
            'mrr_hub': 13,
            'arr': 15,
        }

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Injecter seulement si pas de formule
            data_map = {
                'nb_hackathons': proj['revenue']['hackathon'].get('volume', 0),
                'ca_hackathons': proj['revenue']['hackathon']['revenue'],
                'nb_factory': proj['revenue']['factory'].get('volume', 0),
                'ca_factory': proj['revenue']['factory']['revenue'],
                'nb_clients_hub': proj['revenue']['enterprise_hub']['customers']['total'],
                'mrr_hub': proj['revenue']['enterprise_hub']['mrr'],
                'arr': proj['revenue']['enterprise_hub']['arr'],
            }

            for row_key, value in data_map.items():
                if row_key not in row_mapping:
                    continue

                cell = ws[f'{col}{row_mapping[row_key]}']
                if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.value = value

        logger.info("✓ Ventes adapté")

    def adapt_charges_personnel_sheet(self):
        """Adapter le sheet Charges de personnel et FG"""
        logger.info("\n👥 Adaptation sheet Charges de personnel et FG...")

        ws = self.wb['Charges de personnel et FG']

        # Injecter les charges de personnel par mois
        # Ligne approximative pour le total
        total_row = 5

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            cell = ws[f'{col}{total_row}']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                total_personnel = proj['costs']['personnel']['total']
                cell.value = total_personnel

        logger.info("✓ Charges personnel adaptées")

    def adapt_infrastructure_sheet(self):
        """Adapter le sheet Infrastructure technique"""
        logger.info("\n☁️ Adaptation sheet Infrastructure technique...")

        ws = self.wb['Infrastructure technique']

        # Lignes pour cloud et SaaS
        cloud_row = 3
        saas_row = 5

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            # Cloud
            cell = ws[f'{col}{cloud_row}']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['infrastructure'].get('cloud', 0)

            # SaaS tools
            cell = ws[f'{col}{saas_row}']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['infrastructure'].get('saas_tools', 0)

        logger.info("✓ Infrastructure adaptée")

    def adapt_marketing_sheet(self):
        """Adapter le sheet Marketing"""
        logger.info("\n📢 Adaptation sheet Marketing...")

        ws = self.wb['Marketing']

        # Ligne approximative pour le total marketing
        total_row = 3

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            cell = ws[f'{col}{total_row}']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.value = proj['costs']['marketing']['total']

        logger.info("✓ Marketing adapté")

    def adapt_sous_traitance_sheet(self):
        """Adapter le sheet Sous traitance"""
        logger.info("\n🔧 Adaptation sheet Sous traitance...")

        ws = self.wb['Sous traitance']

        # Ligne pour le total
        total_row = 3

        for month in range(1, 51):
            if month not in self.month_to_col:
                continue

            col = self.month_to_col[month]
            proj = self.projections[month - 1]

            cell = ws[f'{col}{total_row}']
            if not (isinstance(cell.value, str) and cell.value.startswith('=')):
                # Sous-traitance = freelance dans notre modèle
                cell.value = proj['costs']['personnel'].get('freelance', 0)

        logger.info("✓ Sous traitance adapté")

    def adapt_synthese_sheet(self):
        """
        Le sheet Synthèse référence le P&L avec des formules
        On ne touche à rien ici, les formules vont recalculer automatiquement
        """
        logger.info("\n📊 Sheet Synthèse: Conservation des formules (auto-recalcul)")
        # Rien à faire, les formules ='P&L'!A3 vont récupérer les nouvelles valeurs

    def adapt_all(self):
        """Adapter tous les sheets pertinents"""
        logger.info("\n🔨 ADAPTATION COMPLÈTE DU FICHIER SOURCE")
        logger.info("=" * 60)

        self.adapt_pl_sheet()
        self.adapt_ventes_sheet()
        self.adapt_charges_personnel_sheet()
        self.adapt_infrastructure_sheet()
        self.adapt_marketing_sheet()
        self.adapt_sous_traitance_sheet()
        self.adapt_synthese_sheet()

        logger.info("\n" + "=" * 60)
        logger.info("✅ ADAPTATION TERMINÉE")

    def save(self, output_path: Path):
        """Sauvegarder le workbook adapté"""
        logger.info(f"\n💾 Sauvegarde: {output_path}")
        self.wb.save(output_path)
        size_kb = output_path.stat().st_size / 1024
        logger.info(f"✓ Fichier sauvegardé: {size_kb:.1f} KB")


def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   ADAPTATION FICHIER SOURCE AVEC PROJECTIONS 50M[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent

    # Charger les projections
    projections_path = base_path / "data" / "structured" / "projections_50m.json"
    console.print(f"[yellow]📂 Chargement projections:[/yellow] {projections_path.name}")
    with open(projections_path, 'r', encoding='utf-8') as f:
        projections = json.load(f)
    console.print(f"[green]✓ {len(projections)} mois chargés[/green]")

    # Charger les assumptions
    assumptions_path = base_path / "data" / "structured" / "assumptions.yaml"
    console.print(f"[yellow]📂 Chargement assumptions:[/yellow] {assumptions_path.name}")
    import yaml
    with open(assumptions_path, 'r', encoding='utf-8') as f:
        assumptions = yaml.safe_load(f)
    console.print(f"[green]✓ Assumptions chargées[/green]\n")

    # Fichiers
    source_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    output_file = base_path / "data" / "outputs" / "BP_50M_Adapted_Nov2025-Dec2029.xlsx"

    # Créer l'adaptateur et traiter
    adapter = SourceExcelAdapter(source_file, projections, assumptions)
    adapter.adapt_all()
    adapter.save(output_file)

    console.print(f"\n[bold green]✅ FICHIER ADAPTÉ GÉNÉRÉ[/bold green]")
    console.print(f"[green]📁 {output_file}[/green]")
    console.print(f"\n[cyan]→ Toutes les formules Excel préservées[/cyan]")
    console.print(f"[cyan]→ Données injectées depuis projections_50m.json[/cyan]")
    console.print(f"[cyan]→ Ouvrir dans Excel pour voir les formules recalculer[/cyan]\n")


if __name__ == "__main__":
    main()
