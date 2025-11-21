#!/usr/bin/env python3
"""
Extraction complète de la structure du fichier Excel source
pour permettre sa reproduction exacte avec adaptation 50 mois
"""

import openpyxl
from pathlib import Path
import json
from rich.console import Console
from rich.progress import track
import re

console = Console()

def safe_value(val):
    """Convertir une valeur en type JSON-sérialisable"""
    if val is None:
        return None
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)

def analyze_column_structure(ws):
    """Analyser la structure des colonnes (mois, années, totaux)"""
    headers = {}

    # Ligne 1: années
    row1 = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(1, col_idx)
        val = safe_value(cell.value)
        row1.append({
            'col': openpyxl.utils.get_column_letter(col_idx),
            'value': val,
            'formula': val if isinstance(val, str) and val.startswith('=') else None,
            'merged': cell.coordinate in [str(m) for m in ws.merged_cells]
        })

    headers['row1_years'] = row1

    # Ligne 2: sous-périodes/descriptions
    row2 = []
    for col_idx in range(1, min(ws.max_column + 1, 150)):  # Limiter à 150 cols
        cell = ws.cell(2, col_idx)
        val = safe_value(cell.value)
        row2.append({
            'col': openpyxl.utils.get_column_letter(col_idx),
            'value': val,
            'formula': val if isinstance(val, str) and val.startswith('=') else None
        })

    headers['row2_periods'] = row2

    # Ligne 3: labels ou mois
    row3 = []
    for col_idx in range(1, min(ws.max_column + 1, 150)):
        cell = ws.cell(3, col_idx)
        val = safe_value(cell.value)
        row3.append({
            'col': openpyxl.utils.get_column_letter(col_idx),
            'value': val,
            'formula': val if isinstance(val, str) and val.startswith('=') else None
        })

    headers['row3_labels'] = row3

    return headers

def extract_formulas_patterns(ws, max_samples=50):
    """Extraire les patterns de formules pour comprendre la logique"""
    formulas = []

    for row_idx in range(1, min(ws.max_row + 1, 100)):  # Premières 100 lignes
        for col_idx in range(1, min(ws.max_column + 1, 50)):  # Premières 50 colonnes
            cell = ws.cell(row_idx, col_idx)
            cell_val = cell.value

            # Convertir les objets spéciaux en string
            if cell_val and not isinstance(cell_val, (str, int, float, bool, type(None))):
                cell_val = str(cell_val)

            if isinstance(cell_val, str) and cell_val.startswith('='):
                formulas.append({
                    'cell': cell.coordinate,
                    'row': row_idx,
                    'col': col_idx,
                    'col_letter': openpyxl.utils.get_column_letter(col_idx),
                    'formula': cell_val,
                    'references': extract_cell_references(cell_val)
                })

                if len(formulas) >= max_samples:
                    break

        if len(formulas) >= max_samples:
            break

    return formulas

def extract_cell_references(formula):
    """Extraire les références de cellules d'une formule"""
    # Pattern pour détecter les références: Sheet!A1, A1, $A$1, etc.
    pattern = r"(?:'?([^'!]+)'?!)?(\$?[A-Z]+\$?\d+)"
    matches = re.findall(pattern, formula)

    refs = []
    for sheet, cell in matches:
        refs.append({
            'sheet': sheet if sheet else 'same',
            'cell': cell
        })

    return refs

def extract_row_structure(ws, max_rows=100):
    """Extraire la structure des lignes (labels, groupes)"""
    rows = []

    for row_idx in range(1, min(ws.max_row + 1, max_rows)):
        col_a = safe_value(ws.cell(row_idx, 1).value)
        col_b = safe_value(ws.cell(row_idx, 2).value)
        col_c = safe_value(ws.cell(row_idx, 3).value)

        # Détecter si c'est une formule dans colonne A
        is_formula_a = isinstance(col_a, str) and col_a.startswith('=')

        row_info = {
            'row': row_idx,
            'col_a': col_a,
            'col_b': col_b,
            'col_c': col_c,
            'is_formula_a': is_formula_a,
            'is_header': row_idx <= 3,
            'is_empty': not col_a and not col_b and not col_c
        }

        rows.append(row_info)

    return rows

def analyze_sheet_deep(ws, sheet_name):
    """Analyse approfondie d'un sheet"""
    console.print(f"\n[cyan]📊 Analyse approfondie: {sheet_name}[/cyan]")

    analysis = {
        'name': sheet_name,
        'dimensions': {
            'max_row': ws.max_row,
            'max_col': ws.max_column
        },
        'column_structure': analyze_column_structure(ws),
        'row_structure': extract_row_structure(ws, max_rows=150),
        'formula_patterns': extract_formulas_patterns(ws, max_samples=100),
        'merged_cells': [str(m) for m in ws.merged_cells]
    }

    # Statistiques
    total_formulas = len([f for f in analysis['formula_patterns']])
    inter_sheet_refs = len([f for f in analysis['formula_patterns']
                            if any(r['sheet'] != 'same' for r in f['references'])])

    console.print(f"  ✓ Dimensions: {ws.max_column} cols × {ws.max_row} rows")
    console.print(f"  ✓ Formules extraites: {total_formulas}")
    console.print(f"  ✓ Références inter-sheets: {inter_sheet_refs}")
    console.print(f"  ✓ Cellules fusionnées: {len(analysis['merged_cells'])}")

    return analysis

def main():
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]   EXTRACTION STRUCTURE COMPLÈTE - Excel Source[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    base_path = Path(__file__).parent.parent
    source_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"

    console.print(f"[yellow]📂 Chargement:[/yellow] {source_file.name}\n")
    wb = openpyxl.load_workbook(source_file, data_only=False)

    full_structure = {
        'source_file': source_file.name,
        'total_sheets': len(wb.sheetnames),
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }

    # Analyser chaque sheet
    for sheet_name in track(wb.sheetnames, description="Analyse sheets..."):
        ws = wb[sheet_name]
        full_structure['sheets'][sheet_name] = analyze_sheet_deep(ws, sheet_name)

    # Sauvegarder l'analyse complète
    output_file = base_path / "data" / "outputs" / "source_structure_complete.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(full_structure, f, indent=2, ensure_ascii=False)

    console.print(f"\n[green]✅ Structure complète extraite:[/green] {output_file}")
    console.print(f"[green]   Taille:[/green] {output_file.stat().st_size / 1024:.1f} KB")

    # Créer aussi un rapport lisible
    report_file = base_path / "data" / "outputs" / "source_structure_report.txt"
    with open(report_file, 'w', encoding='utf-8') as f:
        f.write("═" * 80 + "\n")
        f.write("STRUCTURE COMPLÈTE DU FICHIER SOURCE\n")
        f.write(f"Fichier: {source_file.name}\n")
        f.write("═" * 80 + "\n\n")

        for sheet_name in wb.sheetnames:
            sheet_data = full_structure['sheets'][sheet_name]
            f.write(f"\n{'─' * 80}\n")
            f.write(f"SHEET: {sheet_name}\n")
            f.write(f"{'─' * 80}\n")
            f.write(f"Dimensions: {sheet_data['dimensions']['max_col']} colonnes × {sheet_data['dimensions']['max_row']} lignes\n")
            f.write(f"Formules: {len(sheet_data['formula_patterns'])}\n")
            f.write(f"Cellules fusionnées: {len(sheet_data['merged_cells'])}\n\n")

            # Structure colonnes (en-têtes)
            f.write("En-têtes colonnes (Ligne 1 - Années):\n")
            for col_info in sheet_data['column_structure']['row1_years'][:20]:  # Premières 20 cols
                if col_info['value']:
                    f.write(f"  {col_info['col']}: {col_info['value']}\n")

            f.write("\n")

            # Quelques formules exemples
            f.write("Exemples de formules:\n")
            for formula in sheet_data['formula_patterns'][:10]:  # Premières 10
                f.write(f"  {formula['cell']}: {formula['formula'][:80]}\n")

            f.write("\n")

            # Structure lignes (labels)
            f.write("Structure lignes (labels colonne A):\n")
            for row_info in sheet_data['row_structure'][:30]:  # Premières 30 lignes
                if row_info['col_a'] and not row_info['is_empty']:
                    label = str(row_info['col_a'])[:50]
                    f.write(f"  L{row_info['row']}: {label}\n")

            f.write("\n")

    console.print(f"[green]✅ Rapport lisible créé:[/green] {report_file}\n")

if __name__ == "__main__":
    main()
