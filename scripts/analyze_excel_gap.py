#!/usr/bin/env python3
"""
Analyse comparative entre Excel source et Excel généré
"""

import openpyxl
from pathlib import Path
from rich.console import Console
from rich.table import Table
from rich import box
import json

console = Console()

def analyze_excel(file_path: Path):
    """Analyser un fichier Excel et extraire sa structure"""
    wb = openpyxl.load_workbook(file_path, data_only=False)

    analysis = {
        'file': file_path.name,
        'sheets': [],
        'total_sheets': len(wb.sheetnames)
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Compter les lignes/colonnes non vides
        max_row = ws.max_row
        max_col = ws.max_column

        # Extraire les en-têtes (lignes 1-3)
        headers = []
        for row_idx in range(1, min(4, max_row + 1)):
            row_headers = []
            for col_idx in range(1, min(max_col + 1, 70)):  # Limiter à 70 colonnes pour analyse
                cell = ws.cell(row_idx, col_idx)
                row_headers.append({
                    'value': str(cell.value) if cell.value else '',
                    'merged': isinstance(cell, openpyxl.cell.cell.MergedCell)
                })
            headers.append(row_headers)

        # Extraire les labels de lignes (colonne A, B, C)
        row_labels = []
        for row_idx in range(1, min(max_row + 1, 100)):  # Limiter à 100 lignes
            row_data = {}
            for col_idx in range(1, 4):  # Colonnes A, B, C
                cell = ws.cell(row_idx, col_idx)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                row_data[col_letter] = str(cell.value) if cell.value else ''
            row_labels.append(row_data)

        # Détecter les formules
        formula_count = 0
        value_count = 0
        for row in ws.iter_rows(min_row=4, max_row=max_row, min_col=4, max_col=max_col):
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                    else:
                        value_count += 1

        sheet_info = {
            'name': sheet_name,
            'max_row': max_row,
            'max_col': max_col,
            'headers': headers,
            'row_labels': row_labels[:30],  # Limiter à 30 premières lignes
            'formula_count': formula_count,
            'value_count': value_count
        }

        analysis['sheets'].append(sheet_info)

    return analysis

def compare_sheets(source_analysis, generated_analysis):
    """Comparer deux analyses Excel"""
    gaps = []

    # Comparer les noms de sheets
    source_sheets = {s['name'] for s in source_analysis['sheets']}
    gen_sheets = {s['name'] for s in generated_analysis['sheets']}

    missing_sheets = source_sheets - gen_sheets
    extra_sheets = gen_sheets - source_sheets

    if missing_sheets:
        gaps.append({
            'type': 'MISSING_SHEETS',
            'severity': 'HIGH',
            'details': f"Sheets manquants: {', '.join(missing_sheets)}"
        })

    if extra_sheets:
        gaps.append({
            'type': 'EXTRA_SHEETS',
            'severity': 'LOW',
            'details': f"Sheets supplémentaires: {', '.join(extra_sheets)}"
        })

    # Comparer les sheets communs
    common_sheets = source_sheets & gen_sheets

    for sheet_name in common_sheets:
        source_sheet = next(s for s in source_analysis['sheets'] if s['name'] == sheet_name)
        gen_sheet = next(s for s in generated_analysis['sheets'] if s['name'] == sheet_name)

        # Comparer dimensions
        if source_sheet['max_col'] != gen_sheet['max_col']:
            gaps.append({
                'type': 'COLUMN_COUNT_DIFF',
                'severity': 'MEDIUM',
                'sheet': sheet_name,
                'details': f"Colonnes - Source: {source_sheet['max_col']}, Généré: {gen_sheet['max_col']}"
            })

        if abs(source_sheet['max_row'] - gen_sheet['max_row']) > 5:
            gaps.append({
                'type': 'ROW_COUNT_DIFF',
                'severity': 'MEDIUM',
                'sheet': sheet_name,
                'details': f"Lignes - Source: {source_sheet['max_row']}, Généré: {gen_sheet['max_row']}"
            })

        # Comparer formules vs valeurs
        source_has_formulas = source_sheet['formula_count'] > 0
        gen_has_formulas = gen_sheet['formula_count'] > 0

        if source_has_formulas and not gen_has_formulas:
            gaps.append({
                'type': 'FORMULAS_MISSING',
                'severity': 'HIGH',
                'sheet': sheet_name,
                'details': f"Source a {source_sheet['formula_count']} formules, généré a des valeurs statiques"
            })

    return gaps

def print_analysis_summary(source_analysis, generated_analysis, gaps):
    """Afficher un résumé de l'analyse"""

    # Table de comparaison globale
    console.print("\n[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]")
    console.print("[bold cyan]     GAP ANALYSIS: Excel Source vs Généré[/bold cyan]")
    console.print("[bold cyan]═══════════════════════════════════════════════════════[/bold cyan]\n")

    table = Table(title="Comparaison Globale", box=box.ROUNDED)
    table.add_column("Métrique", style="cyan")
    table.add_column("Source", style="yellow")
    table.add_column("Généré", style="green")
    table.add_column("Status", style="bold")

    source_count = source_analysis['total_sheets']
    gen_count = generated_analysis['total_sheets']
    status = "✓" if source_count == gen_count else "✗"

    table.add_row(
        "Nombre de sheets",
        str(source_count),
        str(gen_count),
        status
    )

    console.print(table)

    # Table des sheets
    console.print("\n[bold yellow]Sheets Détail:[/bold yellow]\n")

    sheet_table = Table(box=box.SIMPLE)
    sheet_table.add_column("Sheet Name", style="cyan", no_wrap=True)
    sheet_table.add_column("Source", style="yellow", justify="center")
    sheet_table.add_column("Généré", style="green", justify="center")
    sheet_table.add_column("Cols", justify="center")
    sheet_table.add_column("Rows", justify="center")
    sheet_table.add_column("Formulas", justify="center")

    source_sheets = {s['name']: s for s in source_analysis['sheets']}
    gen_sheets = {s['name']: s for s in generated_analysis['sheets']}

    all_sheet_names = sorted(set(source_sheets.keys()) | set(gen_sheets.keys()))

    for name in all_sheet_names:
        in_source = "✓" if name in source_sheets else "✗"
        in_gen = "✓" if name in gen_sheets else "✗"

        if name in source_sheets and name in gen_sheets:
            s = source_sheets[name]
            g = gen_sheets[name]
            col_match = "✓" if s['max_col'] == g['max_col'] else f"{s['max_col']}≠{g['max_col']}"
            row_diff = abs(s['max_row'] - g['max_row'])
            row_match = "✓" if row_diff <= 5 else f"{s['max_row']}≠{g['max_row']}"
            formula_match = "✓" if (s['formula_count'] > 0) == (g['formula_count'] > 0) else f"{s['formula_count']}≠{g['formula_count']}"
        else:
            col_match = row_match = formula_match = "-"

        sheet_table.add_row(name, in_source, in_gen, col_match, row_match, formula_match)

    console.print(sheet_table)

    # Gaps identifiés
    if gaps:
        console.print(f"\n[bold red]⚠️  {len(gaps)} Gaps Identifiés:[/bold red]\n")

        gap_table = Table(box=box.ROUNDED)
        gap_table.add_column("Sévérité", style="bold")
        gap_table.add_column("Type", style="cyan")
        gap_table.add_column("Sheet", style="yellow")
        gap_table.add_column("Détails", style="white")

        for gap in sorted(gaps, key=lambda x: {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}[x['severity']]):
            severity_color = {
                'HIGH': '[red]🔴 HIGH[/red]',
                'MEDIUM': '[yellow]🟡 MEDIUM[/yellow]',
                'LOW': '[green]🟢 LOW[/green]'
            }[gap['severity']]

            gap_table.add_row(
                severity_color,
                gap['type'],
                gap.get('sheet', '-'),
                gap['details']
            )

        console.print(gap_table)
    else:
        console.print("\n[bold green]✅ Aucun gap majeur identifié![/bold green]\n")

    # Sauvegarder analyse détaillée
    return {
        'source': source_analysis,
        'generated': generated_analysis,
        'gaps': gaps
    }

def main():
    base_path = Path(__file__).parent.parent

    source_file = base_path / "data" / "raw" / "BP FABRIQ_PRODUCT-OCT2025.xlsx"
    generated_file = base_path / "data" / "outputs" / "BP_50M_Nov2025-Dec2029.xlsx"

    console.print(f"\n[cyan]Analyse du fichier source:[/cyan] {source_file.name}")
    source_analysis = analyze_excel(source_file)

    console.print(f"[cyan]Analyse du fichier généré:[/cyan] {generated_file.name}")
    generated_analysis = analyze_excel(generated_file)

    console.print("[cyan]Comparaison en cours...[/cyan]")
    gaps = compare_sheets(source_analysis, generated_analysis)

    full_analysis = print_analysis_summary(source_analysis, generated_analysis, gaps)

    # Sauvegarder analyse JSON
    output_file = base_path / "data" / "outputs" / "excel_gap_analysis.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(full_analysis, f, indent=2, ensure_ascii=False)

    console.print(f"\n[green]✓ Analyse détaillée sauvegardée:[/green] {output_file}")

if __name__ == "__main__":
    main()
