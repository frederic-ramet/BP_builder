#!/usr/bin/env python3
"""
Vérifier sections RH et Volumes Commerciaux dans Paramètres
"""

import openpyxl
from pathlib import Path
from rich.console import Console

console = Console()

base_path = Path(__file__).parent.parent
final_file = base_path / "data" / "outputs" / "BP_50M_FINAL_Nov2025-Dec2029.xlsx"

wb = openpyxl.load_workbook(final_file, data_only=False)
ws = wb['Paramètres']

console.print("\n[bold cyan]═══ VÉRIFICATION SECTIONS PHASE 4 ═══[/bold cyan]\n")

console.print("[bold yellow]COLONNE R-S: Coûts RH[/bold yellow]")
for row in range(1, 15):
    col_r = ws[f'R{row}'].value
    col_s = ws[f'S{row}'].value
    if col_r or col_s:
        console.print(f"  {row}: {col_r or '':<40} {col_s or ''}")

console.print("\n[bold yellow]COLONNE R-S (suite): Volumes Commerciaux[/bold yellow]")
for row in range(6, 12):
    col_r = ws[f'R{row}'].value
    col_s = ws[f'S{row}'].value
    if col_r or col_s:
        console.print(f"  {row}: {col_r or '':<40} {col_s or ''}")

# Vérifier présence charges sociales
charges_found = False
volumes_found = False

for row in range(1, 20):
    for col in ['R', 'S']:
        cell = ws[f'{col}{row}'].value
        if cell and isinstance(cell, str):
            if '45%' in cell or 'charges sociales' in cell.lower():
                charges_found = True
                console.print(f"\n[green]✅ Charges sociales 45% trouvées en {col}{row}: {cell}[/green]")
            if 'hackathon' in cell.lower() and 'volume' in cell.lower():
                volumes_found = True
                console.print(f"[green]✅ Volumes hackathons trouvés en {col}{row}: {cell}[/green]")

if not charges_found:
    console.print("\n[red]❌ Charges sociales 45% NON trouvées[/red]")

if not volumes_found:
    console.print("[red]❌ Volumes hackathons NON trouvés[/red]")

console.print("\n")
